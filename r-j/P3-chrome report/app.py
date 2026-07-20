from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import Flask, Response, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "frequency_report.json"

SOURCE_COLUMNS = [
    "WEEK LABEL",
    "TRANSMISSION",
    "MARKET",
    "GENRE",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "FREQUENCY/LCN NO",
    "BAND",
    "TV CH. No.",
    "AUDIO",
    "VIDEO",
    "LANGUAGE",
    "CRN No.",
    "RANK WITHIN GENRE",
]
FIXED_COLUMNS = [
    "TRANSMISSION",
    "MARKET",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "BAND",
    "TV CH. No.",
    "CRN No.",
]
KEY_COLUMNS = [
    "TRANSMISSION",
    "MARKET",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "CRN No.",
]
DISPLAY_COLUMNS = FIXED_COLUMNS + ["WEEK LABEL", "GENRE", "LANGUAGE", "NAME"]
FREQUENCY_COLUMN = "FREQUENCY/LCN NO"
RANK_COLUMN = "RANK WITHIN GENRE"
FILTER_FIELD_MAP = {
    "market": ("market", "markets"),
    "mso": ("mso", "msos"),
    "mso_type": ("mso_type", "mso_types"),
    "city": ("city", "cities"),
    "head_end": ("head_end", "head_ends"),
    "crn_no": ("crn_no", "crn_numbers"),
    "channel_name": ("channel_name", "channels"),
    "band": ("band", "bands"),
}


def ensure_directories() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def write_report_file(report: dict[str, Any]) -> None:
    temp_file = OUTPUT_FILE.with_suffix(".tmp")
    temp_file.write_text(json.dumps(report, indent=2), encoding="utf-8")
    temp_file.replace(OUTPUT_FILE)


def empty_report(message: str | None = None) -> dict[str, Any]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": [],
        "records": [],
        "summary": {
            "total_channels": 0,
            "increased": 0,
            "decreased": 0,
            "no_change": 0,
        },
        "filters": {
            "markets": [],
            "msos": [],
            "mso_types": [],
            "cities": [],
            "head_ends": [],
            "crn_numbers": [],
            "channels": [],
            "bands": [],
            "weeks": [],
            "change_options": ["Changed", "No Change"],
        },
        "message": message or "Add weekly Excel files to the data folder to generate the report.",
        "data_directory": str(DATA_DIR),
    }


def normalize_report_shape(report: dict[str, Any]) -> dict[str, Any]:
    normalized = empty_report(report.get("message"))
    normalized.update(report)
    normalized["summary"] = {
        **empty_report()["summary"],
        **report.get("summary", {}),
    }
    normalized["filters"] = {
        **empty_report()["filters"],
        **report.get("filters", {}),
    }
    normalized["weeks"] = report.get("weeks", [])
    normalized["records"] = report.get("records", [])
    normalized["message"] = report.get("message", "")
    normalized["data_directory"] = report.get("data_directory", str(DATA_DIR))
    return normalized


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_frequency(value: Any) -> float | int | None:
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group())
    return int(number) if number.is_integer() else number


def normalize_rank(value: Any) -> int | None:
    normalized = normalize_frequency(value)
    if normalized is None:
        return None
    return int(normalized)


def week_sort_key(label: str) -> tuple[int, str]:
    match = re.search(r"(\d+)", label)
    return (int(match.group(1)), label) if match else (10**9, label)


def get_week_files() -> list[Path]:
    ensure_directories()
    files = sorted(DATA_DIR.glob("Week*.xlsx"), key=lambda path: week_sort_key(path.stem))
    return files


def prepare_week_rows(path: Path, fallback_label: str) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    week_label = fallback_label

    for row_index, values in enumerate(sheet.iter_rows(min_row=1, max_col=16, values_only=True), start=1):
        if row_index == 1:
            continue

        row = {
            SOURCE_COLUMNS[index]: values[index] if index < len(values) else None
            for index in range(len(SOURCE_COLUMNS))
        }

        if row_index == 2 and normalize_text(row["WEEK LABEL"]):
            week_label = normalize_text(row["WEEK LABEL"])

        normalized = {column: normalize_text(row.get(column)) for column in DISPLAY_COLUMNS if column != "NAME"}
        normalized["NAME"] = normalize_text(row.get("CHANNEL NAME"))
        normalized[FREQUENCY_COLUMN] = normalize_frequency(row.get(FREQUENCY_COLUMN))
        normalized[RANK_COLUMN] = normalize_rank(row.get(RANK_COLUMN))
        normalized["ROW KEY"] = "||".join(normalized[column] for column in KEY_COLUMNS)

        if not normalized["ROW KEY"].replace("|", ""):
            continue

        rows.append(normalized)

    workbook.close()

    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        deduped.setdefault(row["ROW KEY"], row)

    return week_label, list(deduped.values())


def calculate_pairwise_change(previous: float | int | None, current: float | int | None) -> str:
    if previous is None or current is None:
        return "missing"
    if current > previous:
        return "increase"
    if current < previous:
        return "decrease"
    return "no_change"


def calculate_rank_change(previous: int | None, current: int | None) -> str:
    if previous is None or current is None:
        return "missing"
    if current < previous:
        return "improve"
    if current > previous:
        return "decline"
    return "no_change"


def calculate_band_change(previous: str | None, current: str | None) -> str:
    previous_text = normalize_text(previous)
    current_text = normalize_text(current)
    if not previous_text or not current_text:
        return "missing"
    if current_text == previous_text:
        return "no_change"
    return "change"


def has_any_change(series: dict[str, Any], weeks: list[str]) -> bool:
    values = [series.get(week) for week in weeks if series.get(week) not in (None, "")]
    if len(values) <= 1:
        return False
    return len(set(values)) > 1


def summarize_records(records: list[dict[str, Any]], weeks: list[str], view: str = "frequency") -> dict[str, int]:
    if view == "rank":
        change_key = "rank_changes"
        positive_status = "improve"
        negative_status = "decline"
        positive_label = "improved"
        negative_label = "declined"
    elif view == "band":
        change_key = "band_changes"
        positive_status = "change"
        negative_status = "no_change"
        positive_label = "changed"
        negative_label = "stable"
    else:
        change_key = "changes"
        positive_status = "increase"
        negative_status = "decrease"
        positive_label = "increased"
        negative_label = "decreased"

    increased = 0
    decreased = 0
    no_change = 0

    for record in records:
        for week in weeks[1:]:
            status = record[change_key][week]
            if status == positive_status:
                increased += 1
            elif status == negative_status:
                decreased += 1
            elif status == "no_change":
                no_change += 1

    return {
        "total_channels": len(records),
        positive_label: increased,
        negative_label: decreased,
        "no_change": no_change,
    }


def summarize_focus_channels(records: list[dict[str, Any]], weeks: list[str], view: str) -> list[dict[str, Any]]:
    focus_channels = [
        ("INDIA TV", {"INDIA TV"}),
        ("AAJ TAK", {"AAJ TAK"}),
        ("NEWS 18", {"NEWS 18 INDIA"}),
        ("REPUBLIC BHARAT", {"REPUBLIC BHARAT"}),
    ]

    if view == "rank":
        change_key = "rank_changes"
        positive_status = "improve"
        positive_label = "improved"
        negative_status = "decline"
        negative_label = "declined"
    elif view == "band":
        change_key = "band_changes"
        positive_status = "change"
        positive_label = "changed"
        negative_status = "no_change"
        negative_label = "stable"
    else:
        change_key = "changes"
        positive_status = "increase"
        positive_label = "increased"
        negative_status = "decrease"
        negative_label = "decreased"

    summaries: list[dict[str, Any]] = []
    latest_week = weeks[-1] if weeks else ""

    for label, aliases in focus_channels:
        channel_records = [record for record in records if record["channel_name"] in aliases]
        positive_count = 0
        negative_count = 0
        no_change_count = 0

        for record in channel_records:
            for week in weeks[1:]:
                status = record[change_key].get(week)
                if status == positive_status:
                    positive_count += 1
                elif status == negative_status:
                    negative_count += 1
                elif status == "no_change":
                    no_change_count += 1

        latest_changed = sum(
            1
            for record in channel_records
            if latest_week and record[change_key].get(latest_week) == positive_status
        )
        latest_negative = sum(
            1
            for record in channel_records
            if latest_week and record[change_key].get(latest_week) == negative_status
        )

        summaries.append(
            {
                "label": label,
                "records": len(channel_records),
                "positive": positive_count,
                "positive_label": positive_label,
                "negative": negative_count,
                "negative_label": negative_label,
                "no_change": no_change_count,
                "latest_week": latest_week,
                "latest_positive": latest_changed,
                "latest_negative": latest_negative,
            }
        )

    return summaries


def build_report() -> dict[str, Any]:
    week_files = get_week_files()
    if not week_files:
        report = empty_report()
        write_report_file(report)
        return report

    weekly_data = [prepare_week_rows(path, path.stem) for path in week_files]
    weeks = [label for label, _ in weekly_data]

    merged: dict[str, dict[str, Any]] = {}
    for week_label, rows in weekly_data:
        for row in rows:
            record = merged.setdefault(
                row["ROW KEY"],
                {
                    "row_key": row["ROW KEY"],
                    "transmission": row["TRANSMISSION"],
                    "market": row["MARKET"],
                    "mso_type": row["MSO TYPE"],
                    "city": row["CITY"],
                    "head_end": row["HEAD-END"],
                    "channel_name": row["CHANNEL NAME"],
                    "band": row["BAND"],
                    "tv_ch_no": row["TV CH. No."],
                    "crn_no": row["CRN No."],
                    "genre": row["GENRE"],
                    "language": row["LANGUAGE"],
                    "name": row["NAME"],
                    "week_label": row["WEEK LABEL"],
                    "frequencies": {},
                    "ranks": {},
                    "bands": {},
                    "changes": {},
                    "rank_changes": {},
                    "band_changes": {},
                },
            )
            record["frequencies"][week_label] = row[FREQUENCY_COLUMN]
            record["ranks"][week_label] = row[RANK_COLUMN]
            record["bands"][week_label] = row["BAND"]

    records = list(merged.values())

    for record in records:
        for week in weeks:
            record["frequencies"].setdefault(week, None)
            record["ranks"].setdefault(week, None)
            record["bands"].setdefault(week, "")

        record["changes"][weeks[0]] = "baseline"
        record["rank_changes"][weeks[0]] = "baseline"
        record["band_changes"][weeks[0]] = "baseline"
        for index in range(1, len(weeks)):
            previous_week = weeks[index - 1]
            current_week = weeks[index]
            record["changes"][current_week] = calculate_pairwise_change(
                record["frequencies"][previous_week],
                record["frequencies"][current_week],
            )
            record["rank_changes"][current_week] = calculate_rank_change(
                record["ranks"][previous_week],
                record["ranks"][current_week],
            )
            record["band_changes"][current_week] = calculate_band_change(
                record["bands"][previous_week],
                record["bands"][current_week],
            )

        record["change_status"] = "YES" if has_any_change(record["frequencies"], weeks) else "NO"
        record["rank_change_status"] = "YES" if has_any_change(record["ranks"], weeks) else "NO"
        record["band_change_status"] = "YES" if has_any_change(record["bands"], weeks) else "NO"

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": weeks,
        "records": records,
        "summary": summarize_records(records, weeks, "frequency"),
        "filters": {
            "markets": sorted({record["market"] for record in records if record["market"]}),
            "msos": sorted({record.get("mso", "") for record in records if record.get("mso", "")}),
            "mso_types": sorted({record["mso_type"] for record in records if record["mso_type"]}),
            "cities": sorted({record["city"] for record in records if record["city"]}),
            "head_ends": sorted({record["head_end"] for record in records if record["head_end"]}),
            "crn_numbers": sorted({record["crn_no"] for record in records if record["crn_no"]}),
            "channels": sorted({record["channel_name"] for record in records if record["channel_name"]}),
            "bands": sorted({record["band"] for record in records if record["band"]}),
            "weeks": weeks,
            "change_options": ["Changed", "No Change"],
        },
    }

    write_report_file(report)
    return report


def report_is_stale() -> bool:
    ensure_directories()
    if not OUTPUT_FILE.exists():
        return True
    output_time = OUTPUT_FILE.stat().st_mtime
    week_files = get_week_files()
    if not week_files:
        return False
    return any(path.stat().st_mtime > output_time for path in week_files)


def load_report() -> dict[str, Any]:
    ensure_directories()
    if not get_week_files():
        return build_report()
    if report_is_stale():
        return build_report()
    try:
        report = json.loads(OUTPUT_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return build_report()
    return normalize_report_shape(report)


def get_view_maps(view: str) -> tuple[str, str]:
    if view == "rank":
        return "rank_changes", "rank_change_status"
    if view == "band":
        return "band_changes", "band_change_status"
    return "changes", "change_status"


def apply_filters(records: list[dict[str, Any]], filters: dict[str, str], search: str, view: str = "frequency") -> list[dict[str, Any]]:
    search_value = search.strip().lower()
    filtered = []
    change_key, status_key = get_view_maps(view)

    for record in records:
        if filters.get("market") and record["market"] != filters["market"]:
            continue
        if filters.get("mso") and record.get("mso", "") != filters["mso"]:
            continue
        if filters.get("mso_type") and record["mso_type"] != filters["mso_type"]:
            continue
        if filters.get("city") and record["city"] != filters["city"]:
            continue
        if filters.get("head_end") and record["head_end"] != filters["head_end"]:
            continue
        if filters.get("crn_no") and record["crn_no"] != filters["crn_no"]:
            continue
        if filters.get("channel_name") and record["channel_name"] != filters["channel_name"]:
            continue
        if filters.get("band") and record["band"] != filters["band"]:
            continue
        if filters.get("change"):
            if filters["change"] == "Changed":
                if filters.get("week"):
                    changed_states = {"change"} if view == "band" else {"increase", "decrease", "improve", "decline"}
                    if record[change_key].get(filters["week"]) not in changed_states:
                        continue
                elif record[status_key] != "YES":
                    continue
            elif filters["change"] == "No Change":
                if filters.get("week"):
                    if record[change_key].get(filters["week"]) != "no_change":
                        continue
                elif record[status_key] != "NO":
                    continue
        if search_value:
            haystack = " ".join(
                [
                    record["channel_name"],
                    record["name"],
                    record["market"],
                    record["city"],
                    record["head_end"],
                ]
            ).lower()
            if search_value not in haystack:
                continue
        filtered.append(record)

    return filtered


def serialize_records(records: list[dict[str, Any]], weeks: list[str], view: str) -> list[dict[str, Any]]:
    if view == "rank":
        series_key = "ranks"
        change_key = "rank_changes"
        status_key = "rank_change_status"
    elif view == "band":
        series_key = "bands"
        change_key = "band_changes"
        status_key = "band_change_status"
    else:
        series_key = "frequencies"
        change_key = "changes"
        status_key = "change_status"

    serialized: list[dict[str, Any]] = []
    for record in records:
        item = dict(record)
        item["frequencies"] = {week: record[series_key].get(week) for week in weeks}
        item["changes"] = {week: record[change_key].get(week) for week in weeks}
        item["change_status"] = record[status_key]
        serialized.append(item)
    return serialized


def build_contextual_filters(
    records: list[dict[str, Any]],
    filters: dict[str, str],
    search: str,
    view: str,
    weeks: list[str],
) -> dict[str, list[str]]:
    contextual_filters: dict[str, list[str]] = {
        "weeks": weeks,
        "change_options": ["Changed", "No Change"],
    }

    for filter_name, (record_key, response_key) in FILTER_FIELD_MAP.items():
        sibling_filters = {key: value for key, value in filters.items() if key != filter_name}
        matching_records = apply_filters(records, sibling_filters, search, view)
        contextual_filters[response_key] = sorted(
            {
                str(record.get(record_key, "")).strip()
                for record in matching_records
                if str(record.get(record_key, "")).strip()
            }
        )

    return contextual_filters


def sort_records(
    records: list[dict[str, Any]],
    sort_key: str,
    sort_direction: str,
    weeks: list[str],
    view: str = "frequency",
) -> list[dict[str, Any]]:
    reverse = sort_direction == "desc"

    def value_for(record: dict[str, Any]) -> Any:
        if sort_key in weeks:
            if view == "rank":
                value = record["ranks"].get(sort_key)
            elif view == "band":
                value = record["bands"].get(sort_key)
            else:
                value = record["frequencies"].get(sort_key)
        else:
            value = record.get(sort_key)
        if value is None or value == "":
            return (1, "")
        return (0, value)

    return sorted(records, key=value_for, reverse=reverse)


def paginate_records(records: list[dict[str, Any]], page: int, page_size: int) -> tuple[list[dict[str, Any]], int]:
    total_count = len(records)
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    current_page = min(max(page, 1), total_pages)
    start = (current_page - 1) * page_size
    end = start + page_size
    return records[start:end], total_pages


def create_export_workbook(records: list[dict[str, Any]], weeks: list[str], view: str) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{view.title()} Report"

    headers = [
        "TRANSMISSION",
        "MARKET",
        "MSO TYPE",
        "CITY",
        "HEAD-END",
        "CHANNEL NAME",
        "BAND",
        "TV CH. No.",
        "CRN No.",
        "NAME",
    ]
    for week in weeks:
        headers.extend([week, f"{week} STATUS"])
    sheet.append(headers)

    for record in records:
        row = [
            record["transmission"],
            record["market"],
            record["mso_type"],
            record["city"],
            record["head_end"],
            record["channel_name"],
            record["band"],
            record["tv_ch_no"],
            record["crn_no"],
            record["name"],
        ]
        if view == "rank":
            series = record["ranks"]
            change_map = record["rank_changes"]
        elif view == "band":
            series = record["bands"]
            change_map = record["band_changes"]
        else:
            series = record["frequencies"]
            change_map = record["changes"]
        for week in weeks:
            row.extend([series[week], change_map[week]])
        sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    workbook.close()
    return buffer


def create_shareable_dashboard_html(report: dict[str, Any]) -> str:
    style_text = (BASE_DIR / "static" / "style.css").read_text(encoding="utf-8")
    report_json = json.dumps(report).replace("</", "<\\/")
    html = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Chrome Report</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  <style>
__STYLE__
.offline-note { margin-top: 8px; color: var(--muted); font-size: 0.72rem; }
  </style>
</head>
<body>
  <div class="app-shell">
    <header class="hero">
      <div class="hero-copy">
        <h1>Chrome Report</h1>
      </div>
      <div class="hero-meta">
        <article class="meta-card">
          <span>Generated</span>
          <strong id="generatedAt">--</strong>
        </article>
        <article class="meta-card">
          <span>Total Records</span>
          <strong id="totalRecords">0</strong>
        </article>
        <div class="hero-actions">
          <button id="saveCopyButton" class="ghost-button" type="button">Save Copy</button>
        </div>
      </div>
    </header>

    <section class="panel filter-panel">
      <div class="panel-heading">
        <div>
          <h2>Filter Panel</h2>
        </div>
      </div>
      <div class="filter-grid">
        <label><span>Market</span><select id="marketFilter"></select></label>
        <label><span>City</span><select id="cityFilter"></select></label>
        <label><span>MSO Type</span><select id="msoTypeFilter"></select></label>
        <label><span>MSO</span><select id="msoFilter"></select></label>
        <label><span>Headend</span><select id="headendFilter"></select></label>
        <label><span>CRN No</span><select id="crnFilter"></select></label>
        <label><span>Channel</span><select id="channelFilter"></select></label>
        <label><span>Band</span><select id="bandFilter"></select></label>
        <label><span>Week</span><select id="weekFilter"></select></label>
        <label><span>Change</span><select id="changeFilter"></select></label>
        <div class="action-row">
          <button id="resetButton" class="ghost-button" type="button">Reset Filters</button>
        </div>
      </div>
      <div class="offline-note">This is a downloaded standalone dashboard. Filters and view switching work without the backend.</div>
    </section>

    <section class="panel table-panel">
      <div class="panel-heading table-heading">
        <div>
          <h2 id="tableTitle">Weekly Frequency Analysis</h2>
        </div>
        <div class="table-side">
          <div class="view-switcher">
            <button id="frequencyViewButton" class="switch-button active" type="button">Frequency</button>
            <button id="rankViewButton" class="switch-button" type="button">Rank</button>
            <button id="bandViewButton" class="switch-button" type="button">Band</button>
          </div>
          <div class="table-meta">
            <span id="resultCount">0 records</span>
            <span id="pageInfo">Page 1</span>
          </div>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead id="tableHead"></thead>
          <tbody id="tableBody"></tbody>
        </table>
      </div>
      <div class="pagination-bar">
        <button id="prevPage" class="ghost-button" type="button">Previous</button>
        <button id="nextPage" class="ghost-button" type="button">Next</button>
      </div>
    </section>

    <section class="panel focus-panel">
      <div class="panel-heading">
        <div>
          <h2>Channel Summary</h2>
        </div>
      </div>
      <div id="focusSummary" class="focus-summary"></div>
    </section>

    <section class="kpi-grid bottom-kpis">
      <article class="kpi-card compact-kpi">
        <span>Total Channels</span>
        <strong id="kpiTotal">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span id="kpiLabelOne">Frequency Increased</span>
        <strong id="kpiIncrease">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span id="kpiLabelTwo">Frequency Decreased</span>
        <strong id="kpiDecrease">0</strong>
      </article>
    </section>
  </div>

  <template id="emptyStateTemplate">
    <tr><td colspan="100%" class="empty-state">No records match the current filters.</td></tr>
  </template>

  <script>
const report = __DATA__;
const state = {
  view: "frequency",
  filters: { market: "", city: "", mso_type: "", mso: "", head_end: "", crn_no: "", channel_name: "", band: "", week: "", change: "" },
  sortKey: "channel_name",
  sortDirection: "asc",
  page: 1,
  pageSize: 30,
};

const tableColumns = [
  { key: "market", label: "MARKET" },
  { key: "mso_type", label: "MSO TYPE" },
  { key: "city", label: "CITY" },
  { key: "head_end", label: "HEAD-END" },
  { key: "channel_name", label: "CHANNEL NAME" },
  { key: "band", label: "BAND" },
  { key: "tv_ch_no", label: "TV CH. No." },
  { key: "crn_no", label: "CRN No." },
  { key: "name", label: "NAME" },
];
const filterOrder = ["market", "city", "mso_type", "mso", "head_end", "crn_no", "channel_name", "band", "week", "change"];
const fieldMap = { market: "market", city: "city", mso_type: "mso_type", mso: "mso", head_end: "head_end", crn_no: "crn_no", channel_name: "channel_name", band: "band" };
const filters = {
  market: document.getElementById("marketFilter"),
  city: document.getElementById("cityFilter"),
  mso_type: document.getElementById("msoTypeFilter"),
  mso: document.getElementById("msoFilter"),
  head_end: document.getElementById("headendFilter"),
  crn_no: document.getElementById("crnFilter"),
  channel_name: document.getElementById("channelFilter"),
  band: document.getElementById("bandFilter"),
  week: document.getElementById("weekFilter"),
  change: document.getElementById("changeFilter"),
};
const viewButtons = {
  frequency: document.getElementById("frequencyViewButton"),
  rank: document.getElementById("rankViewButton"),
  band: document.getElementById("bandViewButton"),
};

function formatNumber(value) { return new Intl.NumberFormat().format(value || 0); }
function formatTimestamp(value) {
  const date = new Date(value);
  return Number.isNaN(date.getTime()) ? "--" : date.toLocaleString("en-IN", { year: "numeric", month: "2-digit", day: "2-digit", hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false });
}
function createOption(value, label) {
  const option = document.createElement("option");
  option.value = value;
  option.textContent = label;
  return option;
}
function getViewConfig() {
  if (state.view === "rank") {
    return { series: "ranks", changes: "rank_changes", status: "rank_change_status", positive: "improve", negative: "decline", kpiOne: "Rank Improved", kpiTwo: "Rank Declined", title: "Weekly Rank Analysis" };
  }
  if (state.view === "band") {
    return { series: "bands", changes: "band_changes", status: "band_change_status", positive: "change", negative: "no_change", kpiOne: "Band Changed", kpiTwo: "Band Stable", title: "Weekly Band Analysis" };
  }
  return { series: "frequencies", changes: "changes", status: "change_status", positive: "increase", negative: "decrease", kpiOne: "Frequency Increased", kpiTwo: "Frequency Decreased", title: "Weekly Frequency Analysis" };
}
function filterRecords(ignoreKey = "") {
  const viewConfig = getViewConfig();
  return report.records.filter((record) => {
    for (const [key, field] of Object.entries(fieldMap)) {
      if (key === ignoreKey) continue;
      if (state.filters[key] && String(record[field] || "") !== state.filters[key]) return false;
    }
    if (ignoreKey !== "change" && state.filters.change) {
      if (state.filters.change === "Changed") {
        if (state.filters.week) {
          const changedSet = state.view === "band" ? new Set(["change"]) : new Set(["increase", "decrease", "improve", "decline"]);
          if (!changedSet.has(record[viewConfig.changes][state.filters.week])) return false;
        } else if (record[viewConfig.status] !== "YES") {
          return false;
        }
      }
      if (state.filters.change === "No Change") {
        if (state.filters.week) {
          if (record[viewConfig.changes][state.filters.week] !== "no_change") return false;
        } else if (record[viewConfig.status] !== "NO") {
          return false;
        }
      }
    }
    if (ignoreKey !== "week" && state.filters.week && !record[viewConfig.series][state.filters.week] && record[viewConfig.series][state.filters.week] !== 0) {
      return false;
    }
    return true;
  });
}
function getOptions(key) {
  if (key === "week") return report.weeks.slice();
  if (key === "change") return ["Changed", "No Change"];
  const field = fieldMap[key];
  const values = new Set();
  filterRecords(key).forEach((record) => {
    const value = String(record[field] || "").trim();
    if (value) values.add(value);
  });
  return Array.from(values).sort((a, b) => a.localeCompare(b));
}
function populateSelect(select, values, allLabel, selectedValue) {
  const safeValues = values.filter((value) => value !== null && value !== undefined && String(value).trim() !== "");
  const safeSelectedValue = safeValues.includes(selectedValue) ? selectedValue : "";
  select.innerHTML = "";
  select.appendChild(createOption("", allLabel));
  safeValues.forEach((value) => select.appendChild(createOption(value, value)));
  select.value = safeSelectedValue;
  return safeSelectedValue;
}
function syncFilters() {
  state.filters.market = populateSelect(filters.market, getOptions("market"), "All Markets", state.filters.market);
  state.filters.city = populateSelect(filters.city, getOptions("city"), "All Cities", state.filters.city);
  state.filters.mso_type = populateSelect(filters.mso_type, getOptions("mso_type"), "All MSO Types", state.filters.mso_type);
  state.filters.mso = populateSelect(filters.mso, getOptions("mso"), "All MSO", state.filters.mso);
  state.filters.head_end = populateSelect(filters.head_end, getOptions("head_end"), "All Headend", state.filters.head_end);
  state.filters.crn_no = populateSelect(filters.crn_no, getOptions("crn_no"), "All CRN No", state.filters.crn_no);
  state.filters.channel_name = populateSelect(filters.channel_name, getOptions("channel_name"), "All Channels", state.filters.channel_name);
  state.filters.band = populateSelect(filters.band, getOptions("band"), "All Bands", state.filters.band);
  state.filters.week = populateSelect(filters.week, getOptions("week"), "All Weeks", state.filters.week);
  state.filters.change = populateSelect(filters.change, getOptions("change"), "All Changes", state.filters.change);
}
function buildTableHead() {
  const tr = document.createElement("tr");
  [...tableColumns, ...report.weeks.map((week) => ({ key: week, label: week })), { key: "change_status", label: "CHANGE" }].forEach((column) => {
    const th = document.createElement("th");
    const isActive = state.sortKey === column.key;
    const suffix = isActive ? (state.sortDirection === "asc" ? " ▲" : " ▼") : "";
    th.textContent = column.label + suffix;
    th.className = "sortable";
    th.addEventListener("click", () => {
      if (state.sortKey === column.key) {
        state.sortDirection = state.sortDirection === "asc" ? "desc" : "asc";
      } else {
        state.sortKey = column.key;
        state.sortDirection = "asc";
      }
      state.page = 1;
      render();
    });
    tr.appendChild(th);
  });
  const tableHead = document.getElementById("tableHead");
  tableHead.replaceChildren(tr);
}
function formatWeekValue(value, status, isBaseline) {
  if (value === null || value === undefined || value === "") return "-";
  if (isBaseline || status === "baseline" || status === "missing" || status === "no_change") return String(value);
  if (state.view === "rank") {
    if (status === "improve") return "▲ " + value;
    if (status === "decline") return "▼ " + value;
    return String(value);
  }
  if (state.view === "band") {
    return status === "change" ? "• " + value : String(value);
  }
  if (status === "increase") return "▲ " + value;
  if (status === "decrease") return "▼ " + value;
  return String(value);
}
function sortRecords(records) {
  const viewConfig = getViewConfig();
  return [...records].sort((a, b) => {
    let left;
    let right;
    if (report.weeks.includes(state.sortKey)) {
      left = a[viewConfig.series][state.sortKey];
      right = b[viewConfig.series][state.sortKey];
    } else {
      left = a[state.sortKey];
      right = b[state.sortKey];
    }
    const leftEmpty = left === null || left === undefined || left === "";
    const rightEmpty = right === null || right === undefined || right === "";
    if (leftEmpty && rightEmpty) return 0;
    if (leftEmpty) return 1;
    if (rightEmpty) return -1;
    if (left < right) return state.sortDirection === "asc" ? -1 : 1;
    if (left > right) return state.sortDirection === "asc" ? 1 : -1;
    return 0;
  });
}
function summarize(records) {
  const viewConfig = getViewConfig();
  let positive = 0;
  let negative = 0;
  records.forEach((record) => {
    report.weeks.slice(1).forEach((week) => {
      const status = record[viewConfig.changes][week];
      if (status === viewConfig.positive) positive += 1;
      else if (status === viewConfig.negative) negative += 1;
    });
  });
  return { total: records.length, positive, negative };
}
function renderFocusSummary(records) {
  const focusSummary = document.getElementById("focusSummary");
  const focusChannels = [
    ["INDIA TV", new Set(["INDIA TV"])],
    ["AAJ TAK", new Set(["AAJ TAK"])],
    ["NEWS 18", new Set(["NEWS 18 INDIA"])],
    ["REPUBLIC BHARAT", new Set(["REPUBLIC BHARAT"])],
  ];
  const viewConfig = getViewConfig();
  focusSummary.innerHTML = focusChannels.map(([label, aliases]) => {
    const channelRecords = records.filter((record) => aliases.has(record.channel_name));
    let positive = 0;
    let negative = 0;
    let stable = 0;
    channelRecords.forEach((record) => {
      report.weeks.slice(1).forEach((week) => {
        const status = record[viewConfig.changes][week];
        if (status === viewConfig.positive) positive += 1;
        else if (status === viewConfig.negative) negative += 1;
        else if (status === "no_change") stable += 1;
      });
    });
    return `<div class="focus-line"><strong>${label}</strong><span>${formatNumber(channelRecords.length)} rows, ${formatNumber(positive)} ${viewConfig.kpiOne.toLowerCase()}, ${formatNumber(negative)} ${viewConfig.kpiTwo.toLowerCase()}, ${formatNumber(stable)} stable.</span></div>`;
  }).join("");
}
function render() {
  syncFilters();
  const filtered = sortRecords(filterRecords());
  const summary = summarize(filtered);
  const viewConfig = getViewConfig();
  document.getElementById("generatedAt").textContent = formatTimestamp(report.generated_at);
  document.getElementById("totalRecords").textContent = formatNumber(filtered.length);
  document.getElementById("tableTitle").textContent = viewConfig.title;
  document.getElementById("kpiTotal").textContent = formatNumber(summary.total);
  document.getElementById("kpiIncrease").textContent = formatNumber(summary.positive);
  document.getElementById("kpiDecrease").textContent = formatNumber(summary.negative);
  document.getElementById("kpiLabelOne").textContent = viewConfig.kpiOne;
  document.getElementById("kpiLabelTwo").textContent = viewConfig.kpiTwo;
  Object.entries(viewButtons).forEach(([view, button]) => button.classList.toggle("active", view === state.view));
  buildTableHead();
  renderFocusSummary(filtered);
  const totalPages = Math.max(1, Math.ceil(filtered.length / state.pageSize));
  if (state.page > totalPages) state.page = totalPages;
  const start = (state.page - 1) * state.pageSize;
  const rows = filtered.slice(start, start + state.pageSize);
  document.getElementById("resultCount").textContent = formatNumber(filtered.length) + " records";
  document.getElementById("pageInfo").textContent = "Page " + state.page + " of " + totalPages;
  const tableBody = document.getElementById("tableBody");
  if (!rows.length) {
    tableBody.replaceChildren(document.getElementById("emptyStateTemplate").content.cloneNode(true));
    return;
  }
  const fragment = document.createDocumentFragment();
  rows.forEach((record) => {
    const tr = document.createElement("tr");
    tableColumns.forEach((column) => {
      const td = document.createElement("td");
      td.textContent = record[column.key] || "";
      tr.appendChild(td);
    });
    report.weeks.forEach((week, index) => {
      const td = document.createElement("td");
      const status = index === 0 ? "missing" : record[viewConfig.changes][week];
      td.classList.add("status-" + status);
      td.textContent = formatWeekValue(record[viewConfig.series][week], status, index === 0);
      tr.appendChild(td);
    });
    const changeTd = document.createElement("td");
    changeTd.textContent = record[viewConfig.status];
    changeTd.classList.add(record[viewConfig.status] === "NO" ? "change-no" : "change-yes");
    tr.appendChild(changeTd);
    fragment.appendChild(tr);
  });
  tableBody.replaceChildren(fragment);
}
Object.entries(filters).forEach(([key, select]) => {
  select.addEventListener("change", () => {
    state.filters[key] = select.value;
    const changedIndex = filterOrder.indexOf(key);
    if (changedIndex >= 0) {
      filterOrder.slice(changedIndex + 1).forEach((nextKey) => {
        state.filters[nextKey] = "";
      });
    }
    state.page = 1;
    render();
  });
});
document.getElementById("resetButton").addEventListener("click", () => {
  state.filters = { market: "", city: "", mso_type: "", mso: "", head_end: "", crn_no: "", channel_name: "", band: "", week: "", change: "" };
  state.page = 1;
  render();
});
document.getElementById("prevPage").addEventListener("click", () => { if (state.page > 1) { state.page -= 1; render(); } });
document.getElementById("nextPage").addEventListener("click", () => {
  const totalPages = Math.max(1, Math.ceil(filterRecords().length / state.pageSize));
  if (state.page < totalPages) { state.page += 1; render(); }
});
Object.entries(viewButtons).forEach(([view, button]) => {
  button.addEventListener("click", () => {
    state.view = view;
    state.page = 1;
    render();
  });
});
document.getElementById("saveCopyButton").addEventListener("click", () => window.print());
render();
  </script>
</body>
</html>
"""
    return html.replace("__STYLE__", style_text).replace("__DATA__", report_json)


app = Flask(__name__)


@app.route("/")
def index() -> str:
    load_report()
    return render_template("index.html")


@app.get("/api/frequency")
def api_frequency():
    report = load_report()
    weeks = report["weeks"]
    view = request.args.get("view", "frequency").lower()
    if view not in {"frequency", "rank", "band"}:
        view = "frequency"
    filters = {
        "market": request.args.get("market", ""),
        "mso": request.args.get("mso", ""),
        "mso_type": request.args.get("mso_type", ""),
        "city": request.args.get("city", ""),
        "head_end": request.args.get("head_end", ""),
        "crn_no": request.args.get("crn_no", ""),
        "channel_name": request.args.get("channel_name", ""),
        "band": request.args.get("band", ""),
        "week": request.args.get("week", ""),
        "change": request.args.get("change", ""),
    }
    search = request.args.get("search", "")
    sort_key = request.args.get("sort_key", "channel_name")
    sort_direction = request.args.get("sort_direction", "asc")
    page = max(1, request.args.get("page", default=1, type=int))
    page_size = min(100, max(10, request.args.get("page_size", default=25, type=int)))

    filtered = apply_filters(report["records"], filters, search, view)
    sorted_records = sort_records(filtered, sort_key, sort_direction, weeks, view)
    page_records, total_pages = paginate_records(sorted_records, page, page_size)

    return jsonify(
        {
            "generated_at": report["generated_at"],
            "view": view,
            "weeks": weeks,
            "filters": build_contextual_filters(report["records"], filters, search, view, weeks),
            "summary": summarize_records(filtered, weeks, view),
            "focus_channels": summarize_focus_channels(filtered, weeks, view),
            "message": report.get("message", ""),
            "data_directory": report.get("data_directory", str(DATA_DIR)),
            "table": {
                "records": serialize_records(page_records, weeks, view),
                "page": min(page, total_pages),
                "page_size": page_size,
                "total_count": len(filtered),
                "total_pages": total_pages,
                "sort_key": sort_key,
                "sort_direction": sort_direction,
            },
        }
    )


@app.post("/api/export")
def export_frequency():
    payload = request.get_json(silent=True) or {}
    report = load_report()
    view = str(payload.get("view", "frequency")).lower()
    if view not in {"frequency", "rank", "band"}:
        view = "frequency"
    filters = {
        "market": payload.get("market", ""),
        "mso": payload.get("mso", ""),
        "mso_type": payload.get("mso_type", ""),
        "city": payload.get("city", ""),
        "head_end": payload.get("head_end", ""),
        "crn_no": payload.get("crn_no", ""),
        "channel_name": payload.get("channel_name", ""),
        "band": payload.get("band", ""),
        "week": payload.get("week", ""),
        "change": payload.get("change", ""),
    }
    records = apply_filters(report["records"], filters, payload.get("search", ""), view)
    buffer = create_export_workbook(records, report["weeks"], view)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"chrome_report_{view}_filtered.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/download/dashboard")
def download_dashboard() -> Response:
    report = load_report()
    html = create_shareable_dashboard_html(report)
    buffer = BytesIO(html.encode("utf-8"))
    buffer.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"chrome_report_dashboard_{timestamp}.html",
        mimetype="text/html",
    )


if __name__ == "__main__":
    ensure_directories()
    build_report()
    port = int(os.environ.get("PORT", "9001"))
    debug = os.environ.get("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(host="0.0.0.0", port=port, debug=debug)
