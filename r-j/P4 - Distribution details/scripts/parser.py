from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from headend_id import generate_headend_id, resolve_state

LOGGER = logging.getLogger(__name__)

LABEL_ALIASES = {
    "date": "date",
    "dated": "date",
    "networkname": "network_name",
    "headendlocation": "headend_location",
    "headend": "headend_location",
    "state": "state",
    "barcmarket": "barc_market",
    "stb": "stbs",
    "stbs": "stbs",
    "landingchannel": "landing_channel_1",
    "1stlandingchannel": "landing_channel_1",
    "secondlandingchannel": "landing_channel_2",
    "2ndlandingchannel": "landing_channel_2",
    "barkerchannel": "barker_1",
    "barker": "barker_1",
    "secondbarkerchannel": "barker_2",
    "2ndbarkerchannel": "barker_2",
    "2ndbarker": "barker_2",
}

NULL_LIKE_VALUES = {"", "na", "n/a", "none", "null", "-"}


@dataclass(slots=True)
class ParsedWorkbook:
    distribution_rows: list[dict[str, Any]]
    channel_rows: list[dict[str, Any]]
    week_dates: list[str]


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text or text.lower() in NULL_LIKE_VALUES:
        return None
    return text


def clean_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)

    text = clean_text(value)
    if text is None:
        return None

    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else None


def clean_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)
    if text is None:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%y", "%d-%B-%y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def workbook_to_rows(workbook_path: Path) -> dict[str, list[list[Any]]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for worksheet in workbook.worksheets:
        sheets[worksheet.title] = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return sheets


def detect_label_rows(rows: list[list[Any]]) -> dict[str, int]:
    label_rows: dict[str, int] = {}
    for row_index, row in enumerate(rows[:15]):
        first_cell = row[0] if row else None
        canonical = LABEL_ALIASES.get(normalize_label(first_cell))
        if canonical:
            label_rows[canonical] = row_index
    return label_rows


def detect_blocks(rows: list[list[Any]], lcn_header_row: int) -> list[int]:
    header_row = rows[lcn_header_row] if lcn_header_row < len(rows) else []
    block_starts: list[int] = []
    for col_index in range(max(0, len(header_row) - 1)):
        left = normalize_label(header_row[col_index])
        right = normalize_label(header_row[col_index + 1]) if col_index + 1 < len(header_row) else ""
        if left in {"lcn", "lcnno", "lcnnumber"} and right in {"channel", "channelname", "channelnames"}:
            block_starts.append(col_index)
    return block_starts


def get_cell(rows: list[list[Any]], row_index: int | None, col_index: int) -> Any:
    if row_index is None or row_index < 0 or row_index >= len(rows):
        return None
    row = rows[row_index]
    if col_index >= len(row):
        return None
    return row[col_index]


def parse_sheet(sheet_name: str, rows: list[list[Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if len(rows) < 11:
        return [], []

    label_rows = detect_label_rows(rows)
    lcn_header_row = 10
    block_starts = detect_blocks(rows, lcn_header_row)

    if not block_starts:
        LOGGER.warning("Skipping sheet '%s' because no LCN blocks were detected.", sheet_name)
        return [], []

    distribution_rows: list[dict[str, Any]] = []
    channel_rows: list[dict[str, Any]] = []

    for block_start in block_starts:
        record_date = clean_date(get_cell(rows, label_rows.get("date"), block_start))
        network_name = clean_text(get_cell(rows, label_rows.get("network_name"), block_start))
        headend_location = clean_text(get_cell(rows, label_rows.get("headend_location"), block_start))
        source_state = clean_text(get_cell(rows, label_rows.get("state"), block_start)) or sheet_name
        state = resolve_state(
            source_state=source_state,
            headend_location=headend_location,
            sheet_name=sheet_name,
        )
        barc_market = clean_text(get_cell(rows, label_rows.get("barc_market"), block_start))
        stbs = clean_int(get_cell(rows, label_rows.get("stbs"), block_start))
        landing_channel_1 = clean_text(get_cell(rows, label_rows.get("landing_channel_1"), block_start))
        landing_channel_2 = clean_text(get_cell(rows, label_rows.get("landing_channel_2"), block_start))
        barker_1 = clean_text(get_cell(rows, label_rows.get("barker_1"), block_start))
        barker_2 = clean_text(get_cell(rows, label_rows.get("barker_2"), block_start))

        if not network_name and not headend_location:
            continue

        headend_id = generate_headend_id(
            network_name=network_name,
            headend_location=headend_location,
            state=state,
        )

        distribution_rows.append(
            {
                "headend_id": headend_id,
                "date": record_date,
                "network_name": network_name,
                "headend_location": headend_location,
                "state": state,
                "barc_market": barc_market,
                "stbs": stbs,
                "landing_channel_1": landing_channel_1,
                "landing_channel_2": landing_channel_2,
                "barker_1": barker_1,
                "barker_2": barker_2,
                "source_sheet": sheet_name,
            }
        )

        for row_index in range(lcn_header_row + 1, len(rows)):
            lcn_no = clean_text(get_cell(rows, row_index, block_start))
            channel_name = clean_text(get_cell(rows, row_index, block_start + 1))

            if not lcn_no and not channel_name:
                continue

            if normalize_label(get_cell(rows, row_index, block_start)) in {"lcn", "lcnno", "lcnnumber"}:
                continue

            if lcn_no is None or channel_name is None:
                continue

            channel_rows.append(
                {
                    "headend_id": headend_id,
                    "date": record_date,
                    "network_name": network_name,
                    "headend_location": headend_location,
                    "state": state,
                    "barc_market": barc_market,
                    "lcn_no": lcn_no,
                    "channel_name": channel_name,
                }
            )

    return distribution_rows, channel_rows


def parse_workbook(workbook_path: Path) -> ParsedWorkbook:
    LOGGER.info("Reading workbook: %s", workbook_path)
    workbook_rows = workbook_to_rows(workbook_path)
    distribution_rows: list[dict[str, Any]] = []
    channel_rows: list[dict[str, Any]] = []

    for sheet_name, rows in workbook_rows.items():
        LOGGER.info("Parsing sheet: %s", sheet_name)
        sheet_distribution, sheet_channels = parse_sheet(sheet_name, rows)
        distribution_rows.extend(sheet_distribution)
        channel_rows.extend(sheet_channels)

    distribution_rows = list(
        {
            (
                row["headend_id"],
                row.get("date"),
                row.get("network_name"),
                row.get("headend_location"),
                row.get("state"),
            ): row
            for row in distribution_rows
        }.values()
    )
    channel_rows = list(
        {
            (
                row["headend_id"],
                row.get("date"),
                row.get("lcn_no"),
                row.get("channel_name"),
            ): row
            for row in channel_rows
        }.values()
    )
    week_dates = sorted({row["date"] for row in channel_rows if row.get("date")})

    LOGGER.info(
        "Parsed workbook into %s headend snapshots and %s channel rows across %s weeks.",
        len(distribution_rows),
        len(channel_rows),
        len(week_dates),
    )

    return ParsedWorkbook(
        distribution_rows=distribution_rows,
        channel_rows=channel_rows,
        week_dates=week_dates,
    )
