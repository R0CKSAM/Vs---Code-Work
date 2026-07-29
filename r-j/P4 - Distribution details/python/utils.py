from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

import openpyxl
import pandas as pd

LOGGER = logging.getLogger(__name__)

HEADEND_OUTPUT_COLUMNS = [
    "Network_Name",
    "Headend",
    "State",
    "BARC_Market",
    "STB",
    "Landing_Channel",
    "Second_Landing_Channel",
    "Barker",
    "Second_Barker",
]

LABEL_ALIASES = {
    "dated": "Dated",
    "date": "Dated",
    "networkname": "Network_Name",
    "headendlocation": "Headend",
    "headend": "Headend",
    "state": "State",
    "barcmarket": "BARC_Market",
    "stb": "STB",
    "stbs": "STB",
    "landingchannel": "Landing_Channel",
    "2ndlandingchannel": "Second_Landing_Channel",
    "secondlandingchannel": "Second_Landing_Channel",
    "barkerchannel": "Barker",
    "barker": "Barker",
    "2ndbarkerchannel": "Second_Barker",
    "secondbarkerchannel": "Second_Barker",
    "2ndbarker": "Second_Barker",
}

LCN_HEADER_ALIASES = {"lcn", "lcnno", "lcnnumber"}
CHANNEL_HEADER_ALIASES = {"channel", "channelname", "channelnames"}
POSITION_HEADER_ALIASES = {"position", "channelposition", "channel_position"}
NULL_LIKE_VALUES = {"", "na", "n/a", "none", "null", "-"}


@dataclass(slots=True)
class ParsedWorkbook:
    headends: pd.DataFrame
    channels: pd.DataFrame


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text or text.lower() in NULL_LIKE_VALUES:
        return None
    return text


def clean_stb(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)

    text = clean_text(value)
    if text is None:
        return None

    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else None


def clean_week(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    if text is None:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%y", "%d-%B-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def iter_excel_files(root: Path) -> Iterator[Path]:
    for folder_name in ("data", "Data"):
        folder = root / folder_name
        if not folder.exists():
            continue
        for path in sorted(folder.glob("*.xlsx")):
            if not path.name.startswith("~$"):
                yield path


def detect_label_rows(rows: list[list[Any]]) -> dict[str, int]:
    label_rows: dict[str, int] = {}

    for row_index, row in enumerate(rows):
        for value in row:
            canonical = LABEL_ALIASES.get(normalize_label(value))
            if canonical and canonical not in label_rows:
                label_rows[canonical] = row_index

    return label_rows


def detect_lcn_header_row(rows: list[list[Any]]) -> int:
    best_row = -1
    best_score = -1

    for row_index, row in enumerate(rows):
        normalized_cells = [normalize_label(value) for value in row if value not in (None, "")]
        if not normalized_cells:
            continue

        lcn_hits = sum(cell in LCN_HEADER_ALIASES for cell in normalized_cells)
        channel_hits = sum(cell in CHANNEL_HEADER_ALIASES for cell in normalized_cells)
        score = lcn_hits + channel_hits

        if lcn_hits and channel_hits and score > best_score:
            best_row = row_index
            best_score = score

    if best_row == -1:
        raise ValueError("Unable to detect the LCN header row in the worksheet.")

    return best_row


def get_cell(rows: list[list[Any]], row_index: int | None, col_index: int) -> Any:
    if row_index is None or row_index >= len(rows):
        return None
    row = rows[row_index]
    if col_index >= len(row):
        return None
    return row[col_index]


def detect_headend_blocks(rows: list[list[Any]], lcn_header_row: int) -> list[tuple[int, bool]]:
    header_row = rows[lcn_header_row]
    blocks: list[tuple[int, bool]] = []

    for col_index, value in enumerate(header_row):
        if normalize_label(value) not in LCN_HEADER_ALIASES:
            continue

        right_1 = normalize_label(header_row[col_index + 1]) if col_index + 1 < len(header_row) else ""
        right_2 = normalize_label(header_row[col_index + 2]) if col_index + 2 < len(header_row) else ""

        if right_1 in CHANNEL_HEADER_ALIASES:
            has_explicit_position = right_2 in POSITION_HEADER_ALIASES
            blocks.append((col_index, has_explicit_position))

    if not blocks:
        raise ValueError("Unable to detect any headend blocks from the LCN header row.")

    return blocks


def parse_sheet(sheet_name: str, rows: list[list[Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    label_rows = detect_label_rows(rows)
    lcn_header_row = detect_lcn_header_row(rows)
    blocks = detect_headend_blocks(rows, lcn_header_row)

    if "Network_Name" not in label_rows:
        raise ValueError(f"Sheet '{sheet_name}' is missing the 'Network Name' metadata row.")

    headend_records: list[dict[str, Any]] = []
    channel_records: list[dict[str, Any]] = []

    for block_start, has_explicit_position in blocks:
        metadata = {
            "Sheet_Name": sheet_name,
            "Week": clean_week(get_cell(rows, label_rows.get("Dated"), block_start)),
            "Network_Name": clean_text(get_cell(rows, label_rows.get("Network_Name"), block_start)),
            "Headend": clean_text(get_cell(rows, label_rows.get("Headend"), block_start)),
            "State": clean_text(get_cell(rows, label_rows.get("State"), block_start)) or sheet_name,
            "BARC_Market": clean_text(get_cell(rows, label_rows.get("BARC_Market"), block_start)),
            "STB": clean_stb(get_cell(rows, label_rows.get("STB"), block_start)),
            "Landing_Channel": clean_text(get_cell(rows, label_rows.get("Landing_Channel"), block_start)),
            "Second_Landing_Channel": clean_text(get_cell(rows, label_rows.get("Second_Landing_Channel"), block_start)),
            "Barker": clean_text(get_cell(rows, label_rows.get("Barker"), block_start)),
            "Second_Barker": clean_text(get_cell(rows, label_rows.get("Second_Barker"), block_start)),
        }

        if not metadata["Network_Name"] and not metadata["Headend"]:
            continue

        headend_records.append(metadata)

        running_position = 0
        for row_index in range(lcn_header_row + 1, len(rows)):
            lcn_value = get_cell(rows, row_index, block_start)
            channel_value = get_cell(rows, row_index, block_start + 1)
            position_value = get_cell(rows, row_index, block_start + 2) if has_explicit_position else None

            lcn = clean_text(lcn_value)
            channel_name = clean_text(channel_value)

            if not lcn and not channel_name:
                continue

            # Skip repeated headers or malformed footer text.
            if normalize_label(lcn_value) in LCN_HEADER_ALIASES or normalize_label(channel_value) in CHANNEL_HEADER_ALIASES:
                continue

            running_position += 1
            channel_position = clean_stb(position_value) if has_explicit_position else None

            channel_records.append(
                {
                    "Sheet_Name": sheet_name,
                    "Headend_Key_Network_Name": metadata["Network_Name"],
                    "Headend_Key_Headend": metadata["Headend"],
                    "Headend_Key_State": metadata["State"],
                    "Headend_Key_BARC_Market": metadata["BARC_Market"],
                    "Week": metadata["Week"],
                    "LCN": lcn,
                    "Channel_Name": channel_name,
                    "Channel_Position": channel_position or running_position,
                }
            )

    return headend_records, channel_records


def workbook_to_rows(workbook_path: Path) -> dict[str, list[list[Any]]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    all_rows: dict[str, list[list[Any]]] = {}

    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        all_rows[worksheet.title] = rows

    return all_rows


def parse_workbook(workbook_path: Path) -> ParsedWorkbook:
    LOGGER.info("Reading workbook: %s", workbook_path)
    workbook_rows = workbook_to_rows(workbook_path)

    all_headends: list[dict[str, Any]] = []
    all_channels: list[dict[str, Any]] = []

    for sheet_name, rows in workbook_rows.items():
        LOGGER.info("Parsing sheet: %s", sheet_name)
        headends, channels = parse_sheet(sheet_name=sheet_name, rows=rows)
        all_headends.extend(headends)
        all_channels.extend(channels)

    headends_df = pd.DataFrame(all_headends)
    channels_df = pd.DataFrame(all_channels)

    if headends_df.empty:
        raise ValueError("No headend records were extracted from the workbook.")

    if channels_df.empty:
        LOGGER.warning("No channel records were extracted from the workbook.")

    for column in HEADEND_OUTPUT_COLUMNS:
        if column not in headends_df.columns:
            headends_df[column] = None

    headends_df = (
        headends_df.drop_duplicates(
            subset=["Network_Name", "Headend", "State", "BARC_Market"],
            keep="first",
        )
        .reset_index(drop=True)
    )

    channels_df = channels_df.drop_duplicates(
        subset=[
            "Headend_Key_Network_Name",
            "Headend_Key_Headend",
            "Headend_Key_State",
            "Headend_Key_BARC_Market",
            "Week",
            "LCN",
            "Channel_Name",
            "Channel_Position",
        ],
        keep="first",
    ).reset_index(drop=True)

    invalid_channel_mask = channels_df["LCN"].isna() | channels_df["Channel_Name"].isna()
    invalid_channel_count = int(invalid_channel_mask.sum())
    if invalid_channel_count:
        LOGGER.warning(
            "Skipping %s incomplete channel rows that are missing either LCN or Channel_Name.",
            invalid_channel_count,
        )
        channels_df = channels_df.loc[~invalid_channel_mask].reset_index(drop=True)

    LOGGER.info(
        "Parsed workbook into %s unique headends and %s channel rows.",
        len(headends_df),
        len(channels_df),
    )

    return ParsedWorkbook(headends=headends_df, channels=channels_df)
