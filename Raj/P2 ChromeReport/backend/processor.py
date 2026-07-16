from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .utils import (
    COLUMN_ALIASES,
    DIMENSION_COLUMNS,
    available_measure_columns,
    build_business_key,
    ensure_output_dir,
    measure_column_name,
    normalize_frequency,
    normalize_slug,
    normalize_text,
    scan_week_files,
)


@dataclass
class ProcessingSummary:
    source_files: list[str]
    total_rows: int
    total_weeks: int
    parquet_path: Path


class FrequencyDatasetProcessor:
    def __init__(self, data_dir: Path, output_path: Path) -> None:
        self.data_dir = data_dir
        self.output_path = output_path

    def parquet_is_fresh(self) -> bool:
        if not self.output_path.exists():
            return False
        parquet_mtime = self.output_path.stat().st_mtime
        source_files = scan_week_files(self.data_dir)
        if not source_files:
            return True
        return all(path.stat().st_mtime <= parquet_mtime for path in source_files)

    def process_if_needed(self, force: bool = False) -> ProcessingSummary:
        if force or not self.parquet_is_fresh():
            return self.process()

        df = pd.read_parquet(self.output_path)
        source_files = scan_week_files(self.data_dir)
        return ProcessingSummary(
            source_files=[path.name for path in source_files],
            total_rows=len(df.index),
            total_weeks=len(available_measure_columns(df, "F")),
            parquet_path=self.output_path,
        )

    def process(self) -> ProcessingSummary:
        source_files = scan_week_files(self.data_dir)
        if not source_files:
            raise FileNotFoundError(f"No weekly Excel files found in {self.data_dir}")

        merged_df: pd.DataFrame | None = None

        for week_index, path in enumerate(source_files, start=1):
            weekly_df = self._read_week_file(path, week_index)
            if merged_df is None:
                merged_df = weekly_df
            else:
                merged_df = merged_df.merge(weekly_df, on="Business Key", how="outer", suffixes=("", "__new"))
                for column in DIMENSION_COLUMNS:
                    new_column = f"{column}__new"
                    if new_column in merged_df.columns:
                        merged_df[column] = (
                            merged_df[column]
                            .replace("", pd.NA)
                            .fillna(merged_df[new_column].replace("", pd.NA))
                            .fillna("")
                        )
                        merged_df = merged_df.drop(columns=[new_column])

        assert merged_df is not None
        merged_df = self._finalize_dataset(merged_df)

        ensure_output_dir(self.output_path)
        merged_df.to_parquet(self.output_path, engine="pyarrow", index=False)

        return ProcessingSummary(
            source_files=[path.name for path in source_files],
            total_rows=len(merged_df.index),
            total_weeks=len(available_measure_columns(merged_df, "F")),
            parquet_path=self.output_path,
        )

    def _read_week_file(self, path: Path, week_index: int) -> pd.DataFrame:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            wb.close()
            raise ValueError(f"{path.name} is empty")

        selected_columns: list[tuple[int, str]] = []
        for index, column_name in enumerate(header):
            slug = normalize_slug(column_name)
            if slug in COLUMN_ALIASES:
                selected_columns.append((index, COLUMN_ALIASES[slug]))

        if not any(name == "Frequency" for _, name in selected_columns):
            wb.close()
            raise ValueError(f"{path.name} is missing a Frequency column")

        records: list[dict[str, str | None]] = []
        for raw_row in rows:
            record: dict[str, str | None] = {}
            has_value = False
            for index, canonical_name in selected_columns:
                value = raw_row[index] if index < len(raw_row) else None
                if canonical_name in {"Frequency", "Rank"}:
                    normalized_value = normalize_frequency(value)
                else:
                    normalized_value = normalize_text(value)
                if normalized_value not in {"", None}:
                    has_value = True
                record[canonical_name] = normalized_value
            if has_value:
                records.append(record)
        wb.close()

        weekly_df = pd.DataFrame.from_records(records)
        for column in DIMENSION_COLUMNS:
            if column not in weekly_df.columns:
                weekly_df[column] = ""
            weekly_df[column] = weekly_df[column].map(normalize_text)

        if "Frequency" not in weekly_df.columns:
            weekly_df["Frequency"] = None
        if "Rank" not in weekly_df.columns:
            weekly_df["Rank"] = None
        if "Band" not in weekly_df.columns:
            weekly_df["Band"] = ""

        weekly_df["Frequency"] = weekly_df["Frequency"].map(normalize_frequency)
        weekly_df["Rank"] = weekly_df["Rank"].map(normalize_frequency)
        weekly_df["Band"] = weekly_df["Band"].map(normalize_text)
        weekly_df["Band Measure"] = weekly_df["Band"]
        weekly_df["Business Key"] = weekly_df.apply(build_business_key, axis=1)

        frequency_column = measure_column_name("F", week_index)
        rank_column = measure_column_name("R", week_index)
        band_column = measure_column_name("B", week_index)

        keep_columns = ["Business Key", *DIMENSION_COLUMNS, "Frequency", "Rank", "Band Measure"]
        weekly_df = weekly_df[keep_columns].rename(
            columns={
                "Frequency": frequency_column,
                "Rank": rank_column,
                "Band Measure": band_column,
            }
        )
        weekly_df = weekly_df.drop_duplicates(subset=["Business Key"], keep="last").reset_index(drop=True)
        return weekly_df

    def _finalize_dataset(self, df: pd.DataFrame) -> pd.DataFrame:
        frequency_columns = available_measure_columns(df, "F")
        rank_columns = available_measure_columns(df, "R")
        band_columns = available_measure_columns(df, "B")

        for column in frequency_columns + rank_columns:
            df[column] = df[column].map(lambda value: normalize_frequency(value) or "NA")
        for column in band_columns:
            df[column] = df[column].map(lambda value: normalize_text(value) or "NA")

        df["Frequency Total Changes"] = df.apply(lambda row: self._count_numeric_changes(row, frequency_columns), axis=1)
        df["Frequency Status"] = df.apply(lambda row: self._latest_numeric_status(row, frequency_columns, rank_mode=False), axis=1)
        df["Rank Total Changes"] = df.apply(lambda row: self._count_numeric_changes(row, rank_columns), axis=1)
        df["Rank Status"] = df.apply(lambda row: self._latest_numeric_status(row, rank_columns, rank_mode=True), axis=1)
        df["Band Total Changes"] = df.apply(lambda row: self._count_text_changes(row, band_columns), axis=1)
        df["Band Status"] = df.apply(lambda row: self._latest_band_status(row, band_columns), axis=1)

        # Keep legacy/default frequency aliases for existing UI flows.
        df["Total Changes"] = df["Frequency Total Changes"]
        df["Status"] = df["Frequency Status"]

        ordered_columns = [
            "Market",
            "MSO",
            "City",
            "Head End",
            "Channel Name",
            "CR No",
            "Transmission",
            "MSO Type",
            "Band",
            "TV Channel No",
            *frequency_columns,
            *rank_columns,
            *band_columns,
            "Frequency Total Changes",
            "Frequency Status",
            "Rank Total Changes",
            "Rank Status",
            "Band Total Changes",
            "Band Status",
            "Total Changes",
            "Status",
            "Business Key",
        ]
        existing = [column for column in ordered_columns if column in df.columns]
        remaining = [column for column in df.columns if column not in existing]
        df = df[existing + remaining]
        return df.sort_values(
            by=["Market", "MSO", "City", "Head End", "Channel Name", "CR No"],
            na_position="last",
        ).reset_index(drop=True)

    @staticmethod
    def _count_numeric_changes(row: pd.Series, week_columns: list[str]) -> int:
        values = [row[column] for column in week_columns if row[column] != "NA"]
        if len(values) < 2:
            return 0
        changes = 0
        previous = values[0]
        for current in values[1:]:
            if current != previous:
                changes += 1
            previous = current
        return changes

    @staticmethod
    def _count_text_changes(row: pd.Series, week_columns: list[str]) -> int:
        values = [row[column] for column in week_columns if row[column] != "NA"]
        if len(values) < 2:
            return 0
        changes = 0
        previous = values[0]
        for current in values[1:]:
            if current != previous:
                changes += 1
            previous = current
        return changes

    @staticmethod
    def _latest_numeric_status(row: pd.Series, week_columns: list[str], rank_mode: bool) -> str:
        valid_values = [row[column] for column in week_columns if row[column] != "NA"]
        if len(valid_values) < 2:
            return "→ No Change"
        previous = float(valid_values[-2])
        current = float(valid_values[-1])
        if rank_mode:
            if current < previous:
                return "↑ Improved"
            if current > previous:
                return "↓ Declined"
            return "→ No Change"
        if current > previous:
            return "↑ Increased"
        if current < previous:
            return "↓ Decreased"
        return "→ No Change"

    @staticmethod
    def _latest_band_status(row: pd.Series, week_columns: list[str]) -> str:
        valid_values = [row[column] for column in week_columns if row[column] != "NA"]
        if len(valid_values) < 2:
            return "→ No Change"
        return "↑ Changed" if valid_values[-1] != valid_values[-2] else "→ No Change"
