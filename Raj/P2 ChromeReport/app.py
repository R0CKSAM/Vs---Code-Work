from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from weekly_workbook_builder import ensure_combined_weekly_workbooks


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DISTRIBUTION_SUMMARY_DIR = BASE_DIR / "distribution summary"
NBHD_DATA_DIR = BASE_DIR / "NBHD Data"
LEGACY_NBHD_DATA_DIR = BASE_DIR / "NBHD"
OTS_DATA_DIR = BASE_DIR / "OTS"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_JSON = OUTPUT_DIR / "frequency_report.json"
OUTPUT_HTML = OUTPUT_DIR / "chrome_report_dashboard.html"
HISTORY_DIR = BASE_DIR / "history"
HISTORY_DISTRIBUTION_CSV = HISTORY_DIR / "distribution_history.csv"
HISTORY_NBHD_CSV = HISTORY_DIR / "nbhd_history.csv"
HISTORY_OTS_CSV = HISTORY_DIR / "ots_history.csv"
LEGACY_HISTORY_DISTRIBUTION_CSV = BASE_DIR / "distribution_history.csv"
LEGACY_HISTORY_NBHD_CSV = BASE_DIR / "nbhd_history.csv"
LEGACY_HISTORY_OTS_CSV = BASE_DIR / "ots_history.csv"
STYLE_FILE = BASE_DIR / "static" / "style.css"
NBHD_SCRIPT_FILE = BASE_DIR / "static" / "neighbourhood.js"
OTS_SCRIPT_FILE = BASE_DIR / "static" / "ots.js"
COMPARISON_SCRIPT_FILE = BASE_DIR / "static" / "comparison.js"
NBHD_BENCHMARK_SCRIPT_FILE = BASE_DIR / "static" / "nbhd_benchmark.js"

SOURCE_COLUMNS = [
    "WEEK LABEL",
    "TRANSMISSION",
    "MARKET",
    "GENRE",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "FREQUENCY/LCN NO",
    "BAND",
    "TV CH. No.",
    "AUDIO",
    "VIDEO",
    "LANGUAGE",
    "CRN No.",
    "RANK WITHIN GENRE",
]
KEY_COLUMNS = [
    "TRANSMISSION",
    "MARKET",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "CRN No.",
]
DISPLAY_COLUMNS = [
    "TRANSMISSION",
    "MARKET",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "WEEK LABEL",
    "GENRE",
    "LANGUAGE",
    "NAME",
]
FREQUENCY_COLUMN = "FREQUENCY/LCN NO"
RANK_COLUMN = "RANK WITHIN GENRE"

FOCUS_CHANNELS = {
    "INDIA TV": "India TV",
    "AAJ TAK": "Aaj Tak",
    "NEWS 18 INDIA": "News 18",
    "REPUBLIC BHARAT": "Republic Bharat",
}

REPORT_CACHE: dict[str, Any] = {
    "signature": None,
    "report": None,
}
NBHD_REPORT_CACHE: dict[str, Any] = {
    "signature": None,
    "report": None,
}
OTS_REPORT_CACHE: dict[str, Any] = {
    "signature": None,
    "report": None,
}
COMPARISON_REPORT_CACHE: dict[str, Any] = {
    "signature": None,
    "report": None,
}
NBHD_BENCHMARK_REPORT_CACHE: dict[str, Any] = {
    "signature": None,
    "report": None,
}


def ensure_directories() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    NBHD_DATA_DIR.mkdir(parents=True, exist_ok=True)
    OTS_DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_number(value: Any) -> float | int | None:
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group())
    return int(number) if number.is_integer() else number


def normalize_rank(value: Any) -> int | None:
    number = normalize_number(value)
    if number is None:
        return None
    return int(number)


def week_sort_key(label: str) -> tuple[int, str]:
    match = re.search(r"(\d+)", label)
    if match:
        return int(match.group(1)), label
    return 10**9, label


def parse_workbook_week(path: Path) -> tuple[int, int] | None:
    match = re.search(r"wk[-\s_]*(\d{1,2}).*?(20\d{2})", path.stem, re.IGNORECASE)
    if match:
        return int(match.group(2)), int(match.group(1))
    match = re.search(r"(20\d{2}).*?wk[-\s_]*(\d{1,2})", path.stem, re.IGNORECASE)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def format_week_label(week_number: int, year: int) -> str:
    return f"Wk-{int(week_number):02d}'{str(year)[-2:]}"


def get_signature(files: list[Path]) -> tuple[tuple[str, int, int], ...]:
    return tuple((path.name, int(path.stat().st_mtime), path.stat().st_size) for path in files)


def history_files_ready() -> bool:
    paths = [resolve_history_path(HISTORY_DISTRIBUTION_CSV, LEGACY_HISTORY_DISTRIBUTION_CSV), resolve_history_path(HISTORY_NBHD_CSV, LEGACY_HISTORY_NBHD_CSV), resolve_history_path(HISTORY_OTS_CSV, LEGACY_HISTORY_OTS_CSV)]
    return all(path.exists() and path.stat().st_size > 0 for path in paths)


def get_history_signature() -> tuple[tuple[str, int, int], ...]:
    paths = [
        path
        for path in [
            resolve_history_path(HISTORY_DISTRIBUTION_CSV, LEGACY_HISTORY_DISTRIBUTION_CSV),
            resolve_history_path(HISTORY_NBHD_CSV, LEGACY_HISTORY_NBHD_CSV),
            resolve_history_path(HISTORY_OTS_CSV, LEGACY_HISTORY_OTS_CSV),
        ]
        if path.exists()
    ]
    return get_signature(paths)


def read_history_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, keep_default_na=False, na_values=[""], low_memory=False)


def resolve_history_path(primary: Path, legacy: Path) -> Path:
    if primary.exists():
        return primary
    return legacy


def normalize_header_key(value: Any) -> str:
    text = normalize_text(value).upper()
    return re.sub(r"[^A-Z0-9]+", "", text)


def get_nbhd_input_dir() -> Path:
    ensure_directories()
    primary_files = list(NBHD_DATA_DIR.glob("*.xlsx"))
    if primary_files:
        return NBHD_DATA_DIR
    return LEGACY_NBHD_DATA_DIR if LEGACY_NBHD_DATA_DIR.exists() else NBHD_DATA_DIR


def get_nbhd_source_dir() -> Path:
    ensure_directories()
    combined_files = list(DATA_DIR.glob("*.xlsx"))
    if combined_files:
        return DATA_DIR
    return get_nbhd_input_dir()


def get_ots_input_dir() -> Path:
    ensure_directories()
    return OTS_DATA_DIR


def get_ots_source_dir() -> Path:
    ensure_directories()
    combined_files = list(DATA_DIR.glob("*.xlsx"))
    if combined_files:
        return DATA_DIR
    return get_ots_input_dir()


def get_week_files() -> list[Path]:
    ensure_directories()
    ensure_combined_weekly_workbooks(DATA_DIR, DISTRIBUTION_SUMMARY_DIR, get_nbhd_input_dir(), get_ots_input_dir())
    return sorted(
        [path for path in DATA_DIR.glob("*.xlsx") if not path.name.startswith("~$")],
        key=lambda path: week_sort_key(path.stem),
    )


def get_nbhd_week_files() -> list[Path]:
    source_dir = get_nbhd_source_dir()
    files = [path for path in source_dir.glob("*.xlsx") if not path.name.startswith("~$")]
    return sorted(files, key=lambda path: week_sort_key(path.stem))


def get_ots_week_files() -> list[Path]:
    ensure_directories()
    files = [path for path in get_ots_source_dir().glob("*.xlsx") if not path.name.startswith("~$")]
    return sorted(files, key=lambda path: week_sort_key(path.stem))


def get_highlight_week_files() -> list[Path]:
    ensure_directories()
    files = [
        path
        for pattern in ("*.xlsm", "*.xlsx")
        for path in DATA_DIR.glob(pattern)
        if not path.name.startswith("~$")
    ]
    return sorted(
        {path.resolve(): path for path in files}.values(),
        key=lambda path: parse_workbook_week(path) or (10**9, 10**9),
    )


def empty_nbhd_report(message: str | None = None) -> dict[str, Any]:
    source_dir = get_nbhd_source_dir()
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": [],
        "records": [],
        "message": message or "Add weekly neighbourhood Excel files to the NBHD Data folder.",
        "source_directory": str(source_dir),
    }


def empty_ots_report(message: str | None = None) -> dict[str, Any]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": [],
        "records": [],
        "message": message or "Add weekly OTS Excel files to the data folder.",
        "source_directory": str(get_ots_source_dir()),
    }


def prepare_nbhd_week_rows(path: Path) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["NBDH Data"] if "NBDH Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_map = {normalize_header_key(value): index for index, value in enumerate(header_row)}

    field_aliases = {
        "type": ("TYPE",),
        "frequency": ("FREQU", "FREQUENCY"),
        "tv_ch_no": ("TVCHANNELNO", "TVCHNO"),
        "market": ("MARKET",),
        "city": ("CITY",),
        "head_end": ("HEADEND",),
        "channel": ("CHANNEL",),
        "genre": ("GENRE",),
    }

    field_indexes: dict[str, int] = {}
    missing_fields: list[str] = []
    for field, aliases in field_aliases.items():
        index = next((header_map[alias] for alias in aliases if alias in header_map), None)
        if index is None:
            missing_fields.append(field)
        else:
            field_indexes[field] = index

    if missing_fields:
        workbook.close()
        raise ValueError(f"{path.name} is missing required columns: {', '.join(missing_fields)}")

    week_label = normalize_text(path.stem).upper()
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {
            field: values[index] if index < len(values) else None
            for field, index in field_indexes.items()
        }
        market = normalize_text(row["market"])
        city = normalize_text(row["city"])
        head_end = normalize_text(row["head_end"])
        channel = normalize_text(row["channel"])
        if not market or not city or not head_end or not channel:
            continue
        rows.append(
            {
                "type": normalize_text(row["type"]),
                "market": market,
                "city": city,
                "head_end": head_end,
                "channel": channel,
                "genre": normalize_text(row["genre"]),
                "frequency": normalize_number(row["frequency"]),
                "tv_ch_no": normalize_number(row["tv_ch_no"]),
                "order_token": normalize_number(row["tv_ch_no"]) if normalize_number(row["tv_ch_no"]) is not None else normalize_number(row["frequency"]),
            }
        )

    workbook.close()
    return week_label, rows


def normalize_ots_percentage(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    number = normalize_number(value)
    if number is None:
        return None
    normalized = float(number)
    if "%" not in text and abs(normalized) <= 1:
        normalized *= 100
    return round(normalized, 2)


def prepare_ots_week_rows(path: Path) -> tuple[str, list[dict[str, Any]]]:
    # Read one normalized weekly OTS workbook into market/channel/value rows.
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "OTS Data" in workbook.sheetnames:
        sheet_name = "OTS Data"
    elif "Table1" in workbook.sheetnames:
        sheet_name = "Table1"
    else:
        sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_map = {normalize_header_key(value): index for index, value in enumerate(header_row)}

    required_headers = {
        "market": ("MARKET",),
        "channel": ("CHANNEL", "ATTRIBUTE"),
        "ots": ("OTS", "VALUE"),
    }
    field_indexes: dict[str, int] = {}
    missing_fields: list[str] = []
    for field, aliases in required_headers.items():
        index = next((header_map[alias] for alias in aliases if alias in header_map), None)
        if index is None:
            missing_fields.append(field)
        else:
            field_indexes[field] = index

    if missing_fields:
        workbook.close()
        raise ValueError(f"{path.name} is missing required columns: {', '.join(missing_fields)}")

    week_index = next((header_map[alias] for alias in ("WEEK",) if alias in header_map), None)

    week_label = normalize_text(path.stem)
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        market = normalize_text(values[field_indexes["market"]] if field_indexes["market"] < len(values) else None)
        channel = normalize_text(values[field_indexes["channel"]] if field_indexes["channel"] < len(values) else None)
        ots_value = normalize_ots_percentage(values[field_indexes["ots"]] if field_indexes["ots"] < len(values) else None)
        if not market or not channel:
            continue
        week_value = normalize_text(values[week_index] if week_index is not None and week_index < len(values) else None)
        if week_value:
            week_label = week_value.replace("Week", "Week ").replace("week", "Week ").replace("  ", " ").strip()

        rows.append(
            {
                "market": market,
                "channel": channel,
                "ots": ots_value,
            }
        )

    workbook.close()

    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        deduped.setdefault(f"{row['market']}||{row['channel']}", row)
    return week_label, list(deduped.values())


def get_nbhd_sorted_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    indexed_rows = list(enumerate(rows))
    indexed_rows.sort(
        key=lambda item: (
            item[1].get("order_token") is None,
            item[1].get("order_token") if item[1].get("order_token") is not None else item[0],
            item[0],
        )
    )
    return [row for _, row in indexed_rows]


def extract_nbhd_window(rows: list[dict[str, Any]], radius: int = 4) -> list[tuple[int, dict[str, Any]]]:
    india_rows = [row for row in rows if normalize_text(row["channel"]).upper() == "INDIA TV"]
    if not india_rows:
        return []

    india_row = india_rows[0]
    sorted_rows = get_nbhd_sorted_rows(rows)
    india_index = next((index for index, row in enumerate(sorted_rows) if row is india_row), None)
    if india_index is None:
        return []

    window_rows: list[tuple[int, dict[str, Any]]] = []
    for offset in range(-radius, radius + 1):
        row_index = india_index + offset
        if 0 <= row_index < len(sorted_rows):
            window_rows.append((offset, sorted_rows[row_index]))
    return window_rows


def build_nbhd_report() -> dict[str, Any]:
    if history_files_ready():
        history_path = resolve_history_path(HISTORY_NBHD_CSV, LEGACY_HISTORY_NBHD_CSV)
        dataframe = read_history_csv(history_path)
        if dataframe.empty:
            return empty_nbhd_report()

        weeks = sorted(dataframe["Week"].dropna().astype(str).unique().tolist(), key=week_sort_key)
        merged: dict[str, dict[str, Any]] = {}

        for week_label in weeks:
            week_frame = dataframe[dataframe["Week"].astype(str) == week_label].copy()
            rows = []
            for row in week_frame.to_dict(orient="records"):
                market = normalize_text(row.get("Market"))
                city = normalize_text(row.get("City"))
                head_end = normalize_text(row.get("Head-End"))
                channel = normalize_text(row.get("Channel"))
                if not market or not city or not head_end or not channel:
                    continue
                tv_ch_no = normalize_number(row.get("TV CH. No."))
                frequency = normalize_number(row.get("Frequency"))
                rows.append(
                    {
                        "type": normalize_text(row.get("Type")),
                        "market": market,
                        "city": city,
                        "head_end": head_end,
                        "channel": channel,
                        "genre": normalize_text(row.get("Genre")),
                        "frequency": frequency,
                        "tv_ch_no": tv_ch_no,
                        "order_token": tv_ch_no if tv_ch_no is not None else frequency,
                    }
                )

            grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
            for row in rows:
                grouped.setdefault((row["market"], row["city"], row["head_end"]), []).append(row)

            for (market, city, head_end), group_rows in grouped.items():
                for offset, nbhd_row in extract_nbhd_window(group_rows):
                    row_key = "||".join((market, city, head_end, str(offset)))
                    record = merged.setdefault(
                        row_key,
                        {
                            "row_key": row_key,
                            "market": market,
                            "city": city,
                            "head_end": head_end,
                            "position": offset,
                            "is_reference": offset == 0,
                            "channels": {},
                            "genres": {},
                            "frequencies": {},
                        },
                    )
                    record["channels"][week_label] = normalize_text(nbhd_row["channel"])
                    record["genres"][week_label] = normalize_text(nbhd_row["genre"])
                    record["frequencies"][week_label] = nbhd_row["frequency"]

        records = list(merged.values())
        for record in records:
            for week in weeks:
                record["channels"].setdefault(week, "")
                record["genres"].setdefault(week, "")
                record["frequencies"].setdefault(week, None)

        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "weeks": weeks,
            "records": sorted(
                records,
                key=lambda item: (
                    item["market"].lower(),
                    item["city"].lower(),
                    item["head_end"].lower(),
                    item.get("position", 0),
                ),
            ),
            "message": "",
            "source_directory": str(history_path),
        }

    week_files = get_nbhd_week_files()
    if not week_files:
        return empty_nbhd_report()

    weekly_data = [prepare_nbhd_week_rows(path) for path in week_files]
    weeks = [label for label, _ in weekly_data]
    merged: dict[str, dict[str, Any]] = {}

    for week_label, rows in weekly_data:
        grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        for row in rows:
            grouped.setdefault((row["market"], row["city"], row["head_end"]), []).append(row)

        for (market, city, head_end), group_rows in grouped.items():
            for offset, nbhd_row in extract_nbhd_window(group_rows):
                row_key = "||".join((market, city, head_end, str(offset)))
                record = merged.setdefault(
                    row_key,
                    {
                        "row_key": row_key,
                        "market": market,
                        "city": city,
                        "head_end": head_end,
                        "position": offset,
                        "is_reference": offset == 0,
                        "channels": {},
                        "genres": {},
                        "frequencies": {},
                    },
                )
                record["channels"][week_label] = normalize_text(nbhd_row["channel"])
                record["genres"][week_label] = normalize_text(nbhd_row["genre"])
                record["frequencies"][week_label] = nbhd_row["frequency"]

    records = list(merged.values())
    for record in records:
        for week in weeks:
            record["channels"].setdefault(week, "")
            record["genres"].setdefault(week, "")
            record["frequencies"].setdefault(week, None)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": weeks,
        "records": sorted(
            records,
            key=lambda item: (
                item["market"].lower(),
                item["city"].lower(),
                item["head_end"].lower(),
                item.get("position", 0),
            ),
        ),
        "message": "",
        "source_directory": str(get_nbhd_source_dir()),
    }


def load_nbhd_report(force: bool = False) -> dict[str, Any]:
    signature = get_history_signature() if history_files_ready() else get_signature(get_nbhd_week_files())
    if not force and NBHD_REPORT_CACHE["report"] is not None and NBHD_REPORT_CACHE["signature"] == signature:
        return NBHD_REPORT_CACHE["report"]

    report = build_nbhd_report()
    NBHD_REPORT_CACHE["signature"] = signature
    NBHD_REPORT_CACHE["report"] = report
    return report


def build_nbhd_benchmark_report() -> dict[str, Any]:
    if history_files_ready():
        history_path = resolve_history_path(HISTORY_NBHD_CSV, LEGACY_HISTORY_NBHD_CSV)
        dataframe = read_history_csv(history_path)
        if dataframe.empty:
            return {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "weeks": [],
                "records": [],
                "message": "NBHD history CSV is empty.",
                "source_directory": str(history_path),
            }

        weeks = sorted(dataframe["Week"].dropna().astype(str).unique().tolist(), key=week_sort_key)
        merged: dict[str, dict[str, Any]] = {}
        for row in dataframe.to_dict(orient="records"):
            week_label = normalize_text(row.get("Week"))
            market = normalize_text(row.get("Market"))
            city = normalize_text(row.get("City"))
            head_end = normalize_text(row.get("Head-End"))
            channel = normalize_text(row.get("Channel"))
            if not market or not city or not head_end or not channel or not week_label:
                continue
            row_key = comparison_record_key(market, city, head_end, channel)
            record = merged.setdefault(
                row_key,
                {
                    "market": market,
                    "city": city,
                    "head_end": head_end,
                    "channel": channel,
                    "frequencies": {},
                },
            )
            record["frequencies"][week_label] = normalize_number(row.get("Frequency"))

        records = list(merged.values())
        for record in records:
            for week in weeks:
                record["frequencies"].setdefault(week, None)

        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "weeks": weeks,
            "records": sorted(
                records,
                key=lambda item: (
                    item["market"].lower(),
                    item["city"].lower(),
                    item["head_end"].lower(),
                    item["channel"].lower(),
                ),
            ),
            "message": "",
            "source_directory": str(history_path),
        }

    week_files = get_nbhd_week_files()
    if not week_files:
        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "weeks": [],
            "records": [],
            "message": "Add weekly NBHD files to generate the INDIA TV comparison report.",
            "source_directory": str(get_nbhd_source_dir()),
        }

    weekly_data = [prepare_nbhd_week_rows(path) for path in week_files]
    weeks = [label for label, _ in weekly_data]
    merged: dict[str, dict[str, Any]] = {}

    for week_label, rows in weekly_data:
        for row in rows:
            market = normalize_text(row.get("market"))
            city = normalize_text(row.get("city"))
            head_end = normalize_text(row.get("head_end"))
            channel = normalize_text(row.get("channel"))
            if not market or not city or not head_end or not channel:
                continue
            row_key = comparison_record_key(market, city, head_end, channel)
            record = merged.setdefault(
                row_key,
                {
                    "market": market,
                    "city": city,
                    "head_end": head_end,
                    "channel": channel,
                    "frequencies": {},
                },
            )
            record["frequencies"][week_label] = row.get("frequency")

    records = list(merged.values())
    for record in records:
        for week in weeks:
            record["frequencies"].setdefault(week, None)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": weeks,
        "records": sorted(
            records,
            key=lambda item: (
                item["market"].lower(),
                item["city"].lower(),
                item["head_end"].lower(),
                item["channel"].lower(),
            ),
        ),
        "message": "",
        "source_directory": str(get_nbhd_source_dir()),
    }


def load_nbhd_benchmark_report(force: bool = False) -> dict[str, Any]:
    signature = get_history_signature() if history_files_ready() else get_signature(get_nbhd_week_files())
    if not force and NBHD_BENCHMARK_REPORT_CACHE["report"] is not None and NBHD_BENCHMARK_REPORT_CACHE["signature"] == signature:
        return NBHD_BENCHMARK_REPORT_CACHE["report"]

    report = build_nbhd_benchmark_report()
    NBHD_BENCHMARK_REPORT_CACHE["signature"] = signature
    NBHD_BENCHMARK_REPORT_CACHE["report"] = report
    return report


def build_ots_report() -> dict[str, Any]:
    if history_files_ready():
        history_path = resolve_history_path(HISTORY_OTS_CSV, LEGACY_HISTORY_OTS_CSV)
        dataframe = read_history_csv(history_path)
        if dataframe.empty:
            return empty_ots_report()

        weeks = sorted(dataframe["Week"].dropna().astype(str).unique().tolist(), key=week_sort_key)
        merged: dict[str, dict[str, Any]] = {}

        for week_label in weeks:
            week_frame = dataframe[dataframe["Week"].astype(str) == week_label]
            for row in week_frame.to_dict(orient="records"):
                market = normalize_text(row.get("Market"))
                channel = normalize_text(row.get("Channel"))
                if not market or not channel:
                    continue
                row_key = f"{market}||{channel}"
                record = merged.setdefault(
                    row_key,
                    {
                        "row_key": row_key,
                        "market": market,
                        "channel": channel,
                        "ots_values": {},
                    },
                )
                record["ots_values"][week_label] = normalize_ots_percentage(row.get("OTS"))

        records = list(merged.values())
        for record in records:
            for week in weeks:
                record["ots_values"].setdefault(week, None)

        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "weeks": weeks,
            "records": sorted(records, key=lambda item: (item["market"].lower(), item["channel"].lower())),
            "message": "",
            "source_directory": str(history_path),
        }

    # Merge all weekly OTS workbooks into one dynamic week matrix keyed by market + channel.
    week_files = get_ots_week_files()
    if not week_files:
        return empty_ots_report()

    weekly_data = [prepare_ots_week_rows(path) for path in week_files]
    weeks = [label for label, _ in weekly_data]
    merged: dict[str, dict[str, Any]] = {}

    for week_label, rows in weekly_data:
        for row in rows:
            row_key = f"{row['market']}||{row['channel']}"
            record = merged.setdefault(
                row_key,
                {
                    "row_key": row_key,
                    "market": row["market"],
                    "channel": row["channel"],
                    "ots_values": {},
                },
            )
            record["ots_values"][week_label] = row["ots"]

    records = list(merged.values())
    for record in records:
        for week in weeks:
            record["ots_values"].setdefault(week, None)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": weeks,
        "records": sorted(records, key=lambda item: (item["market"].lower(), item["channel"].lower())),
        "message": "",
        "source_directory": str(get_ots_source_dir()),
    }


def load_ots_report(force: bool = False) -> dict[str, Any]:
    # Reuse a signature cache so repeated dashboard refreshes do not re-parse unchanged OTS files.
    signature = get_history_signature() if history_files_ready() else get_signature(get_ots_week_files())
    if not force and OTS_REPORT_CACHE["report"] is not None and OTS_REPORT_CACHE["signature"] == signature:
        return OTS_REPORT_CACHE["report"]

    report = build_ots_report()
    OTS_REPORT_CACHE["signature"] = signature
    OTS_REPORT_CACHE["report"] = report
    return report


def empty_report(message: str | None = None) -> dict[str, Any]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": [],
        "records": [],
        "message": message or "Add weekly Excel files to the data folder and refresh the dashboard.",
    }


def prepare_week_rows(path: Path, fallback_label: str) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    week_label = fallback_label

    for row_index, values in enumerate(sheet.iter_rows(min_row=1, max_col=16, values_only=True), start=1):
        if row_index == 1:
            continue

        row = {
            SOURCE_COLUMNS[index]: values[index] if index < len(values) else None
            for index in range(len(SOURCE_COLUMNS))
        }

        if row_index == 2 and normalize_text(row["WEEK LABEL"]):
            week_label = normalize_text(row["WEEK LABEL"])

        normalized = {column: normalize_text(row.get(column)) for column in DISPLAY_COLUMNS if column != "NAME"}
        normalized["NAME"] = normalize_text(row.get("CHANNEL NAME"))
        normalized[FREQUENCY_COLUMN] = normalize_number(row.get(FREQUENCY_COLUMN))
        normalized[RANK_COLUMN] = normalize_rank(row.get(RANK_COLUMN))
        normalized["BAND"] = normalize_text(row.get("BAND"))
        normalized["TV CH. No."] = normalize_text(row.get("TV CH. No."))
        normalized["CRN No."] = normalize_text(row.get("CRN No."))
        normalized["ROW KEY"] = "||".join(normalized[column] for column in KEY_COLUMNS)

        if not normalized["ROW KEY"].replace("|", ""):
            continue

        rows.append(normalized)

    workbook.close()

    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        deduped.setdefault(row["ROW KEY"], row)
    return week_label, list(deduped.values())


def calculate_frequency_change(previous: float | int | None, current: float | int | None) -> str:
    if previous is None or current is None:
        return "missing"
    if current > previous:
        return "increase"
    if current < previous:
        return "decrease"
    return "no_change"


def calculate_rank_change(previous: int | None, current: int | None) -> str:
    if previous is None or current is None:
        return "missing"
    if current < previous:
        return "improve"
    if current > previous:
        return "decline"
    return "no_change"


def calculate_band_change(previous: str | None, current: str | None) -> str:
    previous_text = normalize_text(previous)
    current_text = normalize_text(current)
    if not previous_text or not current_text:
        return "missing"
    if current_text == previous_text:
        return "no_change"
    return "change"


def has_any_change(series: dict[str, Any], weeks: list[str]) -> bool:
    values = [series.get(week) for week in weeks if series.get(week) not in (None, "")]
    return len(values) > 1 and len(set(values)) > 1


def build_report() -> dict[str, Any]:
    if history_files_ready():
        dataframe = read_history_csv(resolve_history_path(HISTORY_DISTRIBUTION_CSV, LEGACY_HISTORY_DISTRIBUTION_CSV))
        if dataframe.empty:
            report = empty_report()
            report["message"] = "Distribution history CSV is empty."
            return report

        weeks = sorted(dataframe["Week"].dropna().astype(str).unique().tolist(), key=week_sort_key)
        merged: dict[str, dict[str, Any]] = {}
        for row in dataframe.to_dict(orient="records"):
            transmission = normalize_text(row.get("Transmission"))
            market = normalize_text(row.get("Market"))
            mso_type = normalize_text(row.get("MSO Type"))
            city = normalize_text(row.get("City"))
            head_end = normalize_text(row.get("Head-End"))
            channel_name = normalize_text(row.get("Channel Name"))
            crn_no = normalize_text(row.get("CRN No."))
            row_key = "||".join((transmission, market, mso_type, city, head_end, channel_name, crn_no))
            if not row_key.replace("|", ""):
                continue

            week_label = normalize_text(row.get("Week"))
            record = merged.setdefault(
                row_key,
                {
                    "row_key": row_key,
                    "transmission": transmission,
                    "market": market,
                    "mso_type": mso_type,
                    "mso": transmission,
                    "city": city,
                    "head_end": head_end,
                    "channel_name": channel_name,
                    "band": normalize_text(row.get("Band")),
                    "tv_ch_no": normalize_text(row.get("TV CH. No.")),
                    "crn_no": crn_no,
                    "genre": normalize_text(row.get("Genre")),
                    "language": normalize_text(row.get("Language")),
                    "name": channel_name,
                    "week_label": week_label,
                    "frequencies": {},
                    "ranks": {},
                    "bands": {},
                    "changes": {},
                    "rank_changes": {},
                    "band_changes": {},
                },
            )
            record["frequencies"][week_label] = normalize_number(row.get("Frequency/LCN No"))
            record["ranks"][week_label] = normalize_rank(row.get("Rank Within Genre"))
            record["bands"][week_label] = normalize_text(row.get("Band"))

        records = list(merged.values())
        for record in records:
            for week in weeks:
                record["frequencies"].setdefault(week, None)
                record["ranks"].setdefault(week, None)
                record["bands"].setdefault(week, "")

            if weeks:
                first_week = weeks[0]
                record["changes"][first_week] = "baseline"
                record["rank_changes"][first_week] = "baseline"
                record["band_changes"][first_week] = "baseline"

            for index in range(1, len(weeks)):
                previous_week = weeks[index - 1]
                current_week = weeks[index]
                record["changes"][current_week] = calculate_frequency_change(
                    record["frequencies"][previous_week],
                    record["frequencies"][current_week],
                )
                record["rank_changes"][current_week] = calculate_rank_change(
                    record["ranks"][previous_week],
                    record["ranks"][current_week],
                )
                record["band_changes"][current_week] = calculate_band_change(
                    record["bands"][previous_week],
                    record["bands"][current_week],
                )

            record["change_status"] = "YES" if has_any_change(record["frequencies"], weeks) else "NO"
            record["rank_change_status"] = "YES" if has_any_change(record["ranks"], weeks) else "NO"
            record["band_change_status"] = "YES" if has_any_change(record["bands"], weeks) else "NO"

        report = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "weeks": weeks,
            "records": records,
            "message": "",
        }
        return report

    week_files = get_week_files()
    if not week_files:
        report = empty_report()
        report["message"] = "Add weekly Excel files to the data folder to generate the report."
        write_standalone_dashboard(report)
        return report

    weekly_data = [prepare_week_rows(path, path.stem) for path in week_files]
    weeks = [label for label, _ in weekly_data]

    merged: dict[str, dict[str, Any]] = {}
    for week_label, rows in weekly_data:
        for row in rows:
            record = merged.setdefault(
                row["ROW KEY"],
                {
                    "row_key": row["ROW KEY"],
                    "transmission": row["TRANSMISSION"],
                    "market": row["MARKET"],
                    "mso_type": row["MSO TYPE"],
                    "mso": row["TRANSMISSION"],
                    "city": row["CITY"],
                    "head_end": row["HEAD-END"],
                    "channel_name": row["CHANNEL NAME"],
                    "band": row["BAND"],
                    "tv_ch_no": row["TV CH. No."],
                    "crn_no": row["CRN No."],
                    "genre": row["GENRE"],
                    "language": row["LANGUAGE"],
                    "name": row["NAME"],
                    "week_label": row["WEEK LABEL"],
                    "frequencies": {},
                    "ranks": {},
                    "bands": {},
                    "changes": {},
                    "rank_changes": {},
                    "band_changes": {},
                },
            )
            record["frequencies"][week_label] = row[FREQUENCY_COLUMN]
            record["ranks"][week_label] = row[RANK_COLUMN]
            record["bands"][week_label] = row["BAND"]

    records = list(merged.values())

    for record in records:
        for week in weeks:
            record["frequencies"].setdefault(week, None)
            record["ranks"].setdefault(week, None)
            record["bands"].setdefault(week, "")

        if weeks:
            first_week = weeks[0]
            record["changes"][first_week] = "baseline"
            record["rank_changes"][first_week] = "baseline"
            record["band_changes"][first_week] = "baseline"

        for index in range(1, len(weeks)):
            previous_week = weeks[index - 1]
            current_week = weeks[index]
            record["changes"][current_week] = calculate_frequency_change(
                record["frequencies"][previous_week],
                record["frequencies"][current_week],
            )
            record["rank_changes"][current_week] = calculate_rank_change(
                record["ranks"][previous_week],
                record["ranks"][current_week],
            )
            record["band_changes"][current_week] = calculate_band_change(
                record["bands"][previous_week],
                record["bands"][current_week],
            )

        record["change_status"] = "YES" if has_any_change(record["frequencies"], weeks) else "NO"
        record["rank_change_status"] = "YES" if has_any_change(record["ranks"], weeks) else "NO"
        record["band_change_status"] = "YES" if has_any_change(record["bands"], weeks) else "NO"

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": weeks,
        "records": records,
        "message": "",
    }
    OUTPUT_JSON.write_text(compact_json_text(report), encoding="utf-8")
    write_standalone_dashboard(report)
    return report


def load_report(force: bool = False) -> dict[str, Any]:
    signature = get_history_signature() if history_files_ready() else get_signature(get_week_files())
    if not force and REPORT_CACHE["report"] is not None and REPORT_CACHE["signature"] == signature:
        return REPORT_CACHE["report"]

    report = build_report()
    REPORT_CACHE["signature"] = signature
    REPORT_CACHE["report"] = report
    return report


def normalize_frequency_cell(value: Any) -> float | int | None:
    text = normalize_text(value).upper()
    if text in {"", "NA", "N/A", "NOT AVAILABLE", "-"}:
        return None
    return normalize_number(value)


def comparison_record_key(market: str, city: str, head_end: str, channel: str) -> str:
    return "||".join(
        (
            normalize_text(market).upper(),
            normalize_text(city).upper(),
            normalize_text(head_end).upper(),
            normalize_text(channel).upper(),
        )
    )


def parse_weekly_highlight_sheet(path: Path) -> tuple[str, str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = "Weekly Highlights" if "Weekly Highlights" in workbook.sheetnames else "Weekly Highlight"
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{path.name} does not contain a Weekly Highlights sheet.")

        sheet = workbook[sheet_name]
        header_row_index = None
        headers: list[Any] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            row_values = list(row)
            header_keys = [normalize_header_key(value) for value in row_values]
            if {"MARKET", "CITY", "HEADEND", "CHANNEL"}.issubset(set(header_keys)):
                if sum(1 for value in row_values if "FREQUENCY" in normalize_text(value).upper()) >= 2:
                    header_row_index = row_index
                    headers = row_values
                    break

        if header_row_index is None:
            raise ValueError(f"Could not detect Weekly Highlights headers in {path.name}.")

        header_map = {normalize_header_key(value): index for index, value in enumerate(headers)}
        week_columns: list[tuple[int, int]] = []
        workbook_week = parse_workbook_week(path)
        workbook_year = workbook_week[0] if workbook_week else datetime.now().year

        for index, header in enumerate(headers):
            match = re.search(r"FREQUENCY\s*WK[-\s_]*(\d{1,2})", normalize_text(header), re.IGNORECASE)
            if match:
                week_columns.append((index, int(match.group(1))))

        if len(week_columns) < 2:
            raise ValueError(f"Could not detect two frequency week columns in Weekly Highlights for {path.name}.")

        week_columns = sorted(week_columns[:2], key=lambda item: item[1])
        previous_week_label = format_week_label(week_columns[0][1], workbook_year)
        current_week_label = format_week_label(week_columns[1][1], workbook_year)

        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            market = normalize_text(values[header_map["MARKET"]] if header_map["MARKET"] < len(values) else None)
            city = normalize_text(values[header_map["CITY"]] if header_map["CITY"] < len(values) else None)
            head_end = normalize_text(values[header_map["HEADEND"]] if header_map["HEADEND"] < len(values) else None)
            channel = normalize_text(values[header_map["CHANNEL"]] if header_map["CHANNEL"] < len(values) else None)
            if not market or not city or not head_end or not channel:
                continue
            rows.append(
                {
                    "market": market,
                    "city": city,
                    "head_end": head_end,
                    "channel": channel,
                    "frequency_previous": normalize_frequency_cell(values[week_columns[0][0]] if week_columns[0][0] < len(values) else None),
                    "frequency_current": normalize_frequency_cell(values[week_columns[1][0]] if week_columns[1][0] < len(values) else None),
                }
            )

        deduped: dict[str, dict[str, Any]] = {}
        for row in rows:
            deduped.setdefault(
                comparison_record_key(row["market"], row["city"], row["head_end"], row["channel"]),
                row,
            )
        return previous_week_label, current_week_label, list(deduped.values())
    finally:
        workbook.close()


def build_comparison_report(report: dict[str, Any] | None = None) -> dict[str, Any]:
    frequency_report = report if report is not None else load_report(force=True)
    week_files = get_highlight_week_files()
    if not week_files:
        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "weeks": frequency_report.get("weeks", []),
            "pairs": [],
            "rows_by_pair": {},
            "message": "Add weekly workbook files to the data folder to generate the comparison table.",
        }

    rank_index = {
        comparison_record_key(record["market"], record["city"], record["head_end"], record["channel_name"]): record
        for record in frequency_report.get("records", [])
    }

    pairs: list[dict[str, Any]] = []
    rows_by_pair: dict[str, list[dict[str, Any]]] = {}
    available_weeks = set(frequency_report.get("weeks", []))

    for path in week_files:
        try:
            previous_week, current_week, highlight_rows = parse_weekly_highlight_sheet(path)
        except Exception:
            continue
        if previous_week not in available_weeks or current_week not in available_weeks:
            continue

        pair_key = f"{previous_week}||{current_week}"
        merged_rows: list[dict[str, Any]] = []
        for row in highlight_rows:
            record = rank_index.get(comparison_record_key(row["market"], row["city"], row["head_end"], row["channel"]))
            if not record:
                continue
            merged_rows.append(
                {
                    "market": row["market"],
                    "city": row["city"],
                    "head_end": row["head_end"],
                    "channel": row["channel"],
                    "frequency_previous": row["frequency_previous"],
                    "frequency_current": row["frequency_current"],
                    "rank_previous": record["ranks"].get(previous_week),
                    "rank_current": record["ranks"].get(current_week),
                }
            )

        merged_rows.sort(
            key=lambda item: (
                item["market"].lower(),
                item["city"].lower(),
                item["head_end"].lower(),
                item["channel"].lower(),
            )
        )
        rows_by_pair[pair_key] = merged_rows
        pairs.append(
            {
                "pair_key": pair_key,
                "week_from": previous_week,
                "week_to": current_week,
                "row_count": len(merged_rows),
            }
        )

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": frequency_report.get("weeks", []),
        "pairs": sorted(pairs, key=lambda item: (week_sort_key(item["week_from"]), week_sort_key(item["week_to"]))),
        "rows_by_pair": rows_by_pair,
        "message": "",
    }


def load_comparison_report(force: bool = False, report: dict[str, Any] | None = None) -> dict[str, Any]:
    source_files = get_highlight_week_files()
    signature = get_signature(source_files) + tuple((f"report::{week}", index, index) for index, week in enumerate((report or load_report(force=False)).get("weeks", [])))
    if not force and COMPARISON_REPORT_CACHE["report"] is not None and COMPARISON_REPORT_CACHE["signature"] == signature:
        return COMPARISON_REPORT_CACHE["report"]

    comparison = build_comparison_report(report=report)
    COMPARISON_REPORT_CACHE["signature"] = signature
    COMPARISON_REPORT_CACHE["report"] = comparison
    return comparison


def get_view_config(view: str) -> dict[str, str]:
    if view == "rank":
        return {
            "series": "ranks",
            "changes": "rank_changes",
            "status": "rank_change_status",
            "positive": "improve",
            "negative": "decline",
            "positive_label": "improved",
            "negative_label": "declined",
        }
    if view == "band":
        return {
            "series": "bands",
            "changes": "band_changes",
            "status": "band_change_status",
            "positive": "change",
            "negative": "no_change",
            "positive_label": "changed",
            "negative_label": "stable",
        }
    return {
        "series": "frequencies",
        "changes": "changes",
        "status": "change_status",
        "positive": "increase",
        "negative": "decrease",
        "positive_label": "increased",
        "negative_label": "decreased",
    }


def filter_records(records: list[dict[str, Any]], view: str, filters: dict[str, str], ignore_key: str = "") -> list[dict[str, Any]]:
    config = get_view_config(view)
    changed_states = {"frequency": {"increase", "decrease"}, "rank": {"improve", "decline"}, "band": {"change"}}
    changed_set = changed_states.get(view, changed_states["frequency"])

    filtered: list[dict[str, Any]] = []
    for record in records:
        if filters["market"] and record["market"] != filters["market"] and ignore_key != "market":
            continue
        if filters["city"] and record["city"] != filters["city"] and ignore_key != "city":
            continue
        if filters["mso_type"] and record["mso_type"] != filters["mso_type"] and ignore_key != "mso_type":
            continue
        if filters["head_end"] and record["head_end"] != filters["head_end"] and ignore_key != "head_end":
            continue
        if filters["crn_no"] and record["crn_no"] != filters["crn_no"] and ignore_key != "crn_no":
            continue
        if filters["channel_name"] and record["channel_name"] != filters["channel_name"] and ignore_key != "channel_name":
            continue
        if filters["band"] and record["band"] != filters["band"] and ignore_key != "band":
            continue

        if ignore_key != "week" and filters["week"]:
            value = record[config["series"]].get(filters["week"])
            if value in (None, ""):
                continue

        if ignore_key != "change" and filters["change"]:
            if filters["change"] == "Changed":
                if filters["week"]:
                    if record[config["changes"]].get(filters["week"]) not in changed_set:
                        continue
                elif record[config["status"]] != "YES":
                    continue
            elif filters["change"] == "No Change":
                if filters["week"]:
                    if record[config["changes"]].get(filters["week"]) != "no_change":
                        continue
                elif record[config["status"]] != "NO":
                    continue

        filtered.append(record)

    return filtered


def sort_value(record: dict[str, Any], sort_key: str, view: str) -> Any:
    if sort_key == "flow_order":
        return (
            0,
            (
                str(record.get("market") or "").lower(),
                str(record.get("city") or "").lower(),
                str(record.get("head_end") or "").lower(),
                str(record.get("channel_name") or "").lower(),
            ),
        )
    if sort_key in record:
        value = record[sort_key]
    elif sort_key in record.get("frequencies", {}):
        config = get_view_config(view)
        value = record[config["series"]].get(sort_key)
    else:
        value = ""

    if value is None:
        return (1, "")
    if isinstance(value, (int, float)):
        return (0, value)
    return (0, str(value).lower())


def sort_records(records: list[dict[str, Any]], sort_key: str, sort_direction: str, view: str) -> list[dict[str, Any]]:
    reverse = sort_direction == "desc"
    return sorted(records, key=lambda record: sort_value(record, sort_key, view), reverse=reverse)


def paginate_records(records: list[dict[str, Any]], page: int, page_size: int) -> tuple[list[dict[str, Any]], int]:
    total_count = len(records)
    if total_count == 0:
        return [], 0
    start = max(page - 1, 0) * page_size
    end = start + page_size
    return records[start:end], total_count


def build_filters(records: list[dict[str, Any]], view: str, current_filters: dict[str, str], weeks: list[str]) -> dict[str, list[str]]:
    def values_for(key: str, field: str) -> list[str]:
        values = {
            normalize_text(record.get(field))
            for record in filter_records(records, view, current_filters, ignore_key=key)
            if normalize_text(record.get(field))
        }
        return sorted(values, key=lambda value: value.lower())

    return {
        "markets": values_for("market", "market"),
        "cities": values_for("city", "city"),
        "mso_types": values_for("mso_type", "mso_type"),
        "head_ends": values_for("head_end", "head_end"),
        "crn_numbers": values_for("crn_no", "crn_no"),
        "channels": values_for("channel_name", "channel_name"),
        "bands": values_for("band", "band"),
        "weeks": weeks,
        "change_options": ["Changed", "No Change"],
    }


def summarize_records(records: list[dict[str, Any]], view: str, weeks: list[str]) -> dict[str, int]:
    config = get_view_config(view)
    summary = {
        "total_channels": len(records),
    }

    positive = 0
    negative = 0
    stable = 0

    if view == "band":
        for record in records:
            if record[config["status"]] == "YES":
                positive += 1
            else:
                stable += 1
        summary["changed"] = positive
        summary["stable"] = stable
        return summary

    final_week = weeks[-1] if weeks else ""
    for record in records:
        status = record[config["changes"]].get(final_week, "no_change")
        if status == config["positive"]:
            positive += 1
        elif status == config["negative"]:
            negative += 1
        else:
            stable += 1

    if view == "rank":
        summary["improved"] = positive
        summary["declined"] = negative
        summary["no_change"] = stable
    else:
        summary["increased"] = positive
        summary["decreased"] = negative
        summary["no_change"] = stable
    return summary


def summarize_focus_channels(records: list[dict[str, Any]], view: str, weeks: list[str]) -> list[dict[str, Any]]:
    config = get_view_config(view)
    items: list[dict[str, Any]] = []

    for match_name, label in FOCUS_CHANNELS.items():
        selected = [record for record in records if normalize_text(record["channel_name"]).upper() == match_name]
        if not selected:
            continue

        positive = 0
        negative = 0
        no_change = 0
        latest_positive = 0
        latest_negative = 0
        latest_week = weeks[-1] if weeks else ""

        for record in selected:
            if not weeks:
                continue
            latest_status = record[config["changes"]].get(latest_week)
            if latest_status == config["positive"]:
                latest_positive += 1
            elif latest_status == config["negative"]:
                latest_negative += 1

            for week in weeks[1:]:
                status = record[config["changes"]].get(week)
                if status == config["positive"]:
                    positive += 1
                elif status == config["negative"]:
                    negative += 1
                elif status == "no_change":
                    no_change += 1

        items.append(
            {
                "label": label,
                "records": len(selected),
                "positive": positive,
                "negative": negative,
                "no_change": no_change,
                "latest_positive": latest_positive,
                "latest_negative": latest_negative,
                "latest_week": latest_week,
                "positive_label": config["positive_label"],
                "negative_label": config["negative_label"],
            }
        )

    return items


def serialize_records(records: list[dict[str, Any]], weeks: list[str]) -> list[dict[str, Any]]:
    serialized: list[dict[str, Any]] = []
    for record in records:
        serialized.append(
            {
                "market": record["market"],
                "mso_type": record["mso_type"],
                "mso": record["mso"],
                "city": record["city"],
                "head_end": record["head_end"],
                "channel_name": record["channel_name"],
                "band": record["band"],
                "tv_ch_no": record["tv_ch_no"],
                "crn_no": record["crn_no"],
                "name": record["name"],
                "frequencies": {week: record["frequencies"].get(week) for week in weeks},
                "ranks": {week: record["ranks"].get(week) for week in weeks},
                "bands": {week: record["bands"].get(week) for week in weeks},
                "changes": {week: record["changes"].get(week, "missing") for week in weeks},
                "rank_changes": {week: record["rank_changes"].get(week, "missing") for week in weeks},
                "band_changes": {week: record["band_changes"].get(week, "missing") for week in weeks},
                "change_status": record["change_status"],
                "rank_change_status": record["rank_change_status"],
                "band_change_status": record["band_change_status"],
            }
        )
    return serialized


def read_style() -> str:
    if STYLE_FILE.exists():
        return STYLE_FILE.read_text(encoding="utf-8")
    return "body { font-family: sans-serif; }"


def read_nbhd_script() -> str:
    if NBHD_SCRIPT_FILE.exists():
        return NBHD_SCRIPT_FILE.read_text(encoding="utf-8")
    return ""


def read_ots_script() -> str:
    if OTS_SCRIPT_FILE.exists():
        return OTS_SCRIPT_FILE.read_text(encoding="utf-8")
    return ""


def read_comparison_script() -> str:
    if COMPARISON_SCRIPT_FILE.exists():
        return COMPARISON_SCRIPT_FILE.read_text(encoding="utf-8")
    return ""


def read_nbhd_benchmark_script() -> str:
    if NBHD_BENCHMARK_SCRIPT_FILE.exists():
        return NBHD_BENCHMARK_SCRIPT_FILE.read_text(encoding="utf-8")
    return ""


def compact_json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def compact_inline_json(value: Any) -> str:
    return compact_json_text(value).replace("</", "<\\/")


def build_dashboard_bundle(report: dict[str, Any] | None = None) -> dict[str, Any]:
    frequency_report = report if report is not None else load_report(force=True)
    return {
        "frequency": frequency_report,
        "comparison": load_comparison_report(force=True, report=frequency_report),
        "nbhd_benchmark": load_nbhd_benchmark_report(force=True),
        "nbhd": build_nbhd_api_payload({"market": "", "city": "", "head_end": ""}, "", force_refresh=True),
        "ots": build_ots_api_payload(
            {"markets": [], "channels": [], "week_from": "", "week_to": "", "change": "", "search": ""},
            force_refresh=True,
        ),
    }


def write_frequency_report_json(report: dict[str, Any] | None = None) -> Path:
    bundle = build_dashboard_bundle(report)
    OUTPUT_JSON.write_text(f"window.__CHROME_REPORT_DATA__ = {compact_inline_json(bundle)};", encoding="utf-8")
    return OUTPUT_JSON


def write_standalone_dashboard(report: dict[str, Any]) -> None:
    OUTPUT_HTML.write_text(create_standalone_dashboard(report), encoding="utf-8")


def create_standalone_dashboard(report: dict[str, Any]) -> str:
    style_text = read_style()
    nbhd_benchmark_script_text = read_nbhd_benchmark_script()
    nbhd_script_text = read_nbhd_script()
    ots_script_text = read_ots_script()
    comparison_script_text = read_comparison_script()

    html = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Chrome Report</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  <style>
__STYLE__
  </style>
</head>
<body>
  <div class="app-shell">
    <div class="table1-scope">
      <section class="panel filter-panel">
        <div class="panel-heading"><div><h2>Filter Panel</h2></div></div>
        <div class="filter-grid">
          <label class="filter-select-field"><span>Market</span><div class="filter-select"><button id="marketFilter" class="filter-select-button" type="button">All Markets</button><div id="marketFilterMenu" class="filter-select-menu" hidden><input id="marketFilterSearch" class="filter-menu-search" type="text" placeholder="Search market..." autocomplete="off" /><div id="marketFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>City</span><div class="filter-select"><button id="cityFilter" class="filter-select-button" type="button">All Cities</button><div id="cityFilterMenu" class="filter-select-menu" hidden><input id="cityFilterSearch" class="filter-menu-search" type="text" placeholder="Search city..." autocomplete="off" /><div id="cityFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>MSO Type</span><div class="filter-select"><button id="msoTypeFilter" class="filter-select-button" type="button">All MSO Types</button><div id="msoTypeFilterMenu" class="filter-select-menu" hidden><input id="msoTypeFilterSearch" class="filter-menu-search" type="text" placeholder="Search MSO type..." autocomplete="off" /><div id="msoTypeFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>Headend</span><div class="filter-select"><button id="headendFilter" class="filter-select-button" type="button">All Headend</button><div id="headendFilterMenu" class="filter-select-menu" hidden><input id="headendFilterSearch" class="filter-menu-search" type="text" placeholder="Search headend..." autocomplete="off" /><div id="headendFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>CRN No</span><div class="filter-select"><button id="crnFilter" class="filter-select-button" type="button">All CRN No</button><div id="crnFilterMenu" class="filter-select-menu" hidden><input id="crnFilterSearch" class="filter-menu-search" type="text" placeholder="Search CRN..." autocomplete="off" /><div id="crnFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>Channel</span><div class="filter-select"><button id="channelFilter" class="filter-select-button" type="button">All Channels</button><div id="channelFilterMenu" class="filter-select-menu" hidden><input id="channelFilterSearch" class="filter-menu-search" type="text" placeholder="Search channel..." autocomplete="off" /><div id="channelFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>Band</span><div class="filter-select"><button id="bandFilter" class="filter-select-button" type="button">All Bands</button><div id="bandFilterMenu" class="filter-select-menu" hidden><input id="bandFilterSearch" class="filter-menu-search" type="text" placeholder="Search band..." autocomplete="off" /><div id="bandFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>From Week</span><div class="filter-select"><button id="weekFromFilter" class="filter-select-button" type="button">From Week</button><div id="weekFromFilterMenu" class="filter-select-menu" hidden><input id="weekFromFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="weekFromFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>To Week</span><div class="filter-select"><button id="weekToFilter" class="filter-select-button" type="button">To Week</button><div id="weekToFilterMenu" class="filter-select-menu" hidden><input id="weekToFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="weekToFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>Change</span><div class="filter-select"><button id="changeFilter" class="filter-select-button" type="button">All Changes</button><div id="changeFilterMenu" class="filter-select-menu" hidden><input id="changeFilterSearch" class="filter-menu-search" type="text" placeholder="Search change..." autocomplete="off" /><div id="changeFilterOptions" class="filter-options-list"></div></div></div></label>
          <div class="action-row table1-filter-actions">
            <button id="resetButton" class="ghost-button" type="button">Reset Filters</button>
            <button id="fullscreenButton" class="primary-button" type="button">Full Screen</button>
          </div>
        </div>
      </section>

      <section class="panel table-panel">
        <div class="panel-heading table-heading">
          <div><h2 id="tableTitle">Weekly Frequency Analysis</h2></div>
          <div class="table-side">
            <div class="view-switcher">
              <button id="frequencyViewButton" class="switch-button active" type="button">Frequency</button>
              <button id="rankViewButton" class="switch-button" type="button">Rank</button>
              <button id="bandViewButton" class="switch-button" type="button">Band</button>
            </div>
            <div class="table-meta">
              <span id="resultCount">0 records</span>
              <span id="pageInfo">Page 1</span>
            </div>
          </div>
        </div>
        <div class="table-wrap">
          <table>
            <thead id="tableHead"></thead>
            <tbody id="tableBody"></tbody>
          </table>
        </div>
        <div class="pagination-bar">
          <button id="prevPage" class="ghost-button" type="button">Previous</button>
          <button id="exitFullscreenButton" class="ghost-button table-exit-fullscreen" type="button" hidden>Exit Full Screen</button>
          <button id="channelReportToggleButton" class="primary-button" type="button">Show Report</button>
          <button id="nextPage" class="ghost-button" type="button">Next</button>
        </div>
      </section>

      <section class="panel channel-report-panel">
        <div class="panel-heading">
          <div><h2>Channel Report</h2></div>
          <div class="table-meta">
            <span id="channelReportCount">0 channels</span>
          </div>
        </div>
        <div class="channel-report-toolbar">
          <label><span>Report Channel</span><select id="channelReportChannelFilter"></select></label>
          <label><span>Week From</span><select id="channelReportWeekFromFilter"></select></label>
          <label><span>Week To</span><select id="channelReportWeekToFilter"></select></label>
          <div class="action-row channel-report-actions">
            <button id="channelReportResetButton" class="ghost-button" type="button">Reset</button>
            <button id="channelReportHideButton" class="ghost-button" type="button">Hide</button>
          </div>
        </div>
        <div id="channelReportContainer" class="channel-report-stack"></div>
      </section>

    </div>

    <section class="panel nbhd-panel">
      <div class="panel-heading nbhd-heading">
        <div><h2>Neighbourhood Comparison</h2></div>
        <div class="table-meta">
          <span id="nbhdResultCount">0 rows</span>
        </div>
      </div>
        <div class="nbhd-toolbar">
          <label class="filter-select-field"><span>Market</span><div class="filter-select"><button id="nbhdMarketFilter" class="filter-select-button" type="button">All Markets</button><div id="nbhdMarketFilterMenu" class="filter-select-menu" hidden><input id="nbhdMarketFilterSearch" class="filter-menu-search" type="text" placeholder="Search market..." autocomplete="off" /><div id="nbhdMarketFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>City</span><div class="filter-select"><button id="nbhdCityFilter" class="filter-select-button" type="button">All Cities</button><div id="nbhdCityFilterMenu" class="filter-select-menu" hidden><input id="nbhdCityFilterSearch" class="filter-menu-search" type="text" placeholder="Search city..." autocomplete="off" /><div id="nbhdCityFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>Headend</span><div class="filter-select"><button id="nbhdHeadendFilter" class="filter-select-button" type="button">All Headends</button><div id="nbhdHeadendFilterMenu" class="filter-select-menu" hidden><input id="nbhdHeadendFilterSearch" class="filter-menu-search" type="text" placeholder="Search headend..." autocomplete="off" /><div id="nbhdHeadendFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>From Week</span><div class="filter-select"><button id="nbhdWeekFromFilter" class="filter-select-button" type="button">From Week</button><div id="nbhdWeekFromFilterMenu" class="filter-select-menu" hidden><input id="nbhdWeekFromFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="nbhdWeekFromFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>To Week</span><div class="filter-select"><button id="nbhdWeekToFilter" class="filter-select-button" type="button">To Week</button><div id="nbhdWeekToFilterMenu" class="filter-select-menu" hidden><input id="nbhdWeekToFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="nbhdWeekToFilterOptions" class="filter-options-list"></div></div></div></label>
          <label class="filter-select-field"><span>Change</span><div class="filter-select"><button id="nbhdChangeFilter" class="filter-select-button" type="button">All Changes</button><div id="nbhdChangeFilterMenu" class="filter-select-menu" hidden><input id="nbhdChangeFilterSearch" class="filter-menu-search" type="text" placeholder="Search change..." autocomplete="off" /><div id="nbhdChangeFilterOptions" class="filter-options-list"></div></div></div></label>
          <div class="action-row nbhd-actions">
            <button id="nbhdResetButton" class="ghost-button" type="button">Reset Filters</button>
            <button id="nbhdFullscreenButton" class="primary-button" type="button">Full Screen</button>
            <button id="nbhdRefreshButton" class="ghost-button" type="button">Refresh</button>
          </div>
        </div>
        <div id="nbhdStatusMessage" class="status-message" hidden></div>
        <div class="nbhd-table-wrap">
          <table id="nbhdTable" class="nbhd-table">
            <thead id="nbhdTableHead"></thead>
            <tbody id="nbhdTableBody"></tbody>
        </table>
      </div>
      <div class="pagination-bar nbhd-pagination-bar">
        <button id="nbhdPrevPage" class="ghost-button" type="button">Previous</button>
        <span id="nbhdPageInfo">Page 1 of 1</span>
          <button id="nbhdNextPage" class="ghost-button" type="button">Next</button>
          <button id="nbhdExitFullscreenButton" class="ghost-button nbhd-exit-fullscreen" type="button" hidden>Exit Full Screen</button>
        </div>
        <div id="nbhdReportLauncher" class="nbhd-report-launcher">
          <button id="nbhdReportToggleButton" class="primary-button" type="button">Show Report</button>
        </div>
        <section id="nbhdReportPanel" class="panel nbhd-report-panel" hidden>
          <div class="panel-heading nbhd-report-heading">
            <div>
              <h3>NBHD Position Change Report</h3>
              <p id="nbhdReportMeta" class="panel-subtitle">Compare the selected channels with their previous and current neighbourhood positions.</p>
            </div>
            <div class="table-meta">
              <span id="nbhdReportCount">0 narratives</span>
            </div>
          </div>
          <div class="nbhd-report-toolbar">
            <label class="ots-multiselect-field">
              <span>Headend</span>
              <div class="ots-multiselect">
                <button id="nbhdReportHeadendFilter" class="ots-select-button" type="button">All Headends</button>
                <div id="nbhdReportHeadendFilterMenu" class="ots-multiselect-menu" hidden>
                  <input id="nbhdReportHeadendFilterSearch" class="filter-menu-search" type="text" placeholder="Search headend..." autocomplete="off" />
                  <div id="nbhdReportHeadendFilterOptions" class="ots-options-list"></div>
                </div>
              </div>
            </label>
            <label class="ots-multiselect-field">
              <span>Channel</span>
              <div class="ots-multiselect">
                <button id="nbhdReportChannelFilter" class="ots-select-button" type="button">Default 4 Channels</button>
                <div id="nbhdReportChannelFilterMenu" class="ots-multiselect-menu" hidden>
                  <input id="nbhdReportChannelFilterSearch" class="filter-menu-search" type="text" placeholder="Search channel..." autocomplete="off" />
                  <div id="nbhdReportChannelFilterOptions" class="ots-options-list"></div>
                </div>
              </div>
            </label>
            <label class="filter-select-field">
              <span>Previous Week</span>
              <div class="filter-select">
                <button id="nbhdReportWeekFromFilter" class="filter-select-button" type="button">Previous Week</button>
                <div id="nbhdReportWeekFromFilterMenu" class="filter-select-menu" hidden>
                  <input id="nbhdReportWeekFromFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" />
                  <div id="nbhdReportWeekFromFilterOptions" class="filter-options-list"></div>
                </div>
              </div>
            </label>
            <label class="filter-select-field">
              <span>Current Week</span>
              <div class="filter-select">
                <button id="nbhdReportWeekToFilter" class="filter-select-button" type="button">Current Week</button>
                <div id="nbhdReportWeekToFilterMenu" class="filter-select-menu" hidden>
                  <input id="nbhdReportWeekToFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" />
                  <div id="nbhdReportWeekToFilterOptions" class="filter-options-list"></div>
                </div>
              </div>
            </label>
            <div class="action-row nbhd-report-actions">
              <button id="nbhdReportResetButton" class="ghost-button" type="button">Reset</button>
              <button id="nbhdReportDownloadButton" class="ghost-button" type="button">Download</button>
              <button id="nbhdReportPrintButton" class="ghost-button" type="button">Print</button>
              <button id="nbhdReportHideButton" class="primary-button" type="button">Hide</button>
            </div>
          </div>
          <div id="nbhdReportStatusMessage" class="status-message" hidden></div>
          <div id="nbhdReportContent" class="nbhd-report-stack"></div>
        </section>
      </section>

      <section class="panel ots-panel">
        <div class="panel-heading ots-heading">
          <div><h2>OTS Comparison</h2></div>
        <div class="table-meta">
          <span id="otsResultCount">0 records</span>
        </div>
      </div>
      <div class="ots-toolbar">
        <label class="ots-multiselect-field">
          <span>Market</span>
          <div class="ots-multiselect">
            <button id="otsMarketButton" class="ots-select-button" type="button">All Markets</button>
            <div id="otsMarketMenu" class="ots-multiselect-menu" hidden><input id="otsMarketSearch" class="ots-menu-search" type="text" placeholder="Search market..." autocomplete="off" /><div id="otsMarketOptions" class="ots-options-list"></div></div>
          </div>
        </label>
        <label class="ots-multiselect-field">
          <span>Channel</span>
          <div class="ots-multiselect">
            <button id="otsChannelButton" class="ots-select-button" type="button">All Channels</button>
            <div id="otsChannelMenu" class="ots-multiselect-menu" hidden><input id="otsChannelSearch" class="ots-menu-search" type="text" placeholder="Search channel..." autocomplete="off" /><div id="otsChannelOptions" class="ots-options-list"></div></div>
          </div>
        </label>
        <label class="filter-select-field"><span>Week From</span><div class="filter-select"><button id="otsWeekFromFilter" class="filter-select-button" type="button">From Week</button><div id="otsWeekFromFilterMenu" class="filter-select-menu" hidden><input id="otsWeekFromFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="otsWeekFromFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>Week To</span><div class="filter-select"><button id="otsWeekToFilter" class="filter-select-button" type="button">To Week</button><div id="otsWeekToFilterMenu" class="filter-select-menu" hidden><input id="otsWeekToFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="otsWeekToFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>Change</span><div class="filter-select"><button id="otsChangeFilter" class="filter-select-button" type="button">All Changes</button><div id="otsChangeFilterMenu" class="filter-select-menu" hidden><input id="otsChangeFilterSearch" class="filter-menu-search" type="text" placeholder="Search change..." autocomplete="off" /><div id="otsChangeFilterOptions" class="filter-options-list"></div></div></div></label>
        <div class="action-row ots-actions">
          <button id="otsResetButton" class="ghost-button" type="button">Reset Filters</button>
          <button id="otsRefreshButton" class="ghost-button" type="button">Refresh</button>
          <button id="otsFullscreenButton" class="primary-button" type="button">Full Screen</button>
        </div>
      </div>
      <div id="otsStatusMessage" class="status-message" hidden></div>
      <div class="ots-table-wrap">
        <table id="otsTable" class="ots-table">
          <thead id="otsTableHead"></thead>
          <tbody id="otsTableBody"></tbody>
        </table>
      </div>
      <div class="pagination-bar ots-pagination-bar">
        <button id="otsPrevPage" class="ghost-button" type="button">Previous</button>
        <span id="otsPageInfo">Page 1 of 1</span>
        <button id="otsNextPage" class="ghost-button" type="button">Next</button>
        <button id="otsExitFullscreenButton" class="ghost-button ots-exit-fullscreen" type="button" hidden>Exit Full Screen</button>
      </div>
      <div id="otsReportLauncher" class="ots-report-launcher">
        <button id="otsReportToggleButton" class="primary-button" type="button">Show Report</button>
      </div>
      <section id="otsReportPanel" class="panel ots-report-panel" hidden>
        <div class="panel-heading ots-report-heading">
          <div>
            <h3>OTS Change Report</h3>
            <p id="otsReportMeta" class="panel-subtitle">Compare channel-wise OTS movement between the previous and current visible weeks.</p>
          </div>
          <div class="table-meta">
            <span id="otsReportCount">0 narratives</span>
          </div>
        </div>
        <div class="ots-report-toolbar">
          <label class="ots-multiselect-field">
            <span>Market</span>
            <div class="ots-multiselect">
              <button id="otsReportMarketFilter" class="ots-select-button" type="button">All Markets</button>
              <div id="otsReportMarketFilterMenu" class="ots-multiselect-menu" hidden>
                <input id="otsReportMarketFilterSearch" class="filter-menu-search" type="text" placeholder="Search market..." autocomplete="off" />
                <div id="otsReportMarketFilterOptions" class="ots-options-list"></div>
              </div>
            </div>
          </label>
          <label class="ots-multiselect-field">
            <span>Channel</span>
            <div class="ots-multiselect">
              <button id="otsReportChannelFilter" class="ots-select-button" type="button">Default 4 Channels</button>
              <div id="otsReportChannelFilterMenu" class="ots-multiselect-menu" hidden>
                <input id="otsReportChannelFilterSearch" class="filter-menu-search" type="text" placeholder="Search channel..." autocomplete="off" />
                <div id="otsReportChannelFilterOptions" class="ots-options-list"></div>
              </div>
            </div>
          </label>
          <div class="action-row ots-report-actions">
            <button id="otsReportResetButton" class="ghost-button" type="button">Reset</button>
            <button id="otsReportDownloadButton" class="ghost-button" type="button">Download</button>
            <button id="otsReportPrintButton" class="ghost-button" type="button">Print</button>
            <button id="otsReportHideButton" class="primary-button" type="button">Hide</button>
          </div>
        </div>
        <div id="otsReportStatusMessage" class="status-message" hidden></div>
        <div id="otsReportContent" class="ots-report-stack"></div>
      </section>
    </section>

    <section class="panel comparison-panel">
      <div class="panel-heading comparison-heading">
        <div><h2>Weekly Frequency & Rank Comparison</h2></div>
        <div class="table-meta">
          <span id="comparisonResultCount">0 rows</span>
        </div>
      </div>
      <div class="comparison-toolbar">
        <label class="filter-select-field"><span>Market</span><div class="filter-select"><button id="comparisonMarketFilter" class="filter-select-button" type="button">All Markets</button><div id="comparisonMarketFilterMenu" class="filter-select-menu" hidden><input id="comparisonMarketFilterSearch" class="filter-menu-search" type="text" placeholder="Search market..." autocomplete="off" /><div id="comparisonMarketFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>City</span><div class="filter-select"><button id="comparisonCityFilter" class="filter-select-button" type="button">All Cities</button><div id="comparisonCityFilterMenu" class="filter-select-menu" hidden><input id="comparisonCityFilterSearch" class="filter-menu-search" type="text" placeholder="Search city..." autocomplete="off" /><div id="comparisonCityFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>Headend</span><div class="filter-select"><button id="comparisonHeadendFilter" class="filter-select-button" type="button">All Headends</button><div id="comparisonHeadendFilterMenu" class="filter-select-menu" hidden><input id="comparisonHeadendFilterSearch" class="filter-menu-search" type="text" placeholder="Search headend..." autocomplete="off" /><div id="comparisonHeadendFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>Channel</span><div class="filter-select"><button id="comparisonChannelFilter" class="filter-select-button" type="button">All Channels</button><div id="comparisonChannelFilterMenu" class="filter-select-menu" hidden><input id="comparisonChannelFilterSearch" class="filter-menu-search" type="text" placeholder="Search channel..." autocomplete="off" /><div id="comparisonChannelFilterOptions" class="filter-options-list"></div></div></div></label>
        <label class="filter-select-field"><span>Week</span><div class="filter-select"><button id="comparisonWeekFilter" class="filter-select-button" type="button">Select Week</button><div id="comparisonWeekFilterMenu" class="filter-select-menu" hidden><input id="comparisonWeekFilterSearch" class="filter-menu-search" type="text" placeholder="Search week..." autocomplete="off" /><div id="comparisonWeekFilterOptions" class="filter-options-list"></div></div></div></label>
        <div class="action-row comparison-actions">
          <button id="comparisonResetButton" class="ghost-button" type="button">Reset Filters</button>
          <button id="comparisonFullscreenButton" class="primary-button" type="button">Full Screen</button>
        </div>
      </div>
      <div id="comparisonStatusMessage" class="status-message" hidden></div>
      <div class="comparison-table-wrap">
        <table id="comparisonTable" class="comparison-table">
          <thead id="comparisonTableHead"></thead>
          <tbody id="comparisonTableBody"></tbody>
        </table>
      </div>
      <div class="pagination-bar comparison-pagination-bar">
        <button id="comparisonPrevPage" class="ghost-button" type="button">Previous</button>
        <span id="comparisonPageInfo">Page 1 of 1</span>
        <button id="comparisonNextPage" class="ghost-button" type="button">Next</button>
        <button id="comparisonExitFullscreenButton" class="ghost-button comparison-exit-fullscreen" type="button" hidden>Exit Full Screen</button>
      </div>
    </section>

    <section class="kpi-grid bottom-kpis">
      <article class="kpi-card compact-kpi">
        <span>Total Rows</span>
        <strong id="kpiTotalRows">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span>Total Market</span>
        <strong id="kpiTotalMarket">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span>Total City</span>
        <strong id="kpiTotalCity">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span>Total MSO Type</span>
        <strong id="kpiTotalMsoType">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span>Total Headend</span>
        <strong id="kpiTotalHeadend">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span>Total Channel</span>
        <strong id="kpiTotalChannel">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span>Total Band</span>
        <strong id="kpiTotalBand">0</strong>
      </article>
    </section>

    <section class="download-bar">
      <button id="downloadDashboardButton" class="ghost-button" type="button">Download Dashboard</button>
    </section>
  </div>

  <template id="emptyStateTemplate">
    <tr><td colspan="100%" class="empty-state">No records match the current filters.</td></tr>
  </template>

  <script src="./frequency_report.json"></script>
  <script>
const reportBundle = window.__CHROME_REPORT_DATA__ || {
  frequency: { generated_at: "", weeks: [], records: [], message: "Dashboard data file could not be loaded." },
  comparison: { generated_at: "", weeks: [], pairs: [], rows_by_pair: {}, message: "Comparison data file could not be loaded." },
  nbhd_benchmark: { generated_at: "", weeks: [], records: [], message: "INDIA TV comparison data file could not be loaded.", source_directory: "" },
  nbhd: { generated_at: "", weeks: [], filters: { markets: [], cities: [], head_ends: [] }, table: { records: [], total_count: 0 }, message: "Neighbourhood data file could not be loaded.", source_directory: "" },
  ots: { generated_at: "", weeks: [], visible_weeks: [], filters: { markets: [], channels: [] }, table: { records: [], total_count: 0 }, message: "OTS data file could not be loaded.", source_directory: "" }
};
const report = reportBundle.frequency;
function normalizeWeeks(weeks) {
  return Array.isArray(weeks) ? weeks.filter((week) => String(week || "").trim() !== "") : [];
}
const state = {
  view: "frequency",
  filters: { market: "", city: "", mso_type: "", head_end: "", crn_no: "", channel_name: "", band: "", week_from: "", week_to: "", change: "" },
  sortKey: "flow_order",
  sortDirection: "asc",
  page: 1,
  pageSize: 30,
};
const tableColumns = [
  { key: "market", label: "MARKET" },
  { key: "city", label: "CITY" },
  { key: "mso_type", label: "MSO TYPE" },
  { key: "head_end", label: "HEAD-END" },
  { key: "crn_no", label: "CRN No." },
  { key: "channel_name", label: "CHANNEL NAME" },
];
const filterOrder = ["market", "city", "mso_type", "head_end", "crn_no", "channel_name", "band", "week_from", "week_to", "change"];
const fieldMap = { market: "market", city: "city", mso_type: "mso_type", head_end: "head_end", crn_no: "crn_no", channel_name: "channel_name", band: "band" };
function getSingleSelectControl(id) {
  return {
    button: document.getElementById(id),
    menu: document.getElementById(`${id}Menu`),
    search: document.getElementById(`${id}Search`),
    options: document.getElementById(`${id}Options`),
  };
}
const filterPlaceholders = {
  market: "All Markets",
  city: "All Cities",
  mso_type: "All MSO Types",
  head_end: "All Headend",
  crn_no: "All CRN No",
  channel_name: "All Channels",
  band: "All Bands",
  week_from: "From Week",
  week_to: "To Week",
  change: "All Changes",
};
const filterSearchPlaceholders = {
  market: "Search market...",
  city: "Search city...",
  mso_type: "Search MSO type...",
  head_end: "Search headend...",
  crn_no: "Search CRN...",
  channel_name: "Search channel...",
  band: "Search band...",
  week_from: "Search week...",
  week_to: "Search week...",
  change: "Search change...",
};
const filters = {
  market: getSingleSelectControl("marketFilter"),
  city: getSingleSelectControl("cityFilter"),
  mso_type: getSingleSelectControl("msoTypeFilter"),
  head_end: getSingleSelectControl("headendFilter"),
  crn_no: getSingleSelectControl("crnFilter"),
  channel_name: getSingleSelectControl("channelFilter"),
  band: getSingleSelectControl("bandFilter"),
  week_from: getSingleSelectControl("weekFromFilter"),
  week_to: getSingleSelectControl("weekToFilter"),
  change: getSingleSelectControl("changeFilter"),
};
const viewButtons = {
  frequency: document.getElementById("frequencyViewButton"),
  rank: document.getElementById("rankViewButton"),
  band: document.getElementById("bandViewButton"),
};
const fullscreenButton = document.getElementById("fullscreenButton");
const exitFullscreenButton = document.getElementById("exitFullscreenButton");
const tableFullscreenScope = document.querySelector(".table1-scope");
const filterPanel = document.querySelector(".filter-panel");
const tablePanel = document.querySelector(".table-panel");
const tableWrap = document.querySelector(".table-wrap");
const fullscreenState = {
  active: false,
  usingNativeFullscreen: false,
  windowScrollY: 0,
  tableScrollTop: 0,
  tableScrollLeft: 0,
};
const DEFAULT_CHANNEL_REPORTS = ["INDIA TV", "AAJ TAK", "NEWS 18 INDIA", "REPUBLIC BHARAT"];
const channelReportState = {
  channel: "__default__",
  week_from: "",
  week_to: "",
  open: false,
};
const channelReportControls = {
  channel: document.getElementById("channelReportChannelFilter"),
  week_from: document.getElementById("channelReportWeekFromFilter"),
  week_to: document.getElementById("channelReportWeekToFilter"),
  container: document.getElementById("channelReportContainer"),
  count: document.getElementById("channelReportCount"),
  panel: document.querySelector(".channel-report-panel"),
  toggle: document.getElementById("channelReportToggleButton"),
  reset: document.getElementById("channelReportResetButton"),
  hide: document.getElementById("channelReportHideButton"),
};
function formatNumber(value) { return new Intl.NumberFormat().format(value || 0); }
function formatTimestamp(value) {
  const date = new Date(value);
  return Number.isNaN(date.getTime()) ? "--" : date.toLocaleString("en-IN", { year: "numeric", month: "2-digit", day: "2-digit", hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false });
}
function createOption(value, label) {
  const option = document.createElement("option");
  option.value = value;
  option.textContent = label;
  return option;
}
function populateOptionList(select, options, selectedValue) {
  const safeOptions = Array.isArray(options) ? options.filter((option) => option && String(option.value || "").trim() !== "") : [];
  const fallback = safeOptions.some((option) => option.value === selectedValue) ? selectedValue : (safeOptions[0]?.value || "");
  select.innerHTML = "";
  safeOptions.forEach((option) => select.appendChild(createOption(option.value, option.label)));
  select.value = fallback;
  return fallback;
}
function updateSingleSelectButton(control, value, placeholder, labels = null) {
  if (!control?.button) return;
  control.button.textContent = value ? (labels?.[value] || value) : placeholder;
}
function renderSingleSelectOptions(control, values, selectedValue, placeholder, onSelect, labels = null) {
  if (!control?.options) return;
  const safeValues = Array.isArray(values) ? values.filter((value) => value !== null && value !== undefined && String(value).trim() !== "") : [];
  const query = String(control.search?.value || "").trim().toLowerCase();
  const fragment = document.createDocumentFragment();
  const options = [{ value: "", label: placeholder }, ...safeValues.map((value) => ({ value, label: labels?.[value] || value }))];
  options
    .filter((option) => !query || String(option.label || "").toLowerCase().includes(query))
    .forEach((option) => {
      const item = document.createElement("button");
      item.type = "button";
      item.className = `filter-option-row${option.value === selectedValue ? " active" : ""}`;
      item.textContent = option.label;
      item.addEventListener("click", () => onSelect(option.value));
      fragment.appendChild(item);
    });
  control.options.replaceChildren(fragment);
}
function getConstrainedWeekOptions(key) {
  const weeks = normalizeWeeks(report.weeks || []);
  if (key === "week_from") {
    const toIndex = state.filters.week_to && weeks.includes(state.filters.week_to) ? weeks.indexOf(state.filters.week_to) : weeks.length - 1;
    return weeks.slice(0, toIndex + 1);
  }
  if (key === "week_to") {
    const fromIndex = state.filters.week_from && weeks.includes(state.filters.week_from) ? weeks.indexOf(state.filters.week_from) : 0;
    return weeks.slice(fromIndex);
  }
  return weeks;
}
function getControlOptions(key) {
  if (key === "week_from" || key === "week_to") return getConstrainedWeekOptions(key);
  return getOptions(key);
}
function closeSingleSelectMenus(exceptControl = null) {
  Object.values(filters).forEach((control) => {
    if (control !== exceptControl && control?.menu) control.menu.hidden = true;
  });
}
function syncSingleSelect(control, values, placeholder, selectedValue, onSelect, labels = null) {
  const safeValues = Array.isArray(values) ? values.filter((value) => value !== null && value !== undefined && String(value).trim() !== "") : [];
  const fallback = safeValues.includes(selectedValue) ? selectedValue : "";
  updateSingleSelectButton(control, fallback, placeholder, labels);
  renderSingleSelectOptions(control, safeValues, fallback, placeholder, (value) => {
    onSelect(value);
    closeSingleSelectMenus();
  }, labels);
  return fallback;
}
function bindSingleSelect(control, key, onApply) {
  if (!control?.button) return;
  control.button.addEventListener("click", (event) => {
    event.stopPropagation();
    const next = control.menu?.hidden ?? false;
    closeSingleSelectMenus();
    if (control.menu) control.menu.hidden = !next;
    if (next && control.search) {
      control.search.value = "";
      control.search.dispatchEvent(new Event("input"));
      requestAnimationFrame(() => control.search?.focus());
    }
  });
  if (control.search) {
    control.search.placeholder = filterSearchPlaceholders[key] || "Search...";
    control.search.addEventListener("click", (event) => event.stopPropagation());
    control.search.addEventListener("input", () => {
      renderSingleSelectOptions(control, getControlOptions(key), state.filters[key], filterPlaceholders[key], onApply, key === "change" ? { Changed: "Changed", "No Change": "No Change" } : null);
    });
  }
}
function getVisibleWeeks() {
  const weeks = normalizeWeeks(report.weeks || []);
  if (!weeks.length) return [];
  if (state.filters.week_from || state.filters.week_to) {
    const fromIndex = state.filters.week_from && weeks.includes(state.filters.week_from) ? weeks.indexOf(state.filters.week_from) : 0;
    const toIndex = state.filters.week_to && weeks.includes(state.filters.week_to) ? weeks.indexOf(state.filters.week_to) : weeks.length - 1;
    const start = Math.min(fromIndex, toIndex);
    const end = Math.max(fromIndex, toIndex);
    return weeks.slice(start, end + 1);
  }
  return weeks.slice(Math.max(0, weeks.length - 4));
}
function formatChannelLabel(value) {
  return String(value || "").toUpperCase();
}
function getChannelReportWeekPair() {
  const allWeeks = normalizeWeeks(report.weeks || []);
  if (!allWeeks.length) return [];
  const fallbackTo = allWeeks[allWeeks.length - 1];
  const fallbackFrom = allWeeks[Math.max(0, allWeeks.length - 2)] || fallbackTo;
  const from = allWeeks.includes(channelReportState.week_from) ? channelReportState.week_from : fallbackFrom;
  const to = allWeeks.includes(channelReportState.week_to) ? channelReportState.week_to : fallbackTo;
  let fromIndex = allWeeks.indexOf(from);
  let toIndex = allWeeks.indexOf(to);
  if (fromIndex === toIndex && allWeeks.length > 1) {
    fromIndex = Math.max(0, toIndex - 1);
  }
  if (fromIndex > toIndex) {
    const swap = fromIndex;
    fromIndex = toIndex;
    toIndex = swap;
  }
  channelReportState.week_from = allWeeks[fromIndex];
  channelReportState.week_to = allWeeks[toIndex];
  return [allWeeks[fromIndex], allWeeks[toIndex]];
}
function buildChannelReportOptions() {
  const channels = Array.from(
    new Set(
      (report.records || [])
        .map((record) => String(record.channel_name || "").trim())
        .filter(Boolean)
    )
  ).sort((left, right) => left.localeCompare(right));
  return [
    { value: "__default__", label: "Default 4 Channels" },
    ...channels.map((channel) => ({ value: channel, label: formatChannelLabel(channel) })),
  ];
}
function getChannelReportTargets() {
  if (channelReportState.channel && channelReportState.channel !== "__default__") {
    return [channelReportState.channel];
  }
  const available = new Set((report.records || []).map((record) => String(record.channel_name || "").trim()));
  return DEFAULT_CHANNEL_REPORTS.filter((channel) => available.has(channel));
}
function buildChannelReportRows(channel, weeks) {
  const [previousWeek, currentWeek] = weeks;
  const grouped = new Map();
  (report.records || []).forEach((record) => {
    if (String(record.channel_name || "").trim() !== channel) return;
    const market = String(record.market || "").trim();
    const headend = String(record.head_end || "").trim();
    const key = `${market}||${headend}`;
    if (!grouped.has(key)) grouped.set(key, record);
  });
  return Array.from(grouped.values())
    .map((record) => {
      const previousFrequency = record.frequencies?.[previousWeek];
      const currentFrequency = record.frequencies?.[currentWeek];
      const previousRank = record.ranks?.[previousWeek];
      const currentRank = record.ranks?.[currentWeek];
      const previousMissing = previousFrequency === null || previousFrequency === undefined || previousFrequency === "";
      const currentMissing = currentFrequency === null || currentFrequency === undefined || currentFrequency === "";
      const hasFrequencyChange = previousMissing !== currentMissing || (!previousMissing && !currentMissing && previousFrequency !== currentFrequency);
      return {
        channel_name: record.channel_name,
        market: record.market,
        head_end: record.head_end,
        previousFrequency,
        currentFrequency,
        previousRank,
        currentRank,
        hasFrequencyChange,
      };
    })
    .filter((record) => record.hasFrequencyChange)
    .sort((left, right) => {
      const marketCompare = String(left.market || "").localeCompare(String(right.market || ""));
      if (marketCompare !== 0) return marketCompare;
      return String(left.head_end || "").localeCompare(String(right.head_end || ""));
    });
}
function buildChannelReportNotes(channel, rows) {
  const changed = [];
  const added = [];
  const dropped = [];
  rows.forEach((row) => {
    const previousMissing = row.previousFrequency === null || row.previousFrequency === undefined || row.previousFrequency === "";
    const currentMissing = row.currentFrequency === null || row.currentFrequency === undefined || row.currentFrequency === "";
    const location = `${row.market} - ${row.head_end}`;
    if (!previousMissing && !currentMissing && row.previousFrequency !== row.currentFrequency) changed.push(location);
    if (previousMissing && !currentMissing) added.push(location);
    if (!previousMissing && currentMissing) dropped.push(location);
  });
  const label = formatChannelLabel(channel);
  const notes = [];
  if (changed.length) notes.push(`${label} LCN changed in ${changed.length} head end${changed.length === 1 ? "" : "s"}: ${changed.join(", ")}`);
  if (added.length) notes.push(`${label} became available in ${added.length} head end${added.length === 1 ? "" : "s"}: ${added.join(", ")}`);
  if (dropped.length) notes.push(`${label} dropped from ${dropped.length} head end${dropped.length === 1 ? "" : "s"}: ${dropped.join(", ")}`);
  if (!notes.length) notes.push(`${label} has no frequency movement in the selected weeks.`);
  return notes;
}
function renderChannelReports() {
  if (channelReportControls.panel) {
    channelReportControls.panel.hidden = !channelReportState.open;
  }
  if (channelReportControls.toggle) {
    channelReportControls.toggle.textContent = channelReportState.open ? "Hide Report" : "Show Report";
  }
  if (!channelReportState.open) return;
  const container = channelReportControls.container;
  if (!container) return;
  const weeks = getChannelReportWeekPair();
  const allWeeks = normalizeWeeks(report.weeks || []);
  channelReportState.channel = populateOptionList(channelReportControls.channel, buildChannelReportOptions(), channelReportState.channel || "__default__");
  channelReportState.week_from = populateSelect(channelReportControls.week_from, allWeeks, "From Week", channelReportState.week_from);
  channelReportState.week_to = populateSelect(channelReportControls.week_to, allWeeks, "To Week", channelReportState.week_to);
  const activeWeeks = getChannelReportWeekPair();
  const channels = getChannelReportTargets();
  channelReportControls.count.textContent = `${channels.length} channel${channels.length === 1 ? "" : "s"}`;
  const fragment = document.createDocumentFragment();
  if (!channels.length) {
    const empty = document.createElement("div");
    empty.className = "empty-state";
    empty.textContent = "No channel report data available.";
    container.replaceChildren(empty);
    return;
  }
  channels.forEach((channel) => {
    const rows = buildChannelReportRows(channel, activeWeeks);
    const notes = buildChannelReportNotes(channel, rows);
    const card = document.createElement("section");
    card.className = "channel-report-card";

    const header = document.createElement("div");
    header.className = "channel-report-card-header";
    const title = document.createElement("h3");
    title.textContent = formatChannelLabel(channel);
    header.appendChild(title);
    card.appendChild(header);

    const wrap = document.createElement("div");
    wrap.className = "channel-report-table-wrap";
    const table = document.createElement("table");
    table.className = "channel-report-table";
    const thead = document.createElement("thead");

    const groupRow = document.createElement("tr");
    [
      { text: "CHANNEL NAME", rowSpan: 2 },
      { text: "MARKET", rowSpan: 2 },
      { text: "HEAD-END", rowSpan: 2 },
    ].forEach((column) => {
      const th = document.createElement("th");
      th.textContent = column.text;
      th.rowSpan = column.rowSpan;
      th.className = "channel-report-subhead";
      groupRow.appendChild(th);
    });
    [
      { text: "Freq", colSpan: activeWeeks.length },
      { text: "Rank", colSpan: activeWeeks.length },
    ].forEach((group) => {
      const th = document.createElement("th");
      th.textContent = group.text;
      th.colSpan = Math.max(group.colSpan, 1);
      th.className = "channel-report-group-head";
      groupRow.appendChild(th);
    });

    const weekRow = document.createElement("tr");
    [...activeWeeks, ...activeWeeks].forEach((week, index, allWeeks) => {
      const th = document.createElement("th");
      th.textContent = week;
      const weekGroupLength = Math.max(activeWeeks.length, 1);
      const weekIndex = index % weekGroupLength;
      th.className = `channel-report-subhead${weekIndex === 0 ? " channel-report-group-start" : ""}${weekIndex === weekGroupLength - 1 ? " channel-report-group-end" : ""}`;
      weekRow.appendChild(th);
    });
    thead.append(groupRow, weekRow);

    const tbody = document.createElement("tbody");
    if (!rows.length) {
      const tr = document.createElement("tr");
      const td = document.createElement("td");
      td.colSpan = 3 + activeWeeks.length * 2;
      td.className = "empty-state";
      td.textContent = "No frequency changes found for the selected weeks.";
      tr.appendChild(td);
      tbody.appendChild(tr);
    } else {
      rows.forEach((row) => {
        const tr = document.createElement("tr");
        [row.channel_name, row.market, row.head_end].forEach((value, index) => {
          const td = document.createElement("td");
          td.textContent = value === row.channel_name ? String(value || "").toUpperCase() : (value || "");
          td.className = index === 2 ? "channel-report-leading-end" : "channel-report-leading";
          tr.appendChild(td);
        });
        [
          { previous: row.previousFrequency, current: row.currentFrequency },
          { previous: row.previousRank, current: row.currentRank },
        ].forEach((pair, pairIndex) => {
          [pair.previous, pair.current].forEach((value, index) => {
            const td = document.createElement("td");
            td.textContent = value === null || value === undefined || value === "" ? "NA" : String(value);
            const boundaryClass = `${index === 0 ? " channel-report-group-start" : ""}${index === 1 ? " channel-report-group-end" : ""}`;
            td.className = `channel-report-cell-stable${boundaryClass}${pairIndex === 0 ? " channel-report-freq-cell" : " channel-report-rank-cell"}`;
            if (value === null || value === undefined || value === "") td.className = `channel-report-cell-missing${boundaryClass}${pairIndex === 0 ? " channel-report-freq-cell" : " channel-report-rank-cell"}`;
            if (index === 1 && pair.previous !== pair.current && value !== null && value !== undefined && value !== "") {
              const currentNumber = Number(value);
              const previousNumber = Number(pair.previous);
              if (!Number.isNaN(currentNumber) && !Number.isNaN(previousNumber)) {
                td.className = `${currentNumber > previousNumber ? "channel-report-cell-increase" : "channel-report-cell-decrease"}${boundaryClass}${pairIndex === 0 ? " channel-report-freq-cell" : " channel-report-rank-cell"}`;
              }
            }
            tr.appendChild(td);
          });
        });
        tbody.appendChild(tr);
      });
    }
    table.append(thead, tbody);
    wrap.appendChild(table);
    card.appendChild(wrap);

    const notesList = document.createElement("ul");
    notesList.className = "channel-report-notes";
    notes.forEach((note) => {
      const item = document.createElement("li");
      item.textContent = note;
      notesList.appendChild(item);
    });
    card.appendChild(notesList);
    fragment.appendChild(card);
  });
  container.replaceChildren(fragment);
}
function resetChannelReports() {
  channelReportState.channel = "__default__";
  channelReportState.week_from = "";
  channelReportState.week_to = "";
  renderChannelReports();
}
function getActiveBaseView() {
  return state.view === "report" ? "frequency" : state.view;
}
function isMissingValue(value) {
  return value === null || value === undefined || value === "";
}
function getDisplayStatus(record, viewConfig, weeks, weekIndex) {
  if (weekIndex === 0) return "baseline";
  const week = weeks[weekIndex];
  const previousWeek = weeks[weekIndex - 1];
  const currentValue = record[viewConfig.series]?.[week];
  const previousValue = record[viewConfig.series]?.[previousWeek];
  const currentMissing = isMissingValue(currentValue);
  const previousMissing = isMissingValue(previousValue);
  if (previousMissing && currentMissing) return "no_change";
  if (previousMissing && !currentMissing) {
    return state.view === "rank" ? "improve" : "increase";
  }
  if (!previousMissing && currentMissing) {
    return state.view === "rank" ? "decline" : "decrease";
  }
  if (currentMissing) return "missing";
  return record[viewConfig.changes]?.[week] || "no_change";
}
function isChangedStatus(status) {
  return ["increase", "decrease", "improve", "decline", "change"].includes(status);
}
function hasVisibleChange(record, viewConfig, weeks) {
  if (weeks.length <= 1) return false;
  return weeks.slice(1).some((_week, index) => isChangedStatus(getDisplayStatus(record, viewConfig, weeks, index + 1)));
}
function getViewConfig() {
  const activeView = getActiveBaseView();
  if (activeView === "rank") return { series: "ranks", changes: "rank_changes", status: "rank_change_status", positive: "improve", negative: "decline", kpiOne: "Rank Improved", kpiTwo: "Rank Declined", title: "Weekly Rank Analysis" };
  if (activeView === "band") return { series: "bands", changes: "band_changes", status: "band_change_status", positive: "change", negative: "no_change", kpiOne: "Band Changed", kpiTwo: "Band Stable", title: "Weekly Band Analysis" };
  return { series: "frequencies", changes: "changes", status: "change_status", positive: "increase", negative: "decrease", kpiOne: "Frequency Increased", kpiTwo: "Frequency Decreased", title: "Weekly Frequency Analysis" };
}
function filterRecords(ignoreKey = "") {
  const viewConfig = getViewConfig();
  const visibleWeeks = getVisibleWeeks();
  return report.records.filter((record) => {
    for (const [key, field] of Object.entries(fieldMap)) {
      if (key === ignoreKey) continue;
      if (state.filters[key] && String(record[field] || "") !== state.filters[key]) return false;
    }
    if (ignoreKey !== "change" && state.filters.change) {
      if (state.filters.change === "Changed") {
        if (!hasVisibleChange(record, viewConfig, visibleWeeks)) {
          return false;
        }
      }
      if (state.filters.change === "No Change") {
        if (hasVisibleChange(record, viewConfig, visibleWeeks)) {
          return false;
        }
      }
    }
    return true;
  });
}
function getOptions(key) {
  if (key === "week_from" || key === "week_to") return getConstrainedWeekOptions(key);
  if (key === "change") return ["Changed", "No Change"];
  const field = fieldMap[key];
  const values = new Set();
  filterRecords(key).forEach((record) => {
    const value = String(record[field] || "").trim();
    if (value) values.add(value);
  });
  return Array.from(values).sort((a, b) => a.localeCompare(b));
}
function populateSelect(select, values, allLabel, selectedValue) {
  const safeValues = values.filter((value) => value !== null && value !== undefined && String(value).trim() !== "");
  if (select instanceof HTMLSelectElement) {
    const safeSelectedValue = safeValues.includes(selectedValue) ? selectedValue : "";
    select.innerHTML = "";
    select.appendChild(createOption("", allLabel));
    safeValues.forEach((value) => select.appendChild(createOption(value, value)));
    select.value = safeSelectedValue;
    return safeSelectedValue;
  }
  const safeSelectedValue = safeValues.includes(selectedValue) ? selectedValue : "";
  return syncSingleSelect(select, safeValues, allLabel, safeSelectedValue, () => {}, null);
}
function applyFilterValue(key, value) {
  state.filters[key] = value;
  if (key === "week_from" || key === "week_to") {
    const weeks = normalizeWeeks(report.weeks || []);
    const fromIndex = state.filters.week_from && weeks.includes(state.filters.week_from) ? weeks.indexOf(state.filters.week_from) : -1;
    const toIndex = state.filters.week_to && weeks.includes(state.filters.week_to) ? weeks.indexOf(state.filters.week_to) : -1;
    if (fromIndex >= 0 && toIndex >= 0 && fromIndex > toIndex) {
      if (key === "week_from") state.filters.week_to = state.filters.week_from;
      else state.filters.week_from = state.filters.week_to;
    }
  }
  const changedIndex = filterOrder.indexOf(key);
  if (changedIndex >= 0) filterOrder.slice(changedIndex + 1).forEach((nextKey) => { state.filters[nextKey] = ""; });
  if (key === "week_from" && value) {
    state.filters.week_to = state.filters.week_to && getConstrainedWeekOptions("week_to").includes(state.filters.week_to) ? state.filters.week_to : value;
  }
  if (key === "week_to" && value) {
    state.filters.week_from = state.filters.week_from && getConstrainedWeekOptions("week_from").includes(state.filters.week_from) ? state.filters.week_from : value;
  }
  state.page = 1;
  render();
}
function syncFilters() {
  state.filters.market = syncSingleSelect(filters.market, getOptions("market"), "All Markets", state.filters.market, (value) => applyFilterValue("market", value), null);
  state.filters.city = syncSingleSelect(filters.city, getOptions("city"), "All Cities", state.filters.city, (value) => applyFilterValue("city", value), null);
  state.filters.mso_type = syncSingleSelect(filters.mso_type, getOptions("mso_type"), "All MSO Types", state.filters.mso_type, (value) => applyFilterValue("mso_type", value), null);
  state.filters.head_end = syncSingleSelect(filters.head_end, getOptions("head_end"), "All Headend", state.filters.head_end, (value) => applyFilterValue("head_end", value), null);
  state.filters.crn_no = syncSingleSelect(filters.crn_no, getOptions("crn_no"), "All CRN No", state.filters.crn_no, (value) => applyFilterValue("crn_no", value), null);
  state.filters.channel_name = syncSingleSelect(filters.channel_name, getOptions("channel_name"), "All Channels", state.filters.channel_name, (value) => applyFilterValue("channel_name", value), null);
  state.filters.band = syncSingleSelect(filters.band, getOptions("band"), "All Bands", state.filters.band, (value) => applyFilterValue("band", value), null);
  state.filters.week_from = syncSingleSelect(filters.week_from, getOptions("week_from"), "From Week", state.filters.week_from, (value) => applyFilterValue("week_from", value), null);
  state.filters.week_to = syncSingleSelect(filters.week_to, getOptions("week_to"), "To Week", state.filters.week_to, (value) => applyFilterValue("week_to", value), null);
  state.filters.change = syncSingleSelect(filters.change, getOptions("change"), "All Changes", state.filters.change, (value) => applyFilterValue("change", value), { Changed: "Changed", "No Change": "No Change" });
}
function sortValue(record, sortKey, weeks) {
  if (sortKey === "flow_order") {
    return [0, [
      String(record.market || "").toLowerCase(),
      String(record.city || "").toLowerCase(),
      String(record.head_end || "").toLowerCase(),
      String(record.channel_name || "").toLowerCase(),
    ]];
  }
  const viewConfig = getViewConfig();
  let value = record[sortKey];
  if (value === undefined && weeks.includes(sortKey)) value = record[viewConfig.series][sortKey];
  if (value === null || value === undefined || value === "") return [1, ""];
  if (typeof value === "number") return [0, value];
  return [0, String(value).toLowerCase()];
}
function getFilteredRecords() {
  const visibleWeeks = getVisibleWeeks();
  const items = filterRecords();
  const sorted = items.slice().sort((a, b) => {
    const left = sortValue(a, state.sortKey, visibleWeeks);
    const right = sortValue(b, state.sortKey, visibleWeeks);
    if (left[0] !== right[0]) return left[0] - right[0];
    if (left[1] < right[1]) return state.sortDirection === "asc" ? -1 : 1;
    if (left[1] > right[1]) return state.sortDirection === "asc" ? 1 : -1;
    return 0;
  });
  return sorted;
}
function getReportWeeks() {
  const visibleWeeks = getVisibleWeeks();
  return visibleWeeks.length > 2 ? visibleWeeks.slice(-2) : visibleWeeks.slice();
}
function getReportChannel(records) {
  const selectedChannel = String(state.filters.channel_name || "").trim();
  if (selectedChannel) return selectedChannel;
  const channels = Array.from(
    new Set(
      records
        .map((record) => String(record.channel_name || "").trim())
        .filter(Boolean)
    )
  ).sort((left, right) => left.localeCompare(right));
  return channels[0] || "";
}
function getReportRows(records) {
  const channel = getReportChannel(records);
  const reportWeeks = getReportWeeks();
  const rows = records.filter((record) => String(record.channel_name || "").trim() === channel);
  return { channel, weeks: reportWeeks, rows };
}
function buildReportNotes(rows, weeks, channel) {
  if (!channel) {
    return ["Select a channel in the Channel filter to view the report."];
  }
  if (weeks.length < 2) {
    return [`${channel} report needs at least two visible weeks.`];
  }
  const [previousWeek, currentWeek] = weeks;
  const changedHeadends = [];
  const newHeadends = [];
  const droppedHeadends = [];
  rows.forEach((record) => {
    const previous = record.frequencies?.[previousWeek];
    const current = record.frequencies?.[currentWeek];
    const headend = String(record.head_end || "").trim();
    const previousMissing = previous === null || previous === undefined || previous === "";
    const currentMissing = current === null || current === undefined || current === "";
    if (!previousMissing && !currentMissing && previous !== current) changedHeadends.push(headend);
    if (previousMissing && !currentMissing) newHeadends.push(headend);
    if (!previousMissing && currentMissing) droppedHeadends.push(headend);
  });
  const notes = [];
  if (changedHeadends.length) notes.push(`${channel} LCN changed in ${changedHeadends.length} head end${changedHeadends.length === 1 ? "" : "s"}: ${changedHeadends.join(", ")}`);
  if (newHeadends.length) notes.push(`${channel} became available in ${newHeadends.length} head end${newHeadends.length === 1 ? "" : "s"}: ${newHeadends.join(", ")}`);
  if (droppedHeadends.length) notes.push(`${channel} dropped in ${droppedHeadends.length} head end${droppedHeadends.length === 1 ? "" : "s"}: ${droppedHeadends.join(", ")}`);
  if (!notes.length) notes.push(`${channel} has no frequency movement in ${currentWeek} compared with ${previousWeek}.`);
  return notes;
}
function buildTableHead() {
  const tableHead = document.getElementById("tableHead");
  if (state.view === "report") {
    const reportData = getReportRows(getFilteredRecords());
    const titleRow = document.createElement("tr");
    const titleHead = document.createElement("th");
    titleHead.colSpan = 3 + reportData.weeks.length * 2;
    titleHead.textContent = reportData.channel || "Channel Report";
    titleHead.className = "report-channel-title";
    titleRow.appendChild(titleHead);

    const groupRow = document.createElement("tr");
    const leadingColumns = [
      { label: "CHANNEL NAME", rowSpan: 2 },
      { label: "MARKET", rowSpan: 2 },
      { label: "HEAD-END", rowSpan: 2 },
    ];
    leadingColumns.forEach((column) => {
      const th = document.createElement("th");
      th.textContent = column.label;
      th.rowSpan = column.rowSpan;
      th.className = "report-subhead";
      groupRow.appendChild(th);
    });
    [
      { label: "Freq", span: reportData.weeks.length },
      { label: "Rank", span: reportData.weeks.length },
    ].forEach((group) => {
      const th = document.createElement("th");
      th.textContent = group.label;
      th.colSpan = Math.max(group.span, 1);
      th.className = "report-group-head";
      groupRow.appendChild(th);
    });

    const weekRow = document.createElement("tr");
    [...reportData.weeks, ...reportData.weeks].forEach((week, index) => {
      const th = document.createElement("th");
      th.textContent = week;
      const weekGroupLength = Math.max(reportData.weeks.length, 1);
      const weekIndex = index % weekGroupLength;
      th.className = `report-subhead${weekIndex === 0 ? " report-group-start" : ""}${weekIndex === weekGroupLength - 1 ? " report-group-end" : ""}`;
      weekRow.appendChild(th);
    });
    tableHead.replaceChildren(titleRow, groupRow, weekRow);
    return;
  }
  const tr = document.createElement("tr");
  [...tableColumns, ...getVisibleWeeks().map((week) => ({ key: week, label: week })), { key: "change_status", label: "CHANGE" }].forEach((column) => {
    const th = document.createElement("th");
    const isActive = state.sortKey === column.key;
    const suffix = isActive ? (state.sortDirection === "asc" ? " ▲" : " ▼") : "";
    th.textContent = `${column.label}${suffix}`;
    th.className = "sortable";
    th.addEventListener("click", () => {
      if (state.sortKey === column.key) state.sortDirection = state.sortDirection === "asc" ? "desc" : "asc";
      else { state.sortKey = column.key; state.sortDirection = "asc"; }
      render();
    });
    tr.appendChild(th);
  });
  tableHead.replaceChildren(tr);
}
function formatWeekValue(value, status, isBaseline) {
  if (value === null || value === undefined || value === "") return "NA";
  if (isBaseline || status === "baseline" || status === "missing" || status === "no_change") return String(value);
  if (state.view === "rank") {
    if (status === "improve") return `▲ ${value}`;
    if (status === "decline") return `▼ ${value}`;
    return String(value);
  }
  if (state.view === "band") {
    if (status === "change") return `• ${value}`;
    return String(value);
  }
  if (status === "increase") return `▲ ${value}`;
  if (status === "decrease") return `▼ ${value}`;
  return String(value);
}
function renderFocusSummary(records) {
  const container = document.getElementById("focusSummary");
  if (!container) return;
  if (state.view === "report") {
    const reportData = getReportRows(records);
    const notes = buildReportNotes(reportData.rows, reportData.weeks, reportData.channel);
    container.innerHTML = `<ul class="report-notes">${notes.map((note) => `<li>${note}</li>`).join("")}</ul>`;
    return;
  }
  const labels = { "INDIA TV": "India TV", "AAJ TAK": "Aaj Tak", "NEWS 18 INDIA": "News 18", "REPUBLIC BHARAT": "Republic Bharat" };
  const viewConfig = getViewConfig();
  const visibleWeeks = getVisibleWeeks();
  const items = Object.entries(labels).map(([channel, label]) => {
    const selected = records.filter((record) => String(record.channel_name || "").toUpperCase() === channel);
    if (!selected.length) return "";
    let positive = 0, negative = 0, noChange = 0, latestPositive = 0, latestNegative = 0;
    const latestWeek = visibleWeeks[visibleWeeks.length - 1];
    selected.forEach((record) => {
      visibleWeeks.slice(1).forEach((week, index) => {
        const status = getDisplayStatus(record, viewConfig, visibleWeeks, index + 1);
        if (status === viewConfig.positive) positive += 1;
        else if (status === viewConfig.negative) negative += 1;
        else if (status === "increase") positive += 1;
        else if (status === "decrease") negative += 1;
        else if (status === "no_change") noChange += 1;
      });
      const latestStatus = getDisplayStatus(record, viewConfig, visibleWeeks, visibleWeeks.length - 1);
      if (latestStatus === viewConfig.positive) latestPositive += 1;
      else if (latestStatus === viewConfig.negative) latestNegative += 1;
      else if (latestStatus === "increase") latestPositive += 1;
      else if (latestStatus === "decrease") latestNegative += 1;
    });
    const positiveLabel = state.view === "rank" ? "improved" : state.view === "band" ? "changed" : "increased";
    const negativeLabel = state.view === "rank" ? "declined" : state.view === "band" ? "stable" : "decreased";
    const latestText = latestPositive || latestNegative ? ` Latest: ${latestPositive ? `${formatNumber(latestPositive)} ${positiveLabel}` : ""}${latestPositive && latestNegative ? ", " : ""}${latestNegative ? `${formatNumber(latestNegative)} ${negativeLabel}` : ""} in ${latestWeek}.` : "";
    return `<div class="focus-line"><strong>${label}</strong><span>${formatNumber(selected.length)} rows, ${formatNumber(positive)} ${positiveLabel}, ${formatNumber(negative)} ${negativeLabel}, ${formatNumber(noChange)} stable.${latestText}</span></div>`;
  }).filter(Boolean).join("");
  container.innerHTML = items || '<div class="focus-line">No channel summary available for the current filters.</div>';
}
function renderTable(records) {
  const tableBody = document.getElementById("tableBody");
  if (state.view === "report") {
    const reportData = getReportRows(records);
    const totalPages = Math.max(1, Math.ceil(reportData.rows.length / state.pageSize));
    if (state.page > totalPages) state.page = totalPages;
    const start = (state.page - 1) * state.pageSize;
    const pageItems = reportData.rows.slice(start, start + state.pageSize);
    if (!pageItems.length) {
      tableBody.replaceChildren(document.getElementById("emptyStateTemplate").content.cloneNode(true));
      return;
    }
    const [previousWeek, currentWeek] = reportData.weeks;
    const fragment = document.createDocumentFragment();
    pageItems.forEach((record) => {
      const tr = document.createElement("tr");
      [
        record.channel_name,
        record.market,
        record.head_end,
      ].forEach((value, index) => {
        const td = document.createElement("td");
        td.textContent = value || "";
        td.className = index === 2 ? "report-leading-end" : "report-leading";
        tr.appendChild(td);
      });
      reportData.weeks.forEach((week, weekIndex) => {
        const td = document.createElement("td");
        const value = record.frequencies?.[week];
        td.textContent = value === null || value === undefined || value === "" ? "NA" : String(value);
        const boundaryClass = `${weekIndex === 0 ? " report-group-start" : ""}${weekIndex === reportData.weeks.length - 1 ? " report-group-end" : ""}`;
        td.className = `report-cell-neutral report-freq-cell${boundaryClass}`;
        if (value === null || value === undefined || value === "") td.className = `report-cell-missing report-freq-cell${boundaryClass}`;
        if (week === currentWeek && previousWeek && record.frequencies?.[previousWeek] !== value && value !== null && value !== undefined && value !== "") td.className = `report-cell-latest report-freq-cell${boundaryClass}`;
        tr.appendChild(td);
      });
      reportData.weeks.forEach((week, weekIndex) => {
        const td = document.createElement("td");
        const value = record.ranks?.[week];
        td.textContent = value === null || value === undefined || value === "" ? "NA" : String(value);
        const boundaryClass = `${weekIndex === 0 ? " report-group-start" : ""}${weekIndex === reportData.weeks.length - 1 ? " report-group-end" : ""}`;
        td.className = `report-cell-neutral report-rank-cell${boundaryClass}`;
        if (value === null || value === undefined || value === "") td.className = `report-cell-missing report-rank-cell${boundaryClass}`;
        if (week === currentWeek && previousWeek && record.ranks?.[previousWeek] !== value && value !== null && value !== undefined && value !== "") td.className = `report-cell-latest report-rank-cell${boundaryClass}`;
        tr.appendChild(td);
      });
      fragment.appendChild(tr);
    });
    tableBody.replaceChildren(fragment);
    return;
  }
  const visibleWeeks = getVisibleWeeks();
  const start = (state.page - 1) * state.pageSize;
  const pageItems = records.slice(start, start + state.pageSize);
  if (!pageItems.length) {
    tableBody.replaceChildren(document.getElementById("emptyStateTemplate").content.cloneNode(true));
    return;
  }
  const viewConfig = getViewConfig();
  const fragment = document.createDocumentFragment();
  pageItems.forEach((record) => {
    const tr = document.createElement("tr");
    tableColumns.forEach((column) => {
      const td = document.createElement("td");
      td.textContent = record[column.key] ?? "";
      tr.appendChild(td);
    });
    visibleWeeks.forEach((week, index) => {
      const td = document.createElement("td");
      const value = record[viewConfig.series][week];
      const status = getDisplayStatus(record, viewConfig, visibleWeeks, index);
      td.classList.add(`status-${status}`);
      td.textContent = formatWeekValue(value, status, index === 0);
      tr.appendChild(td);
    });
    const changeTd = document.createElement("td");
    changeTd.textContent = hasVisibleChange(record, viewConfig, visibleWeeks) ? "YES" : "NO";
    changeTd.classList.add(changeTd.textContent === "NO" ? "change-no" : "change-yes");
    tr.appendChild(changeTd);
    fragment.appendChild(tr);
  });
  tableBody.replaceChildren(fragment);
}
function updateKpis(records) {
  const countDistinct = (key) => new Set(
    records
      .map((record) => String(record[key] ?? "").trim())
      .filter(Boolean)
  ).size;
  document.getElementById("kpiTotalRows").textContent = formatNumber(records.length);
  document.getElementById("kpiTotalMarket").textContent = formatNumber(countDistinct("market"));
  document.getElementById("kpiTotalCity").textContent = formatNumber(countDistinct("city"));
  document.getElementById("kpiTotalMsoType").textContent = formatNumber(countDistinct("mso_type"));
  document.getElementById("kpiTotalHeadend").textContent = formatNumber(countDistinct("head_end"));
  document.getElementById("kpiTotalChannel").textContent = formatNumber(countDistinct("channel_name"));
  document.getElementById("kpiTotalBand").textContent = formatNumber(countDistinct("band"));
}
function getPageSize() {
  if (!fullscreenState.active) return 30;
  const viewportHeight = window.innerHeight || 900;
  return Math.max(45, Math.floor((viewportHeight - 230) / 26));
}
function downloadStandaloneDashboard() {
  const embeddedBundle = JSON.stringify(window.__CHROME_REPORT_DATA__ || reportBundle || {}).split("</").join("<\\/");
  const sourceTag = '<scr' + 'ipt src="./frequency_report.json"><\\/scr' + 'ipt>';
  const embeddedTag = '<scr' + 'ipt>window.__CHROME_REPORT_DATA__ = ' + embeddedBundle + ';<\\/scr' + 'ipt>';
  const html = `<!DOCTYPE html>\n${document.documentElement.outerHTML}`
    .replace(sourceTag, embeddedTag);
  const blob = new Blob([html], { type: "text/html;charset=utf-8" });
  const url = URL.createObjectURL(blob);
  const link = document.createElement("a");
  link.href = url;
  link.download = "chrome_report_dashboard.html";
  document.body.appendChild(link);
  link.click();
  link.remove();
  URL.revokeObjectURL(url);
}
function syncFullscreenButtons() {
  const label = fullscreenState.active ? "Exit Full Screen" : "Full Screen";
  fullscreenButton.textContent = label;
  if (exitFullscreenButton) exitFullscreenButton.hidden = !fullscreenState.active;
}
function setFullscreen(active) {
  if (!tableFullscreenScope || !tableWrap || fullscreenState.active === active) return;
  if (active) {
    fullscreenState.windowScrollY = window.scrollY || window.pageYOffset || 0;
    fullscreenState.tableScrollTop = tableWrap.scrollTop;
    fullscreenState.tableScrollLeft = tableWrap.scrollLeft;
    fullscreenState.active = true;
    document.body.classList.add("table-fullscreen-active");
    tableFullscreenScope.classList.add("table1-scope-fullscreen");
    filterPanel?.classList.add("table1-scope-child");
    tablePanel?.classList.add("table1-scope-child");
    state.pageSize = getPageSize();
    state.page = 1;
    render();
    requestAnimationFrame(() => {
      tableWrap.scrollTop = fullscreenState.tableScrollTop;
      tableWrap.scrollLeft = fullscreenState.tableScrollLeft;
    });
  } else {
    fullscreenState.tableScrollTop = tableWrap.scrollTop;
    fullscreenState.tableScrollLeft = tableWrap.scrollLeft;
    fullscreenState.active = false;
    document.body.classList.remove("table-fullscreen-active");
    tableFullscreenScope.classList.remove("table1-scope-fullscreen");
    filterPanel?.classList.remove("table1-scope-child");
    tablePanel?.classList.remove("table1-scope-child");
    state.pageSize = getPageSize();
    state.page = 1;
    render();
    requestAnimationFrame(() => {
      window.scrollTo({ top: fullscreenState.windowScrollY, behavior: "auto" });
      tableWrap.scrollTop = fullscreenState.tableScrollTop;
      tableWrap.scrollLeft = fullscreenState.tableScrollLeft;
    });
  }
  syncFullscreenButtons();
}
async function enterNativeFullscreen() {
  if (!tableFullscreenScope?.requestFullscreen) return false;
  try {
    fullscreenState.usingNativeFullscreen = true;
    await tableFullscreenScope.requestFullscreen();
    return true;
  } catch (error) {
    fullscreenState.usingNativeFullscreen = false;
    return false;
  }
}
async function exitNativeFullscreen() {
  if (!document.fullscreenElement) return false;
  try {
    await document.exitFullscreen();
    return true;
  } catch (error) {
    return false;
  }
}
async function toggleFullscreen() {
  if (fullscreenState.active) {
    if (fullscreenState.usingNativeFullscreen && document.fullscreenElement === tableFullscreenScope) {
      const exited = await exitNativeFullscreen();
      if (!exited) {
        fullscreenState.usingNativeFullscreen = false;
        setFullscreen(false);
      }
      return;
    }
    setFullscreen(false);
    return;
  }
  const entered = await enterNativeFullscreen();
  if (!entered) {
    fullscreenState.usingNativeFullscreen = false;
    setFullscreen(true);
  }
}
function render() {
  syncFilters();
  renderChannelReports();
  const visibleWeeks = getVisibleWeeks();
  if (report.weeks.includes(state.sortKey) && !visibleWeeks.includes(state.sortKey)) {
    state.sortKey = "flow_order";
    state.sortDirection = "asc";
  }
  const records = getFilteredRecords();
  const reportData = state.view === "report" ? getReportRows(records) : null;
  const displayCount = reportData ? reportData.rows.length : records.length;
  const totalPages = Math.max(1, Math.ceil(displayCount / state.pageSize));
  if (state.page > totalPages) state.page = totalPages;
  document.getElementById("resultCount").textContent = `${formatNumber(displayCount)} records`;
  document.getElementById("pageInfo").textContent = `Page ${state.page} of ${totalPages}`;
  document.getElementById("prevPage").disabled = state.page <= 1;
  document.getElementById("nextPage").disabled = state.page >= totalPages;
  document.getElementById("tableTitle").textContent = state.view === "report" ? "Channel Report" : getViewConfig().title;
  Object.entries(viewButtons).forEach(([view, button]) => button.classList.toggle("active", view === state.view));
  buildTableHead();
  renderTable(records);
  renderFocusSummary(records);
  updateKpis(records);
}
Object.entries(filters).forEach(([key, control]) => {
  bindSingleSelect(control, key, (value) => applyFilterValue(key, value));
});
document.addEventListener("click", (event) => {
  if (!event.target.closest(".filter-select") && !event.target.closest(".ots-multiselect")) {
    closeSingleSelectMenus();
  }
});
Object.entries(channelReportControls).forEach(([key, control]) => {
  if (!control || key === "container" || key === "count" || key === "panel" || key === "toggle" || key === "reset" || key === "hide") return;
  control.addEventListener("change", () => {
    channelReportState[key] = control.value;
    renderChannelReports();
  });
});
if (channelReportControls.toggle) {
  channelReportControls.toggle.addEventListener("click", () => {
    channelReportState.open = !channelReportState.open;
    renderChannelReports();
  });
}
if (channelReportControls.hide) {
  channelReportControls.hide.addEventListener("click", () => {
    channelReportState.open = false;
    renderChannelReports();
  });
}
if (channelReportControls.reset) {
  channelReportControls.reset.addEventListener("click", resetChannelReports);
}
document.getElementById("resetButton").addEventListener("click", () => {
  state.filters = { market: "", city: "", mso_type: "", head_end: "", crn_no: "", channel_name: "", band: "", week_from: "", week_to: "", change: "" };
  state.sortKey = "flow_order";
  state.sortDirection = "asc";
  state.page = 1;
  render();
});
document.getElementById("prevPage").addEventListener("click", () => { if (state.page > 1) { state.page -= 1; render(); } });
document.getElementById("nextPage").addEventListener("click", () => {
  const filteredRecords = getFilteredRecords();
  const totalRows = state.view === "report" ? getReportRows(filteredRecords).rows.length : filteredRecords.length;
  const totalPages = Math.max(1, Math.ceil(totalRows / state.pageSize));
  if (state.page < totalPages) { state.page += 1; render(); }
});
Object.entries(viewButtons).forEach(([view, button]) => {
  button.addEventListener("click", () => {
    state.view = view;
    state.page = 1;
    render();
  });
});
fullscreenButton.addEventListener("click", toggleFullscreen);
if (exitFullscreenButton) {
  exitFullscreenButton.addEventListener("click", async () => {
    if (fullscreenState.usingNativeFullscreen && document.fullscreenElement === tableFullscreenScope) {
      const exited = await exitNativeFullscreen();
      if (!exited) {
        fullscreenState.usingNativeFullscreen = false;
        setFullscreen(false);
      }
      return;
    }
    setFullscreen(false);
  });
}
document.addEventListener("fullscreenchange", () => {
  const isTableFullscreen = document.fullscreenElement === tableFullscreenScope;
  fullscreenState.usingNativeFullscreen = isTableFullscreen;
  if (isTableFullscreen && !fullscreenState.active) {
    setFullscreen(true);
    return;
  }
  if (!isTableFullscreen && fullscreenState.active) {
    fullscreenState.usingNativeFullscreen = false;
    setFullscreen(false);
  }
});
window.addEventListener("resize", () => {
  const nextPageSize = getPageSize();
  if (nextPageSize !== state.pageSize) {
    state.pageSize = nextPageSize;
    state.page = 1;
    render();
  }
});
document.getElementById("downloadDashboardButton").addEventListener("click", downloadStandaloneDashboard);
syncFullscreenButtons();
render();
  </script>
  <script>
window.__NBHD_BENCHMARK_INITIAL_DATA__ = reportBundle.nbhd_benchmark;
  </script>
  <script>
__NBHD_BENCHMARK_SCRIPT__
  </script>
  <script>
window.__COMPARISON_INITIAL_DATA__ = reportBundle.comparison;
  </script>
  <script>
__COMPARISON_SCRIPT__
  </script>
  <script>
window.__NBHD_STANDALONE_DATA__ = reportBundle.nbhd;
  </script>
  <script>
__NBHD_SCRIPT__
  </script>
  <script>
window.__OTS_STANDALONE_DATA__ = reportBundle.ots;
  </script>
  <script>
__OTS_SCRIPT__
  </script>
</body>
</html>
"""

    return (
        html.replace("__STYLE__", style_text)
        .replace("__NBHD_BENCHMARK_SCRIPT__", nbhd_benchmark_script_text)
        .replace("__COMPARISON_SCRIPT__", comparison_script_text)
        .replace("__NBHD_SCRIPT__", nbhd_script_text)
        .replace("__OTS_SCRIPT__", ots_script_text)
    )


def build_api_payload(view: str, filters: dict[str, str], page: int, page_size: int, sort_key: str, sort_direction: str, force_refresh: bool = False) -> dict[str, Any]:
    report = load_report(force=force_refresh)
    weeks = report.get("weeks", [])
    records = report.get("records", [])
    filtered = filter_records(records, view, filters)
    sorted_records = sort_records(filtered, sort_key, sort_direction, view)
    page_records, total_count = paginate_records(sorted_records, page, page_size)
    total_pages = max(1, (total_count + page_size - 1) // page_size) if total_count else 1

    return {
        "generated_at": report.get("generated_at"),
        "view": view,
        "weeks": weeks,
        "filters": build_filters(records, view, filters, weeks),
        "summary": summarize_records(filtered, view, weeks),
        "focus_channels": summarize_focus_channels(filtered, view, weeks),
        "message": report.get("message", ""),
        "data_directory": str(DATA_DIR),
        "table": {
            "records": serialize_records(page_records, weeks),
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": total_pages,
            "sort_key": sort_key,
            "sort_direction": sort_direction,
        },
    }


def filter_nbhd_records(records: list[dict[str, Any]], filters: dict[str, str], search: str, ignore_key: str = "") -> list[dict[str, Any]]:
    search_text = normalize_text(search).lower()
    filtered: list[dict[str, Any]] = []
    for record in records:
        if filters["market"] and record["market"] != filters["market"] and ignore_key != "market":
            continue
        if filters["city"] and record["city"] != filters["city"] and ignore_key != "city":
            continue
        if filters["head_end"] and record["head_end"] != filters["head_end"] and ignore_key != "head_end":
            continue
        if search_text:
            haystack = " ".join(
                [
                    record["market"],
                    record["city"],
                    record["head_end"],
                    *[normalize_text(record["channels"].get(week)) for week in record["channels"]],
                    *[normalize_text(record["genres"].get(week)) for week in record["genres"]],
                    *[normalize_text(record["frequencies"].get(week)) for week in record["frequencies"]],
                ]
            ).lower()
            if search_text not in haystack:
                continue
        filtered.append(record)
    return filtered


def build_nbhd_filters(records: list[dict[str, Any]], current_filters: dict[str, str], search: str) -> dict[str, list[str]]:
    def values_for(key: str, field: str) -> list[str]:
        values = {
            normalize_text(record.get(field))
            for record in filter_nbhd_records(records, current_filters, search, ignore_key=key)
            if normalize_text(record.get(field))
        }
        return sorted(values, key=lambda value: value.lower())

    return {
        "markets": values_for("market", "market"),
        "cities": values_for("city", "city"),
        "head_ends": values_for("head_end", "head_end"),
    }


def serialize_nbhd_records(records: list[dict[str, Any]], weeks: list[str]) -> list[dict[str, Any]]:
    return [
        {
            "market": record["market"],
            "city": record["city"],
            "head_end": record["head_end"],
            "position": record.get("position", 0),
            "is_reference": bool(record.get("is_reference")),
            "channels": {week: record["channels"].get(week, "") for week in weeks},
            "genres": {week: record["genres"].get(week, "") for week in weeks},
            "frequencies": {week: record["frequencies"].get(week) for week in weeks},
        }
        for record in records
    ]


def build_nbhd_api_payload(filters: dict[str, str], search: str, force_refresh: bool = False) -> dict[str, Any]:
    report = load_nbhd_report(force=force_refresh)
    weeks = report.get("weeks", [])
    records = report.get("records", [])
    filtered = filter_nbhd_records(records, filters, search)
    return {
        "generated_at": report.get("generated_at"),
        "weeks": weeks,
        "filters": build_nbhd_filters(records, filters, search),
        "search": search,
        "source_directory": report.get("source_directory", str(get_nbhd_source_dir())),
        "message": report.get("message", ""),
        "summary": {
            "total_headends": len(filtered),
        },
        "table": {
            "records": serialize_nbhd_records(filtered, weeks),
            "total_count": len(filtered),
        },
    }


def build_nbhd_export_bytes(filters: dict[str, str], search: str) -> bytes:
    payload = build_nbhd_api_payload(filters, search, force_refresh=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Neighbourhood Comparison"

    weeks = payload["weeks"]
    header_row_one = ["Market", "City", "Headend"]
    header_row_two = ["", "", ""]
    for group_name in ("Channel", "Genre", "Frequency"):
        header_row_one.extend([group_name] + [""] * (max(len(weeks) - 1, 0)))
        header_row_two.extend(weeks)

    sheet.append(header_row_one)
    sheet.append(header_row_two)

    if weeks:
        start_column = 4
        for _group_name in ("Channel", "Genre", "Frequency"):
            end_column = start_column + len(weeks) - 1
            sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
            start_column = end_column + 1
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("C1:C2")

    for record in payload["table"]["records"]:
        row_values = [record["market"], record["city"], record["head_end"]]
        row_values.extend(record["channels"].get(week, "") for week in weeks)
        row_values.extend(record["genres"].get(week, "") for week in weeks)
        row_values.extend(record["frequencies"].get(week) for week in weeks)
        sheet.append(row_values)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_multi_values(query: dict[str, list[str]], key: str) -> list[str]:
    # Support repeated query params and comma-separated values for multi-select filters.
    values: list[str] = []
    for raw in query.get(key, []):
        for part in raw.split(","):
            text = normalize_text(part)
            if text and text not in values:
                values.append(text)
    return values


def ots_visible_weeks(weeks: list[str], week_from: str, week_to: str) -> list[str]:
    # Slice dynamic week columns based on the selected from/to boundaries.
    if not weeks:
        return []
    if not week_from and not week_to:
        return weeks[max(0, len(weeks) - 8) :]
    start_index = weeks.index(week_from) if week_from in weeks else 0
    end_index = weeks.index(week_to) if week_to in weeks else len(weeks) - 1
    if start_index > end_index:
        start_index, end_index = end_index, start_index
    return weeks[start_index : end_index + 1]


def ots_change_delta(record: dict[str, Any], weeks: list[str]) -> float | None:
    # Compare the latest visible week with the previous visible week.
    if len(weeks) < 2:
        return None
    previous = record["ots_values"].get(weeks[-2])
    current = record["ots_values"].get(weeks[-1])
    if previous is None or current is None:
        return None
    return round(float(current) - float(previous), 2)


def ots_change_type(record: dict[str, Any], weeks: list[str]) -> str:
    delta = ots_change_delta(record, weeks)
    if delta is None:
        return "no_change"
    if delta > 0:
        return "increase"
    if delta < 0:
        return "decrease"
    return "no_change"


def ots_matches_change_filter(change_type: str, filter_value: str) -> bool:
    if not filter_value:
        return True
    if filter_value == "changed":
        return change_type in {"increase", "decrease"}
    if filter_value == "no_change":
        return change_type == "no_change"
    return change_type == filter_value


def filter_ots_records(records: list[dict[str, Any]], filters: dict[str, Any], ignore_key: str = "") -> list[dict[str, Any]]:
    # Apply OTS market/channel/search/change filters against the current visible week range.
    visible_weeks = ots_visible_weeks(filters["all_weeks"], filters["week_from"], filters["week_to"])
    search_text = normalize_text(filters["search"]).lower()
    filtered: list[dict[str, Any]] = []
    for record in records:
        if filters["markets"] and record["market"] not in filters["markets"] and ignore_key != "markets":
            continue
        if filters["channels"] and record["channel"] not in filters["channels"] and ignore_key != "channels":
            continue
        if search_text:
            haystack = f"{record['market']} {record['channel']}".lower()
            if search_text not in haystack:
                continue
        if ignore_key != "change" and filters["change"]:
            change_type = ots_change_type(record, visible_weeks)
            if not ots_matches_change_filter(change_type, filters["change"]):
                continue
        filtered.append(record)
    return filtered


def build_ots_filters(records: list[dict[str, Any]], current_filters: dict[str, Any]) -> dict[str, list[str]]:
    # Build dynamic filter options scoped by the current selections.
    def values_for(key: str, field: str) -> list[str]:
        values = {
            normalize_text(record.get(field))
            for record in filter_ots_records(records, current_filters, ignore_key=key)
            if normalize_text(record.get(field))
        }
        return sorted(values, key=lambda value: value.lower())

    return {
        "markets": values_for("markets", "market"),
        "channels": values_for("channels", "channel"),
        "weeks": current_filters["all_weeks"],
        "change_options": ["", "changed", "no_change", "increase", "decrease"],
    }


def serialize_ots_records(records: list[dict[str, Any]], weeks: list[str]) -> list[dict[str, Any]]:
    # Keep only the visible week columns for the frontend payload.
    return [
        {
            "market": record["market"],
            "channel": record["channel"],
            "ots_values": {week: record["ots_values"].get(week) for week in weeks},
        }
        for record in records
    ]


def build_ots_api_payload(filters: dict[str, Any], force_refresh: bool = False) -> dict[str, Any]:
    # Produce one reusable payload for both the live dashboard and standalone HTML.
    report = load_ots_report(force=force_refresh)
    all_weeks = report.get("weeks", [])
    scoped_filters = {
        **filters,
        "all_weeks": all_weeks,
    }
    visible_weeks = ots_visible_weeks(all_weeks, filters["week_from"], filters["week_to"])
    filtered = filter_ots_records(report.get("records", []), scoped_filters)
    return {
        "generated_at": report.get("generated_at"),
        "weeks": all_weeks,
        "visible_weeks": visible_weeks,
        "filters": build_ots_filters(report.get("records", []), scoped_filters),
        "message": report.get("message", ""),
        "source_directory": report.get("source_directory", str(OTS_DATA_DIR)),
        "table": {
            "records": serialize_ots_records(filtered, visible_weeks),
            "total_count": len(filtered),
        },
    }


def build_ots_export_workbook(filters: dict[str, Any]) -> Workbook:
    # Create the OTS Excel export with dynamic week columns.
    payload = build_ots_api_payload(filters, force_refresh=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OTS Comparison"
    header = ["Market", "Channel", *payload["visible_weeks"], "Change"]
    sheet.append(header)

    for record in payload["table"]["records"]:
        delta = ots_change_delta({"ots_values": record["ots_values"]}, payload["visible_weeks"])
        row = [record["market"], record["channel"]]
        row.extend(record["ots_values"].get(week) for week in payload["visible_weeks"])
        row.append(delta)
        sheet.append(row)
    return workbook


def build_ots_export_bytes(filters: dict[str, Any]) -> bytes:
    workbook = build_ots_export_workbook(filters)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_ots_csv_bytes(filters: dict[str, Any]) -> bytes:
    payload = build_ots_api_payload(filters, force_refresh=True)
    lines = [",".join(['"Market"', '"Channel"', *[f'"{week}"' for week in payload["visible_weeks"]], '"Change"'])]
    for record in payload["table"]["records"]:
        delta = ots_change_delta({"ots_values": record["ots_values"]}, payload["visible_weeks"])
        change_text = "" if delta is None else str(delta)
        cells = [record["market"], record["channel"], *[record["ots_values"].get(week) for week in payload["visible_weeks"]], change_text]
        safe_cells = ['"' + str("" if cell is None else cell).replace('"', '""') + '"' for cell in cells]
        lines.append(",".join(safe_cells))
    return "\n".join(lines).encode("utf-8")


def parse_int(value: str, default: int) -> int:
    try:
        return int(value or str(default))
    except (TypeError, ValueError):
        return default


def parse_api_request(query: dict[str, list[str]]) -> dict[str, Any]:
    view = (query.get("view", ["frequency"])[0] or "frequency").strip().lower()
    if view not in {"frequency", "rank", "band"}:
        view = "frequency"

    filters = {
        "market": (query.get("market", [""])[0] or "").strip(),
        "city": (query.get("city", [""])[0] or "").strip(),
        "mso_type": (query.get("mso_type", [""])[0] or "").strip(),
        "head_end": (query.get("head_end", [""])[0] or "").strip(),
        "crn_no": (query.get("crn_no", [""])[0] or "").strip(),
        "channel_name": (query.get("channel_name", [""])[0] or "").strip(),
        "band": (query.get("band", [""])[0] or "").strip(),
        "week": (query.get("week", [""])[0] or "").strip(),
        "change": (query.get("change", [""])[0] or "").strip(),
    }
    page = max(1, parse_int((query.get("page", ["1"])[0] or "1").strip(), 1))
    page_size = max(1, min(200, parse_int((query.get("page_size", ["30"])[0] or "30").strip(), 30)))
    sort_key = ((query.get("sort_key", ["flow_order"])[0] or "flow_order").strip() or "flow_order")
    sort_direction = (query.get("sort_direction", ["asc"])[0] or "asc").strip().lower()
    if sort_direction not in {"asc", "desc"}:
        sort_direction = "asc"
    force_refresh = (query.get("refresh", [""])[0] or "").strip() == "1"

    return build_api_payload(view, filters, page, page_size, sort_key, sort_direction, force_refresh)


def parse_nbhd_api_request(query: dict[str, list[str]]) -> dict[str, Any]:
    filters = {
        "market": (query.get("market", [""])[0] or "").strip(),
        "city": (query.get("city", [""])[0] or "").strip(),
        "head_end": (query.get("head_end", [""])[0] or "").strip(),
    }
    search = (query.get("search", [""])[0] or "").strip()
    force_refresh = (query.get("refresh", [""])[0] or "").strip() == "1"
    return build_nbhd_api_payload(filters, search, force_refresh)


def parse_ots_api_request(query: dict[str, list[str]]) -> dict[str, Any]:
    filters = {
        "markets": parse_multi_values(query, "market"),
        "channels": parse_multi_values(query, "channel"),
        "week_from": (query.get("week_from", [""])[0] or "").strip(),
        "week_to": (query.get("week_to", [""])[0] or "").strip(),
        "change": (query.get("change", [""])[0] or "").strip(),
        "search": (query.get("search", [""])[0] or "").strip(),
    }
    force_refresh = (query.get("refresh", [""])[0] or "").strip() == "1"
    return build_ots_api_payload(filters, force_refresh)


def parse_ots_filters(query: dict[str, list[str]]) -> dict[str, Any]:
    return {
        "markets": parse_multi_values(query, "market"),
        "channels": parse_multi_values(query, "channel"),
        "week_from": (query.get("week_from", [""])[0] or "").strip(),
        "week_to": (query.get("week_to", [""])[0] or "").strip(),
        "change": (query.get("change", [""])[0] or "").strip(),
        "search": (query.get("search", [""])[0] or "").strip(),
    }


def generate_standalone_dashboard() -> tuple[Path, Path]:
    ensure_directories()
    report = load_report(force=True)
    write_frequency_report_json(report)
    write_standalone_dashboard(report)
    return OUTPUT_JSON, OUTPUT_HTML


def generate_frequency_report_json() -> Path:
    ensure_directories()
    report = load_report(force=True)
    return write_frequency_report_json(report)


if __name__ == "__main__":
    json_path, html_path = generate_standalone_dashboard()
    print(f"Weekly data generated in: {DATA_DIR}")
    print(f"Report JSON updated: {json_path}")
    print(f"Standalone dashboard updated: {html_path}")
