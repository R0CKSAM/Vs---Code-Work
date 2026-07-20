from __future__ import annotations

import re
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook, load_workbook


WEEK_PATTERN = re.compile(r"(?P<year>20\d{2}).*?week\s*0*(?P<week>\d{1,2})", re.IGNORECASE)

DISTRIBUTION_SHEET_TITLE = "Distribution Summary"
NBHD_SHEET_TITLE = "NBDH Data"
OTS_SHEET_TITLE = "OTS Data"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_text(value).upper())


def trim_trailing_empty_cells(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    values = list(row)
    while values and normalize_text(values[-1]) == "":
        values.pop()
    return values


def detect_week_key(path: Path) -> tuple[int, int] | None:
    match = WEEK_PATTERN.search(path.stem)
    if not match:
        return None
    return int(match.group("year")), int(match.group("week"))


def build_output_name(year: int, week: int) -> str:
    return f"{year} WEEK{week:02d}.xlsx"


def iter_week_files(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    return sorted(
        [path for path in folder.glob("*.xlsx") if not path.name.startswith("~$")],
        key=lambda path: (detect_week_key(path) or (9999, 9999), path.name.lower()),
    )


def discover_week_bundles(
    distribution_dir: Path,
    nbhd_dir: Path,
    ots_dir: Path,
) -> dict[tuple[int, int], dict[str, Path]]:
    bundles: dict[tuple[int, int], dict[str, Path]] = {}

    for source_key, folder in (("distribution", distribution_dir), ("nbhd", nbhd_dir), ("ots", ots_dir)):
        for path in iter_week_files(folder):
            week_key = detect_week_key(path)
            if week_key is None:
                continue
            bundles.setdefault(week_key, {})[source_key] = path

    return {
        week_key: bundle
        for week_key, bundle in bundles.items()
        if {"distribution", "nbhd", "ots"}.issubset(bundle)
    }


def copy_sheet_values(
    source_path: Path,
    target_sheet,
    source_sheet_name: str | None = None,
    max_columns: int | None = None,
) -> None:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet_name = source_sheet_name or workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            values = trim_trailing_empty_cells(row)
            if max_columns is not None:
                values = values[:max_columns]
            target_sheet.append(values)
    finally:
        workbook.close()


def normalize_ots_number(value: Any) -> float | int | None:
    text = normalize_text(value).replace("%", "").replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if abs(number) <= 1:
        number *= 100
    number = round(number, 2)
    return int(number) if number.is_integer() else number


def extract_ots_rows_from_table(workbook_path: Path) -> list[tuple[str, str, float | int | None]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        table_sheet_name = "Table1" if "Table1" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[table_sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header_map = {normalize_header_key(value): index for index, value in enumerate(header_row)}

        market_index = header_map.get("MARKET")
        channel_index = header_map.get("CHANNEL", header_map.get("ATTRIBUTE"))
        ots_index = header_map.get("OTS", header_map.get("VALUE"))

        if market_index is not None and channel_index is not None and ots_index is not None:
            rows: list[tuple[str, str, float | int | None]] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                market = normalize_text(values[market_index] if market_index < len(values) else None)
                channel = normalize_text(values[channel_index] if channel_index < len(values) else None)
                if not market or not channel:
                    continue
                rows.append((market, channel, normalize_ots_number(values[ots_index] if ots_index < len(values) else None)))
            return rows

        wide_sheet_name = "Sheet1" if "Sheet1" in workbook.sheetnames else workbook.sheetnames[0]
        wide_sheet = workbook[wide_sheet_name]
        wide_header = next(wide_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        channels = [normalize_text(value) for value in wide_header[1:]]

        rows = []
        for values in wide_sheet.iter_rows(min_row=2, values_only=True):
            market = normalize_text(values[0] if values else None)
            if not market:
                continue
            for offset, channel in enumerate(channels, start=1):
                if not channel:
                    continue
                ots_value = normalize_ots_number(values[offset] if offset < len(values) else None)
                rows.append((market, channel, ots_value))
        return rows
    finally:
        workbook.close()


def write_ots_unpivot_sheet(source_path: Path, target_sheet) -> None:
    target_sheet.append(["Market", "Channel", "OTS"])
    for market, channel, ots_value in extract_ots_rows_from_table(source_path):
        target_sheet.append([market, channel, ots_value])


def output_needs_refresh(output_path: Path, source_paths: list[Path]) -> bool:
    if not output_path.exists():
        return True
    output_mtime = output_path.stat().st_mtime
    return any(path.stat().st_mtime > output_mtime for path in source_paths)


def build_combined_weekly_workbook(bundle: dict[str, Path], output_path: Path) -> Path:
    workbook = Workbook(write_only=True)
    distribution_sheet = workbook.create_sheet(DISTRIBUTION_SHEET_TITLE)
    copy_sheet_values(bundle["distribution"], distribution_sheet, max_columns=16)

    nbhd_sheet = workbook.create_sheet(NBHD_SHEET_TITLE)
    copy_sheet_values(bundle["nbhd"], nbhd_sheet, max_columns=8)

    ots_sheet = workbook.create_sheet(OTS_SHEET_TITLE)
    write_ots_unpivot_sheet(bundle["ots"], ots_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=output_path.parent) as handle:
        temp_path = Path(handle.name)
    try:
        workbook.save(temp_path)
        temp_path.replace(output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
    return output_path


def ensure_combined_weekly_workbooks(
    data_dir: Path,
    distribution_dir: Path,
    nbhd_dir: Path,
    ots_dir: Path,
) -> list[Path]:
    data_dir.mkdir(parents=True, exist_ok=True)
    generated: list[Path] = []
    bundles = discover_week_bundles(distribution_dir, nbhd_dir, ots_dir)

    for week_key, bundle in sorted(bundles.items()):
        output_path = data_dir / build_output_name(*week_key)
        source_paths = [bundle["distribution"], bundle["nbhd"], bundle["ots"]]
        if output_needs_refresh(output_path, source_paths):
            generated.append(build_combined_weekly_workbook(bundle, output_path))

    return generated
