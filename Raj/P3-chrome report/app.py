from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_JSON = OUTPUT_DIR / "frequency_report.json"
OUTPUT_HTML = OUTPUT_DIR / "chrome_report_dashboard.html"
STYLE_FILE = BASE_DIR / "static" / "style.css"

SOURCE_COLUMNS = [
    "WEEK LABEL",
    "TRANSMISSION",
    "MARKET",
    "GENRE",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "FREQUENCY/LCN NO",
    "BAND",
    "TV CH. No.",
    "AUDIO",
    "VIDEO",
    "LANGUAGE",
    "CRN No.",
    "RANK WITHIN GENRE",
]
KEY_COLUMNS = [
    "TRANSMISSION",
    "MARKET",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "CRN No.",
]
DISPLAY_COLUMNS = [
    "TRANSMISSION",
    "MARKET",
    "MSO TYPE",
    "CITY",
    "HEAD-END",
    "CHANNEL NAME",
    "WEEK LABEL",
    "GENRE",
    "LANGUAGE",
    "NAME",
]
FREQUENCY_COLUMN = "FREQUENCY/LCN NO"
RANK_COLUMN = "RANK WITHIN GENRE"

FOCUS_CHANNELS = {
    "INDIA TV": "India TV",
    "AAJ TAK": "Aaj Tak",
    "NEWS 18 INDIA": "News 18",
    "REPUBLIC BHARAT": "Republic Bharat",
}

REPORT_CACHE: dict[str, Any] = {
    "signature": None,
    "report": None,
}

app = Flask(__name__, template_folder="templates", static_folder="static")


def ensure_directories() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_number(value: Any) -> float | int | None:
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group())
    return int(number) if number.is_integer() else number


def normalize_rank(value: Any) -> int | None:
    number = normalize_number(value)
    if number is None:
        return None
    return int(number)


def week_sort_key(label: str) -> tuple[int, str]:
    match = re.search(r"(\d+)", label)
    if match:
        return int(match.group(1)), label
    return 10**9, label


def get_week_files() -> list[Path]:
    ensure_directories()
    return sorted(DATA_DIR.glob("Week*.xlsx"), key=lambda path: week_sort_key(path.stem))


def get_signature(files: list[Path]) -> tuple[tuple[str, int, int], ...]:
    return tuple((path.name, int(path.stat().st_mtime), path.stat().st_size) for path in files)


def empty_report(message: str | None = None) -> dict[str, Any]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": [],
        "records": [],
        "message": message or "Add weekly Excel files to the data folder and refresh the dashboard.",
    }


def prepare_week_rows(path: Path, fallback_label: str) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    week_label = fallback_label

    for row_index, values in enumerate(sheet.iter_rows(min_row=1, max_col=16, values_only=True), start=1):
        if row_index == 1:
            continue

        row = {
            SOURCE_COLUMNS[index]: values[index] if index < len(values) else None
            for index in range(len(SOURCE_COLUMNS))
        }

        if row_index == 2 and normalize_text(row["WEEK LABEL"]):
            week_label = normalize_text(row["WEEK LABEL"])

        normalized = {column: normalize_text(row.get(column)) for column in DISPLAY_COLUMNS if column != "NAME"}
        normalized["NAME"] = normalize_text(row.get("CHANNEL NAME"))
        normalized[FREQUENCY_COLUMN] = normalize_number(row.get(FREQUENCY_COLUMN))
        normalized[RANK_COLUMN] = normalize_rank(row.get(RANK_COLUMN))
        normalized["BAND"] = normalize_text(row.get("BAND"))
        normalized["TV CH. No."] = normalize_text(row.get("TV CH. No."))
        normalized["CRN No."] = normalize_text(row.get("CRN No."))
        normalized["ROW KEY"] = "||".join(normalized[column] for column in KEY_COLUMNS)

        if not normalized["ROW KEY"].replace("|", ""):
            continue

        rows.append(normalized)

    workbook.close()

    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        deduped.setdefault(row["ROW KEY"], row)
    return week_label, list(deduped.values())


def calculate_frequency_change(previous: float | int | None, current: float | int | None) -> str:
    if previous is None or current is None:
        return "missing"
    if current > previous:
        return "increase"
    if current < previous:
        return "decrease"
    return "no_change"


def calculate_rank_change(previous: int | None, current: int | None) -> str:
    if previous is None or current is None:
        return "missing"
    if current < previous:
        return "improve"
    if current > previous:
        return "decline"
    return "no_change"


def calculate_band_change(previous: str | None, current: str | None) -> str:
    previous_text = normalize_text(previous)
    current_text = normalize_text(current)
    if not previous_text or not current_text:
        return "missing"
    if current_text == previous_text:
        return "no_change"
    return "change"


def has_any_change(series: dict[str, Any], weeks: list[str]) -> bool:
    values = [series.get(week) for week in weeks if series.get(week) not in (None, "")]
    return len(values) > 1 and len(set(values)) > 1


def build_report() -> dict[str, Any]:
    week_files = get_week_files()
    if not week_files:
        report = empty_report()
        report["message"] = "Add weekly Excel files to the data folder to generate the report."
        return report

    weekly_data = [prepare_week_rows(path, path.stem) for path in week_files]
    weeks = [label for label, _ in weekly_data]

    merged: dict[str, dict[str, Any]] = {}
    for week_label, rows in weekly_data:
        for row in rows:
            record = merged.setdefault(
                row["ROW KEY"],
                {
                    "row_key": row["ROW KEY"],
                    "transmission": row["TRANSMISSION"],
                    "market": row["MARKET"],
                    "mso_type": row["MSO TYPE"],
                    "mso": row["TRANSMISSION"],
                    "city": row["CITY"],
                    "head_end": row["HEAD-END"],
                    "channel_name": row["CHANNEL NAME"],
                    "band": row["BAND"],
                    "tv_ch_no": row["TV CH. No."],
                    "crn_no": row["CRN No."],
                    "genre": row["GENRE"],
                    "language": row["LANGUAGE"],
                    "name": row["NAME"],
                    "week_label": row["WEEK LABEL"],
                    "frequencies": {},
                    "ranks": {},
                    "bands": {},
                    "changes": {},
                    "rank_changes": {},
                    "band_changes": {},
                },
            )
            record["frequencies"][week_label] = row[FREQUENCY_COLUMN]
            record["ranks"][week_label] = row[RANK_COLUMN]
            record["bands"][week_label] = row["BAND"]

    records = list(merged.values())

    for record in records:
        for week in weeks:
            record["frequencies"].setdefault(week, None)
            record["ranks"].setdefault(week, None)
            record["bands"].setdefault(week, "")

        if weeks:
            first_week = weeks[0]
            record["changes"][first_week] = "baseline"
            record["rank_changes"][first_week] = "baseline"
            record["band_changes"][first_week] = "baseline"

        for index in range(1, len(weeks)):
            previous_week = weeks[index - 1]
            current_week = weeks[index]
            record["changes"][current_week] = calculate_frequency_change(
                record["frequencies"][previous_week],
                record["frequencies"][current_week],
            )
            record["rank_changes"][current_week] = calculate_rank_change(
                record["ranks"][previous_week],
                record["ranks"][current_week],
            )
            record["band_changes"][current_week] = calculate_band_change(
                record["bands"][previous_week],
                record["bands"][current_week],
            )

        record["change_status"] = "YES" if has_any_change(record["frequencies"], weeks) else "NO"
        record["rank_change_status"] = "YES" if has_any_change(record["ranks"], weeks) else "NO"
        record["band_change_status"] = "YES" if has_any_change(record["bands"], weeks) else "NO"

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "weeks": weeks,
        "records": records,
        "message": "",
    }
    OUTPUT_JSON.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report


def load_report(force: bool = False) -> dict[str, Any]:
    week_files = get_week_files()
    signature = get_signature(week_files)
    if not force and REPORT_CACHE["report"] is not None and REPORT_CACHE["signature"] == signature:
        return REPORT_CACHE["report"]

    report = build_report()
    REPORT_CACHE["signature"] = signature
    REPORT_CACHE["report"] = report
    return report


def get_view_config(view: str) -> dict[str, str]:
    if view == "rank":
        return {
            "series": "ranks",
            "changes": "rank_changes",
            "status": "rank_change_status",
            "positive": "improve",
            "negative": "decline",
            "positive_label": "improved",
            "negative_label": "declined",
        }
    if view == "band":
        return {
            "series": "bands",
            "changes": "band_changes",
            "status": "band_change_status",
            "positive": "change",
            "negative": "no_change",
            "positive_label": "changed",
            "negative_label": "stable",
        }
    return {
        "series": "frequencies",
        "changes": "changes",
        "status": "change_status",
        "positive": "increase",
        "negative": "decrease",
        "positive_label": "increased",
        "negative_label": "decreased",
    }


def filter_records(records: list[dict[str, Any]], view: str, filters: dict[str, str], ignore_key: str = "") -> list[dict[str, Any]]:
    config = get_view_config(view)
    changed_states = {"frequency": {"increase", "decrease"}, "rank": {"improve", "decline"}, "band": {"change"}}
    changed_set = changed_states.get(view, changed_states["frequency"])

    filtered: list[dict[str, Any]] = []
    for record in records:
        if filters["market"] and record["market"] != filters["market"] and ignore_key != "market":
            continue
        if filters["city"] and record["city"] != filters["city"] and ignore_key != "city":
            continue
        if filters["mso_type"] and record["mso_type"] != filters["mso_type"] and ignore_key != "mso_type":
            continue
        if filters["mso"] and record["mso"] != filters["mso"] and ignore_key != "mso":
            continue
        if filters["head_end"] and record["head_end"] != filters["head_end"] and ignore_key != "head_end":
            continue
        if filters["crn_no"] and record["crn_no"] != filters["crn_no"] and ignore_key != "crn_no":
            continue
        if filters["channel_name"] and record["channel_name"] != filters["channel_name"] and ignore_key != "channel_name":
            continue
        if filters["band"] and record["band"] != filters["band"] and ignore_key != "band":
            continue

        if ignore_key != "week" and filters["week"]:
            value = record[config["series"]].get(filters["week"])
            if value in (None, ""):
                continue

        if ignore_key != "change" and filters["change"]:
            if filters["change"] == "Changed":
                if filters["week"]:
                    if record[config["changes"]].get(filters["week"]) not in changed_set:
                        continue
                elif record[config["status"]] != "YES":
                    continue
            elif filters["change"] == "No Change":
                if filters["week"]:
                    if record[config["changes"]].get(filters["week"]) != "no_change":
                        continue
                elif record[config["status"]] != "NO":
                    continue

        filtered.append(record)

    return filtered


def sort_value(record: dict[str, Any], sort_key: str, view: str) -> Any:
    if sort_key in record:
        value = record[sort_key]
    elif sort_key in record.get("frequencies", {}):
        config = get_view_config(view)
        value = record[config["series"]].get(sort_key)
    else:
        value = ""

    if value is None:
        return (1, "")
    if isinstance(value, (int, float)):
        return (0, value)
    return (0, str(value).lower())


def sort_records(records: list[dict[str, Any]], sort_key: str, sort_direction: str, view: str) -> list[dict[str, Any]]:
    reverse = sort_direction == "desc"
    return sorted(records, key=lambda record: sort_value(record, sort_key, view), reverse=reverse)


def paginate_records(records: list[dict[str, Any]], page: int, page_size: int) -> tuple[list[dict[str, Any]], int]:
    total_count = len(records)
    if total_count == 0:
        return [], 0
    start = max(page - 1, 0) * page_size
    end = start + page_size
    return records[start:end], total_count


def build_filters(records: list[dict[str, Any]], view: str, current_filters: dict[str, str], weeks: list[str]) -> dict[str, list[str]]:
    def values_for(key: str, field: str) -> list[str]:
        values = {
            normalize_text(record.get(field))
            for record in filter_records(records, view, current_filters, ignore_key=key)
            if normalize_text(record.get(field))
        }
        return sorted(values, key=lambda value: value.lower())

    return {
        "markets": values_for("market", "market"),
        "cities": values_for("city", "city"),
        "mso_types": values_for("mso_type", "mso_type"),
        "msos": values_for("mso", "mso"),
        "head_ends": values_for("head_end", "head_end"),
        "crn_numbers": values_for("crn_no", "crn_no"),
        "channels": values_for("channel_name", "channel_name"),
        "bands": values_for("band", "band"),
        "weeks": weeks,
        "change_options": ["Changed", "No Change"],
    }


def summarize_records(records: list[dict[str, Any]], view: str, weeks: list[str]) -> dict[str, int]:
    config = get_view_config(view)
    summary = {
        "total_channels": len(records),
    }

    positive = 0
    negative = 0
    stable = 0

    if view == "band":
        for record in records:
            if record[config["status"]] == "YES":
                positive += 1
            else:
                stable += 1
        summary["changed"] = positive
        summary["stable"] = stable
        return summary

    final_week = weeks[-1] if weeks else ""
    for record in records:
        status = record[config["changes"]].get(final_week, "no_change")
        if status == config["positive"]:
            positive += 1
        elif status == config["negative"]:
            negative += 1
        else:
            stable += 1

    if view == "rank":
        summary["improved"] = positive
        summary["declined"] = negative
        summary["no_change"] = stable
    else:
        summary["increased"] = positive
        summary["decreased"] = negative
        summary["no_change"] = stable
    return summary


def summarize_focus_channels(records: list[dict[str, Any]], view: str, weeks: list[str]) -> list[dict[str, Any]]:
    config = get_view_config(view)
    items: list[dict[str, Any]] = []

    for match_name, label in FOCUS_CHANNELS.items():
        selected = [record for record in records if normalize_text(record["channel_name"]).upper() == match_name]
        if not selected:
            continue

        positive = 0
        negative = 0
        no_change = 0
        latest_positive = 0
        latest_negative = 0
        latest_week = weeks[-1] if weeks else ""

        for record in selected:
            if not weeks:
                continue
            latest_status = record[config["changes"]].get(latest_week)
            if latest_status == config["positive"]:
                latest_positive += 1
            elif latest_status == config["negative"]:
                latest_negative += 1

            for week in weeks[1:]:
                status = record[config["changes"]].get(week)
                if status == config["positive"]:
                    positive += 1
                elif status == config["negative"]:
                    negative += 1
                elif status == "no_change":
                    no_change += 1

        items.append(
            {
                "label": label,
                "records": len(selected),
                "positive": positive,
                "negative": negative,
                "no_change": no_change,
                "latest_positive": latest_positive,
                "latest_negative": latest_negative,
                "latest_week": latest_week,
                "positive_label": config["positive_label"],
                "negative_label": config["negative_label"],
            }
        )

    return items


def serialize_records(records: list[dict[str, Any]], weeks: list[str]) -> list[dict[str, Any]]:
    serialized: list[dict[str, Any]] = []
    for record in records:
        serialized.append(
            {
                "market": record["market"],
                "mso_type": record["mso_type"],
                "mso": record["mso"],
                "city": record["city"],
                "head_end": record["head_end"],
                "channel_name": record["channel_name"],
                "band": record["band"],
                "tv_ch_no": record["tv_ch_no"],
                "crn_no": record["crn_no"],
                "name": record["name"],
                "frequencies": {week: record["frequencies"].get(week) for week in weeks},
                "ranks": {week: record["ranks"].get(week) for week in weeks},
                "bands": {week: record["bands"].get(week) for week in weeks},
                "changes": {week: record["changes"].get(week, "missing") for week in weeks},
                "rank_changes": {week: record["rank_changes"].get(week, "missing") for week in weeks},
                "band_changes": {week: record["band_changes"].get(week, "missing") for week in weeks},
                "change_status": record["change_status"],
                "rank_change_status": record["rank_change_status"],
                "band_change_status": record["band_change_status"],
            }
        )
    return serialized


def read_style() -> str:
    if STYLE_FILE.exists():
        return STYLE_FILE.read_text(encoding="utf-8")
    return "body { font-family: sans-serif; }"


def create_standalone_dashboard(report: dict[str, Any]) -> str:
    style_text = read_style()
    report_json = json.dumps(report).replace("</", "<\\/")

    html = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Chrome Report</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  <style>
__STYLE__
.share-note { margin-top: 8px; font-size: 0.72rem; color: var(--muted); }
  </style>
</head>
<body>
  <div class="app-shell">
    <header class="hero">
      <div class="hero-copy">
        <h1>Chrome Report</h1>
      </div>
      <div class="hero-meta">
        <article class="meta-card">
          <span>Generated</span>
          <strong id="generatedAt">--</strong>
        </article>
        <article class="meta-card">
          <span>Total Records</span>
          <strong id="totalRecords">0</strong>
        </article>
      </div>
    </header>

    <section class="panel filter-panel">
      <div class="panel-heading"><div><h2>Filter Panel</h2></div></div>
      <div class="filter-grid">
        <label><span>Market</span><select id="marketFilter"></select></label>
        <label><span>City</span><select id="cityFilter"></select></label>
        <label><span>MSO Type</span><select id="msoTypeFilter"></select></label>
        <label><span>MSO</span><select id="msoFilter"></select></label>
        <label><span>Headend</span><select id="headendFilter"></select></label>
        <label><span>CRN No</span><select id="crnFilter"></select></label>
        <label><span>Channel</span><select id="channelFilter"></select></label>
        <label><span>Band</span><select id="bandFilter"></select></label>
        <label><span>Week</span><select id="weekFilter"></select></label>
        <label><span>Change</span><select id="changeFilter"></select></label>
        <div class="action-row"><button id="resetButton" class="ghost-button" type="button">Reset Filters</button></div>
      </div>
      <div class="share-note">Standalone dashboard file. Share this HTML as-is. Recipients do not need the source Excel files.</div>
    </section>

    <section class="panel table-panel">
      <div class="panel-heading table-heading">
        <div><h2 id="tableTitle">Weekly Frequency Analysis</h2></div>
        <div class="table-side">
          <div class="view-switcher">
            <button id="frequencyViewButton" class="switch-button active" type="button">Frequency</button>
            <button id="rankViewButton" class="switch-button" type="button">Rank</button>
            <button id="bandViewButton" class="switch-button" type="button">Band</button>
          </div>
          <div class="table-meta">
            <span id="resultCount">0 records</span>
            <span id="pageInfo">Page 1</span>
          </div>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead id="tableHead"></thead>
          <tbody id="tableBody"></tbody>
        </table>
      </div>
      <div class="pagination-bar">
        <button id="prevPage" class="ghost-button" type="button">Previous</button>
        <button id="nextPage" class="ghost-button" type="button">Next</button>
      </div>
    </section>

    <section class="panel focus-panel">
      <div class="panel-heading"><div><h2>Channel Summary</h2></div></div>
      <div id="focusSummary" class="focus-summary"></div>
    </section>

    <section class="kpi-grid bottom-kpis">
      <article class="kpi-card compact-kpi">
        <span>Total Channels</span>
        <strong id="kpiTotal">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span id="kpiLabelOne">Frequency Increased</span>
        <strong id="kpiIncrease">0</strong>
      </article>
      <article class="kpi-card compact-kpi">
        <span id="kpiLabelTwo">Frequency Decreased</span>
        <strong id="kpiDecrease">0</strong>
      </article>
    </section>
  </div>

  <template id="emptyStateTemplate">
    <tr><td colspan="100%" class="empty-state">No records match the current filters.</td></tr>
  </template>

  <script>
const report = __DATA__;
const state = {
  view: "frequency",
  filters: { market: "", city: "", mso_type: "", mso: "", head_end: "", crn_no: "", channel_name: "", band: "", week: "", change: "" },
  sortKey: "channel_name",
  sortDirection: "asc",
  page: 1,
  pageSize: 30,
};
const tableColumns = [
  { key: "market", label: "MARKET" },
  { key: "mso_type", label: "MSO TYPE" },
  { key: "city", label: "CITY" },
  { key: "head_end", label: "HEAD-END" },
  { key: "channel_name", label: "CHANNEL NAME" },
  { key: "band", label: "BAND" },
  { key: "tv_ch_no", label: "TV CH. No." },
  { key: "crn_no", label: "CRN No." },
  { key: "name", label: "NAME" },
];
const filterOrder = ["market", "city", "mso_type", "mso", "head_end", "crn_no", "channel_name", "band", "week", "change"];
const fieldMap = { market: "market", city: "city", mso_type: "mso_type", mso: "mso", head_end: "head_end", crn_no: "crn_no", channel_name: "channel_name", band: "band" };
const filters = {
  market: document.getElementById("marketFilter"),
  city: document.getElementById("cityFilter"),
  mso_type: document.getElementById("msoTypeFilter"),
  mso: document.getElementById("msoFilter"),
  head_end: document.getElementById("headendFilter"),
  crn_no: document.getElementById("crnFilter"),
  channel_name: document.getElementById("channelFilter"),
  band: document.getElementById("bandFilter"),
  week: document.getElementById("weekFilter"),
  change: document.getElementById("changeFilter"),
};
const viewButtons = {
  frequency: document.getElementById("frequencyViewButton"),
  rank: document.getElementById("rankViewButton"),
  band: document.getElementById("bandViewButton"),
};
function formatNumber(value) { return new Intl.NumberFormat().format(value || 0); }
function formatTimestamp(value) {
  const date = new Date(value);
  return Number.isNaN(date.getTime()) ? "--" : date.toLocaleString("en-IN", { year: "numeric", month: "2-digit", day: "2-digit", hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false });
}
function createOption(value, label) {
  const option = document.createElement("option");
  option.value = value;
  option.textContent = label;
  return option;
}
function getViewConfig() {
  if (state.view === "rank") return { series: "ranks", changes: "rank_changes", status: "rank_change_status", positive: "improve", negative: "decline", kpiOne: "Rank Improved", kpiTwo: "Rank Declined", title: "Weekly Rank Analysis" };
  if (state.view === "band") return { series: "bands", changes: "band_changes", status: "band_change_status", positive: "change", negative: "no_change", kpiOne: "Band Changed", kpiTwo: "Band Stable", title: "Weekly Band Analysis" };
  return { series: "frequencies", changes: "changes", status: "change_status", positive: "increase", negative: "decrease", kpiOne: "Frequency Increased", kpiTwo: "Frequency Decreased", title: "Weekly Frequency Analysis" };
}
function filterRecords(ignoreKey = "") {
  const viewConfig = getViewConfig();
  return report.records.filter((record) => {
    for (const [key, field] of Object.entries(fieldMap)) {
      if (key === ignoreKey) continue;
      if (state.filters[key] && String(record[field] || "") !== state.filters[key]) return false;
    }
    if (ignoreKey !== "change" && state.filters.change) {
      if (state.filters.change === "Changed") {
        if (state.filters.week) {
          const changedSet = state.view === "band" ? new Set(["change"]) : new Set(["increase", "decrease", "improve", "decline"]);
          if (!changedSet.has(record[viewConfig.changes][state.filters.week])) return false;
        } else if (record[viewConfig.status] !== "YES") {
          return false;
        }
      }
      if (state.filters.change === "No Change") {
        if (state.filters.week) {
          if (record[viewConfig.changes][state.filters.week] !== "no_change") return false;
        } else if (record[viewConfig.status] !== "NO") {
          return false;
        }
      }
    }
    if (ignoreKey !== "week" && state.filters.week && !record[viewConfig.series][state.filters.week] && record[viewConfig.series][state.filters.week] !== 0) return false;
    return true;
  });
}
function getOptions(key) {
  if (key === "week") return report.weeks.slice();
  if (key === "change") return ["Changed", "No Change"];
  const field = fieldMap[key];
  const values = new Set();
  filterRecords(key).forEach((record) => {
    const value = String(record[field] || "").trim();
    if (value) values.add(value);
  });
  return Array.from(values).sort((a, b) => a.localeCompare(b));
}
function populateSelect(select, values, allLabel, selectedValue) {
  const safeValues = values.filter((value) => value !== null && value !== undefined && String(value).trim() !== "");
  const safeSelectedValue = safeValues.includes(selectedValue) ? selectedValue : "";
  select.innerHTML = "";
  select.appendChild(createOption("", allLabel));
  safeValues.forEach((value) => select.appendChild(createOption(value, value)));
  select.value = safeSelectedValue;
  return safeSelectedValue;
}
function syncFilters() {
  state.filters.market = populateSelect(filters.market, getOptions("market"), "All Markets", state.filters.market);
  state.filters.city = populateSelect(filters.city, getOptions("city"), "All Cities", state.filters.city);
  state.filters.mso_type = populateSelect(filters.mso_type, getOptions("mso_type"), "All MSO Types", state.filters.mso_type);
  state.filters.mso = populateSelect(filters.mso, getOptions("mso"), "All MSO", state.filters.mso);
  state.filters.head_end = populateSelect(filters.head_end, getOptions("head_end"), "All Headend", state.filters.head_end);
  state.filters.crn_no = populateSelect(filters.crn_no, getOptions("crn_no"), "All CRN No", state.filters.crn_no);
  state.filters.channel_name = populateSelect(filters.channel_name, getOptions("channel_name"), "All Channels", state.filters.channel_name);
  state.filters.band = populateSelect(filters.band, getOptions("band"), "All Bands", state.filters.band);
  state.filters.week = populateSelect(filters.week, getOptions("week"), "All Weeks", state.filters.week);
  state.filters.change = populateSelect(filters.change, getOptions("change"), "All Changes", state.filters.change);
}
function sortValue(record, sortKey) {
  const viewConfig = getViewConfig();
  let value = record[sortKey];
  if (value === undefined && report.weeks.includes(sortKey)) value = record[viewConfig.series][sortKey];
  if (value === null || value === undefined || value === "") return [1, ""];
  if (typeof value === "number") return [0, value];
  return [0, String(value).toLowerCase()];
}
function getFilteredRecords() {
  const items = filterRecords();
  const sorted = items.slice().sort((a, b) => {
    const left = sortValue(a, state.sortKey);
    const right = sortValue(b, state.sortKey);
    if (left[0] !== right[0]) return left[0] - right[0];
    if (left[1] < right[1]) return state.sortDirection === "asc" ? -1 : 1;
    if (left[1] > right[1]) return state.sortDirection === "asc" ? 1 : -1;
    return 0;
  });
  return sorted;
}
function buildTableHead() {
  const tableHead = document.getElementById("tableHead");
  const tr = document.createElement("tr");
  [...tableColumns, ...report.weeks.map((week) => ({ key: week, label: week })), { key: "change_status", label: "CHANGE" }].forEach((column) => {
    const th = document.createElement("th");
    const isActive = state.sortKey === column.key;
    const suffix = isActive ? (state.sortDirection === "asc" ? " ▲" : " ▼") : "";
    th.textContent = `${column.label}${suffix}`;
    th.className = "sortable";
    th.addEventListener("click", () => {
      if (state.sortKey === column.key) state.sortDirection = state.sortDirection === "asc" ? "desc" : "asc";
      else { state.sortKey = column.key; state.sortDirection = "asc"; }
      render();
    });
    tr.appendChild(th);
  });
  tableHead.replaceChildren(tr);
}
function formatWeekValue(value, status, isBaseline) {
  if (value === null || value === undefined || value === "") return "-";
  if (isBaseline || status === "baseline" || status === "missing" || status === "no_change") return String(value);
  if (state.view === "rank") {
    if (status === "improve") return `▲ ${value}`;
    if (status === "decline") return `▼ ${value}`;
    return String(value);
  }
  if (state.view === "band") {
    if (status === "change") return `• ${value}`;
    return String(value);
  }
  if (status === "increase") return `▲ ${value}`;
  if (status === "decrease") return `▼ ${value}`;
  return String(value);
}
function renderFocusSummary(records) {
  const container = document.getElementById("focusSummary");
  const labels = { "INDIA TV": "India TV", "AAJ TAK": "Aaj Tak", "NEWS 18 INDIA": "News 18", "REPUBLIC BHARAT": "Republic Bharat" };
  const viewConfig = getViewConfig();
  const items = Object.entries(labels).map(([channel, label]) => {
    const selected = records.filter((record) => String(record.channel_name || "").toUpperCase() === channel);
    if (!selected.length) return "";
    let positive = 0, negative = 0, noChange = 0, latestPositive = 0, latestNegative = 0;
    const latestWeek = report.weeks[report.weeks.length - 1];
    selected.forEach((record) => {
      report.weeks.slice(1).forEach((week) => {
        const status = record[viewConfig.changes][week];
        if (status === viewConfig.positive) positive += 1;
        else if (status === viewConfig.negative) negative += 1;
        else if (status === "no_change") noChange += 1;
      });
      const latestStatus = record[viewConfig.changes][latestWeek];
      if (latestStatus === viewConfig.positive) latestPositive += 1;
      else if (latestStatus === viewConfig.negative) latestNegative += 1;
    });
    const positiveLabel = state.view === "rank" ? "improved" : state.view === "band" ? "changed" : "increased";
    const negativeLabel = state.view === "rank" ? "declined" : state.view === "band" ? "stable" : "decreased";
    const latestText = latestPositive || latestNegative ? ` Latest: ${latestPositive ? `${formatNumber(latestPositive)} ${positiveLabel}` : ""}${latestPositive && latestNegative ? ", " : ""}${latestNegative ? `${formatNumber(latestNegative)} ${negativeLabel}` : ""} in ${latestWeek}.` : "";
    return `<div class="focus-line"><strong>${label}</strong><span>${formatNumber(selected.length)} rows, ${formatNumber(positive)} ${positiveLabel}, ${formatNumber(negative)} ${negativeLabel}, ${formatNumber(noChange)} stable.${latestText}</span></div>`;
  }).filter(Boolean).join("");
  container.innerHTML = items || '<div class="focus-line">No channel summary available for the current filters.</div>';
}
function renderTable(records) {
  const tableBody = document.getElementById("tableBody");
  const start = (state.page - 1) * state.pageSize;
  const pageItems = records.slice(start, start + state.pageSize);
  if (!pageItems.length) {
    tableBody.replaceChildren(document.getElementById("emptyStateTemplate").content.cloneNode(true));
    return;
  }
  const viewConfig = getViewConfig();
  const fragment = document.createDocumentFragment();
  pageItems.forEach((record) => {
    const tr = document.createElement("tr");
    tableColumns.forEach((column) => {
      const td = document.createElement("td");
      td.textContent = record[column.key] ?? "";
      tr.appendChild(td);
    });
    report.weeks.forEach((week, index) => {
      const td = document.createElement("td");
      const status = index === 0 ? "missing" : record[viewConfig.changes][week];
      td.classList.add(`status-${status}`);
      td.textContent = formatWeekValue(record[viewConfig.series][week], status, index === 0);
      tr.appendChild(td);
    });
    const changeTd = document.createElement("td");
    changeTd.textContent = record[viewConfig.status];
    changeTd.classList.add(record[viewConfig.status] === "NO" ? "change-no" : "change-yes");
    tr.appendChild(changeTd);
    fragment.appendChild(tr);
  });
  tableBody.replaceChildren(fragment);
}
function updateKpis(records) {
  document.getElementById("kpiTotal").textContent = formatNumber(records.length);
  const viewConfig = getViewConfig();
  let positive = 0, negative = 0;
  const latestWeek = report.weeks[report.weeks.length - 1];
  records.forEach((record) => {
    const status = record[viewConfig.changes][latestWeek];
    if (status === viewConfig.positive) positive += 1;
    if (status === viewConfig.negative) negative += 1;
  });
  document.getElementById("kpiLabelOne").textContent = viewConfig.kpiOne;
  document.getElementById("kpiLabelTwo").textContent = viewConfig.kpiTwo;
  document.getElementById("kpiIncrease").textContent = formatNumber(positive);
  document.getElementById("kpiDecrease").textContent = formatNumber(negative);
}
function render() {
  syncFilters();
  const records = getFilteredRecords();
  const totalPages = Math.max(1, Math.ceil(records.length / state.pageSize));
  if (state.page > totalPages) state.page = totalPages;
  document.getElementById("generatedAt").textContent = formatTimestamp(report.generated_at);
  document.getElementById("totalRecords").textContent = formatNumber(records.length);
  document.getElementById("resultCount").textContent = `${formatNumber(records.length)} records`;
  document.getElementById("pageInfo").textContent = `Page ${state.page} of ${totalPages}`;
  document.getElementById("tableTitle").textContent = getViewConfig().title;
  Object.entries(viewButtons).forEach(([view, button]) => button.classList.toggle("active", view === state.view));
  buildTableHead();
  renderTable(records);
  renderFocusSummary(records);
  updateKpis(records);
}
Object.entries(filters).forEach(([key, select]) => {
  select.addEventListener("change", () => {
    state.filters[key] = select.value;
    const changedIndex = filterOrder.indexOf(key);
    if (changedIndex >= 0) filterOrder.slice(changedIndex + 1).forEach((nextKey) => { state.filters[nextKey] = ""; });
    state.page = 1;
    render();
  });
});
document.getElementById("resetButton").addEventListener("click", () => {
  state.filters = { market: "", city: "", mso_type: "", mso: "", head_end: "", crn_no: "", channel_name: "", band: "", week: "", change: "" };
  state.sortKey = "channel_name";
  state.sortDirection = "asc";
  state.page = 1;
  render();
});
document.getElementById("prevPage").addEventListener("click", () => { if (state.page > 1) { state.page -= 1; render(); } });
document.getElementById("nextPage").addEventListener("click", () => {
  const totalPages = Math.max(1, Math.ceil(getFilteredRecords().length / state.pageSize));
  if (state.page < totalPages) { state.page += 1; render(); }
});
Object.entries(viewButtons).forEach(([view, button]) => {
  button.addEventListener("click", () => {
    state.view = view;
    state.page = 1;
    render();
  });
});
render();
  </script>
</body>
</html>
"""

    return html.replace("__STYLE__", style_text).replace("__DATA__", report_json)


def build_api_payload(view: str, filters: dict[str, str], page: int, page_size: int, sort_key: str, sort_direction: str, force_refresh: bool = False) -> dict[str, Any]:
    report = load_report(force=force_refresh)
    weeks = report.get("weeks", [])
    records = report.get("records", [])
    filtered = filter_records(records, view, filters)
    sorted_records = sort_records(filtered, sort_key, sort_direction, view)
    page_records, total_count = paginate_records(sorted_records, page, page_size)
    total_pages = max(1, (total_count + page_size - 1) // page_size) if total_count else 1

    return {
        "generated_at": report.get("generated_at"),
        "view": view,
        "weeks": weeks,
        "filters": build_filters(records, view, filters, weeks),
        "summary": summarize_records(filtered, view, weeks),
        "focus_channels": summarize_focus_channels(filtered, view, weeks),
        "message": report.get("message", ""),
        "data_directory": str(DATA_DIR),
        "table": {
            "records": serialize_records(page_records, weeks),
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": total_pages,
            "sort_key": sort_key,
            "sort_direction": sort_direction,
        },
    }


@app.get("/")
def index() -> str:
    return render_template("index.html")


@app.get("/api/frequency")
def api_frequency():
    view = request.args.get("view", "frequency").strip().lower()
    if view not in {"frequency", "rank", "band"}:
        view = "frequency"

    filters = {
        "market": request.args.get("market", "").strip(),
        "city": request.args.get("city", "").strip(),
        "mso_type": request.args.get("mso_type", "").strip(),
        "mso": request.args.get("mso", "").strip(),
        "head_end": request.args.get("head_end", "").strip(),
        "crn_no": request.args.get("crn_no", "").strip(),
        "channel_name": request.args.get("channel_name", "").strip(),
        "band": request.args.get("band", "").strip(),
        "week": request.args.get("week", "").strip(),
        "change": request.args.get("change", "").strip(),
    }
    page = max(1, int(request.args.get("page", "1") or "1"))
    page_size = max(1, min(200, int(request.args.get("page_size", "30") or "30")))
    sort_key = request.args.get("sort_key", "channel_name").strip() or "channel_name"
    sort_direction = request.args.get("sort_direction", "asc").strip().lower()
    if sort_direction not in {"asc", "desc"}:
        sort_direction = "asc"
    force_refresh = request.args.get("refresh", "").strip() == "1"

    return jsonify(build_api_payload(view, filters, page, page_size, sort_key, sort_direction, force_refresh))


@app.get("/download/dashboard")
def download_dashboard():
    report = load_report(force=True)
    standalone_html = create_standalone_dashboard(report)
    OUTPUT_HTML.write_text(standalone_html, encoding="utf-8")
    return send_file(
        OUTPUT_HTML,
        as_attachment=True,
        download_name="chrome_report_dashboard.html",
        mimetype="text/html",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "9001"))
    app.run(host="0.0.0.0", port=port, debug=False)
