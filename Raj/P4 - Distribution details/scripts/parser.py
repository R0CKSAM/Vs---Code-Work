from __future__ import annotations

import logging
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from headend_id import generate_headend_id

LOGGER = logging.getLogger(__name__)

LABEL_ALIASES = {
    "date": "date",
    "dated": "date",
    "networkname": "network_name",
    "headendlocation": "headend_location",
    "headend": "headend_location",
    "state": "state",
    "barcmarket": "barc_market",
    "stb": "stbs",
    "stbs": "stbs",
    "landingchannel": "landing_channel_1",
    "1stlandingchannel": "landing_channel_1",
    "secondlandingchannel": "landing_channel_2",
    "2ndlandingchannel": "landing_channel_2",
    "barkerchannel": "barker_1",
    "barker": "barker_1",
    "secondbarkerchannel": "barker_2",
    "2ndbarkerchannel": "barker_2",
    "2ndbarker": "barker_2",
}

NULL_LIKE_VALUES = {"", "na", "n/a", "none", "null", "-"}
UNMAPPED_REVIEW = "Unmapped - Review"


@dataclass(slots=True)
class ParsedWorkbook:
    distribution_rows: list[dict[str, Any]]
    channel_rows: list[dict[str, Any]]
    week_dates: list[str]
    normalization_warnings: dict[str, dict[str, int]]


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text or text.lower() in NULL_LIKE_VALUES:
        return None
    return text


def clean_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)

    text = clean_text(value)
    if text is None:
        return None

    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else None


def clean_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)
    if text is None:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%y", "%d-%B-%y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def normalize_lookup_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[._]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def smart_title_case(value: str | None) -> str | None:
    if value is None:
        return None

    def format_token(token: str) -> str:
        if not token:
            return token
        if token.isdigit():
            return token
        if len(token) <= 3 and token.isupper():
            return token
        return token[:1].upper() + token[1:].lower()

    pieces = re.split(r"(\s+|/|-|,|\(|\)|&)", value.strip())
    return "".join(format_token(piece) if re.search(r"[A-Za-z]", piece) else piece for piece in pieces).strip()


def build_normalization_map(raw_mapping: dict[str, str]) -> dict[str, str]:
    return {normalize_lookup_key(raw): canonical for raw, canonical in raw_mapping.items()}


STATE_NORMALIZATION = build_normalization_map(
    {
        "AGARTALA": "Tripura",
        "All India": "All India",
        "ANDHRA PRADESH": "Andhra Pradesh",
        "AP": "Andhra Pradesh",
        "Asaam": "Assam",
        "Assam": "Assam",
        "Bihar": "Bihar",
        "Chhattisdarh": "Chhattisgarh",
        "Chhattisgarh": "Chhattisgarh",
        "Delhi": "Delhi",
        "Goa": "Goa",
        "GUJARAT": "Gujarat",
        "Haryana": "Haryana",
        "Himachal Pradesh": "Himachal Pradesh",
        "Jammu & Kashmir": "Jammu & Kashmir",
        "Jharkhand": "Jharkhand",
        "Karnataka": "Karnataka",
        "M.P.": "Madhya Pradesh",
        "Maharashtra": "Maharashtra",
        "MAHARASHTRA": "Maharashtra",
        "MANIPUR": "Manipur",
        "Manipur": "Manipur",
        "MIZORAM": "Mizoram",
        "Mizoram": "Mizoram",
        "Nagaland": "Nagaland",
        "Odisha": "Odisha",
        "Punjab": "Punjab",
        "Rajasthan": "Rajasthan",
        "TELANGANA": "Telangana",
        "Telangana": "Telangana",
        "U.P": "Uttar Pradesh",
        "U.P.": "Uttar Pradesh",
        "UP": "Uttar Pradesh",
        "Up": "Uttar Pradesh",
        "up": "Uttar Pradesh",
        "Uk": "Uttarakhand",
        "West Bengal": "West Bengal",
    }
)

STATE_CONTEXTUAL_KEYS = {
    normalize_lookup_key(value)
    for value in {
        "Bihar/JHR",
        "District -Uttara Kannada",
        "District Bagalkot",
        "District Davanagere",
        "District Raichur",
        "District Tumkur",
        "District tumkur",
        "Ghatsila",
    }
}

SHEET_STATE_NORMALIZATION = build_normalization_map(
    {
        "Delhi": "Delhi",
        "U.P.": "Uttar Pradesh",
        "U.K.": "Uttarakhand",
        "Punjab": "Punjab",
        "Haryana": "Haryana",
        "H.P.": "Himachal Pradesh",
        "J&K": "Jammu & Kashmir",
        "Rajasthan": "Rajasthan",
        "Bihar": "Bihar",
        "Jharkhand": "Jharkhand",
        "W.B.": "West Bengal",
        "Odisha": "Odisha",
        "MP": "Madhya Pradesh",
        "CG": "Chhattisgarh",
        "Gujarat": "Gujarat",
        "A.P.": "Andhra Pradesh",
        "Telangana": "Telangana",
        "Karnataka": "Karnataka",
    }
)

CITY_NORMALIZATION = build_normalization_map(
    {
        "Ahemdabad": "Ahmedabad",
        "AHMEDABAD": "Ahmedabad",
        "BARODA": "Baroda",
        "Baroda": "Baroda",
        "HYDERABAD": "Hyderabad",
        "Hyderabad": "Hyderabad",
        "KAWARDHA": "Kawardha",
        "Kawardha": "Kawardha",
        "MUMBAI": "Mumbai",
        "Mumbai": "Mumbai",
        "RAJKOT": "Rajkot",
        "Rajkot": "Rajkot",
        "Uttar Kashi": "Uttarkashi",
        "UttarKashi": "Uttarkashi",
        "VIKARABAD": "Vikarabad",
        "Vikarabad": "Vikarabad",
        "WARANGAL": "Warangal",
        "Warangal": "Warangal",
        "AIZAWL- MIZORAM": "Aizawl",
    }
)

BARC_MARKET_NORMALIZATION = build_normalization_map(
    {
        "10-75/Urban/Rural": "10-75L / Urban / Rural",
        "10-75L": "10-75L",
        "10-75L & RURAL": "10-75L & Rural",
        "10-75Lacs": "10-75L",
        "10-75LACS & RURAL": "10-75L & Rural",
        "All India": "All India",
        "All Market": "All Market",
        "ALL MARKET": "All Market",
        "Ap / Telangana - Rural": "AP / Telangana - Rural",
        "Ap/Telangana - Below 75L Urban": "AP / Telangana - Below 75L Urban",
        "Ap/Telangana - Urban": "AP / Telangana - Urban",
        "ASSAM/NORTHEAST/SIKKIM": "Assam / Northeast / Sikkim",
        "ASSAM/NORTHEAST/SIKKIM- RURAL": "Assam / Northeast / Sikkim - Rural",
        "ASSAM/NORTHEAST/SIKKIM- URBAN": "Assam / Northeast / Sikkim - Urban",
        "ASSAM/NORTHEAST/SIKKIM-RURAL": "Assam / Northeast / Sikkim - Rural",
        "ASSAM/NORTHEAST/SIKKIM-URBAN": "Assam / Northeast / Sikkim - Urban",
        "Bangalore": "Bangalore",
        "Below 10 L Urban/Rural": "Below 10L Urban / Rural",
        "BELOW 10 LAC": "Below 10L",
        "Below 10 Lac & Rural": "Below 10L & Rural",
        "BELOW 10 LAC & RURAL": "Below 10L & Rural",
        "BELOW 10L URBAN": "Below 10L Urban",
        "Bihar/Jharkhand - Rural": "Bihar / Jharkhand - Rural",
        "Bihar/Jharkhand - Urban": "Bihar / Jharkhand - Urban",
        "Bihar/Jharkhand - Urban/Rural": "Bihar / Jharkhand - Urban / Rural",
        "Bihar/Jharkhand -Rural": "Bihar / Jharkhand - Rural",
        "Delhi": "Delhi",
        "Entire State CG": "Entire State CG",
        "Haryana": "Haryana",
        "HYDERABAD": "Hyderabad",
        "Hyderabad": "Hyderabad",
        "Karnataka": "Karnataka",
        "Karnataka Rural": "Karnataka - Rural",
        "Karnataka- Urban": "Karnataka - Urban",
        "Karnataka-Below 10L": "Karnataka - Below 10L",
        "Karnataka-Rural": "Karnataka - Rural",
        "M.P/C.G.- All Mkts": "MP / CG - All Markets",
        "Mah/Goa -  Urban/Rural": "Maharashtra / Goa - Urban / Rural",
        "Mah/Goa - Rural": "Maharashtra / Goa - Rural",
        "Mah/Goa -All Mkts": "Maharashtra / Goa - All Markets",
        "Mah/Goa 10-75 L": "Maharashtra / Goa - 10-75L",
        "Mah/Goa Rural": "Maharashtra / Goa - Rural",
        "Mah/Goa Urban": "Maharashtra / Goa - Urban",
        "Mega City": "Mega City",
        "Megacity": "Mega City",
        "Mp 1-10": "MP 1-10",
        "Mp 10+": "MP 10+",
        "MP 10+": "MP 10+",
        "MP 10-75": "MP 10-75",
        "Mp Lc1": "MP LC1",
        "MP LC1": "MP LC1",
        "MpCg Rural": "MP / CG - Rural",
        "NAGALAND/NORTHEAST/": "Nagaland / Northeast",
        "NORTHEAST /Urban-Rural": "Northeast - Urban / Rural",
        "Odisha Rural": "Odisha Rural",
        "Odisha Urban": "Odisha Urban",
        "Odisha Urban/Rural": "Odisha Urban / Rural",
        "Rajasthan-Rural": "Rajasthan - Rural",
        "Rajasthan-Urban": "Rajasthan - Urban",
        "Rajasthan-Urban-Rural": "Rajasthan - Urban / Rural",
        "Rajasthan-Urban/Rural": "Rajasthan - Urban / Rural",
        "Rural": "Rural",
        "RURAL": "Rural",
        "U.P./U.K. - All Mkts": "UP / Uttarakhand - All Markets",
        "UP/UK 10 to 75L": "UP / Uttarakhand - 10-75L",
        "UP/UK Below 10 L Urban": "UP / Uttarakhand - Below 10L Urban",
        "UP/UK Below 10L Urban": "UP / Uttarakhand - Below 10L Urban",
        "UP/UK Rural": "UP / Uttarakhand - Rural",
        "UP/Uttarakhand - 10-75L": "UP / Uttarakhand - 10-75L",
        "UP/Uttarakhand - Below 10 L Urban": "UP / Uttarakhand - Below 10L Urban",
        "UP/Uttarakhand - Below 10L Urban": "UP / Uttarakhand - Below 10L Urban",
        "UP/Uttarakhand - Rural": "UP / Uttarakhand - Rural",
        "UP/Uttarakhand - Urban": "UP / Uttarakhand - Urban",
        "UP/Uttarakhand - Urban/Rural": "UP / Uttarakhand - Urban / Rural",
        "UP/Uttarakhand -10 to 75L": "UP / Uttarakhand - 10-75L",
        "Urban": "Urban",
        "Urban-Rural": "Urban / Rural",
        "Urban/Rural": "Urban / Rural",
        "West Bengal -Rural": "West Bengal - Rural",
        "West Bengal Rural": "West Bengal - Rural",
        "West Bengal Urban-Rural": "West Bengal - Urban / Rural",
    }
)


def workbook_to_rows(workbook_path: Path) -> dict[str, list[list[Any]]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for worksheet in workbook.worksheets:
        sheets[worksheet.title] = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return sheets


def detect_label_rows(rows: list[list[Any]]) -> dict[str, int]:
    label_rows: dict[str, int] = {}
    for row_index, row in enumerate(rows[:15]):
        first_cell = row[0] if row else None
        canonical = LABEL_ALIASES.get(normalize_label(first_cell))
        if canonical:
            label_rows[canonical] = row_index
    return label_rows


def detect_blocks(rows: list[list[Any]], lcn_header_row: int) -> list[int]:
    header_row = rows[lcn_header_row] if lcn_header_row < len(rows) else []
    block_starts: list[int] = []
    for col_index in range(max(0, len(header_row) - 1)):
        left = normalize_label(header_row[col_index])
        right = normalize_label(header_row[col_index + 1]) if col_index + 1 < len(header_row) else ""
        if left in {"lcn", "lcnno", "lcnnumber"} and right in {"channel", "channelname", "channelnames"}:
            block_starts.append(col_index)
    return block_starts


def get_cell(rows: list[list[Any]], row_index: int | None, col_index: int) -> Any:
    if row_index is None or row_index < 0 or row_index >= len(rows):
        return None
    row = rows[row_index]
    if col_index >= len(row):
        return None
    return row[col_index]


def create_warning_counters() -> dict[str, Counter[str]]:
    return {
        "state": Counter(),
        "headend_location": Counter(),
        "barc_market": Counter(),
    }


def record_unmatched(counters: dict[str, Counter[str]], field: str, raw_value: str | None) -> None:
    if raw_value:
        counters[field][raw_value] += 1


def sheet_state_from_context(sheet_name: str | None) -> str | None:
    if not sheet_name:
        return None
    return SHEET_STATE_NORMALIZATION.get(normalize_lookup_key(sheet_name))


def normalize_state(
    raw_value: Any,
    *,
    sheet_name: str | None,
    headend_location: str | None,
    counters: dict[str, Counter[str]],
) -> str:
    raw_text = clean_text(raw_value)
    sheet_state = sheet_state_from_context(sheet_name)

    if raw_text is None:
        return sheet_state or UNMAPPED_REVIEW

    lookup_key = normalize_lookup_key(raw_text)
    if lookup_key in STATE_NORMALIZATION:
        return STATE_NORMALIZATION[lookup_key]

    if lookup_key in STATE_CONTEXTUAL_KEYS or lookup_key.startswith("district "):
        if sheet_state:
            return sheet_state
        record_unmatched(counters, "state", raw_text)
        return UNMAPPED_REVIEW

    record_unmatched(counters, "state", raw_text)
    return smart_title_case(raw_text) or UNMAPPED_REVIEW


def normalize_headend_location(raw_value: Any, *, counters: dict[str, Counter[str]]) -> str | None:
    raw_text = clean_text(raw_value)
    if raw_text is None:
        return None

    lookup_key = normalize_lookup_key(raw_text)
    if lookup_key in CITY_NORMALIZATION:
        return CITY_NORMALIZATION[lookup_key]

    record_unmatched(counters, "headend_location", raw_text)
    return smart_title_case(raw_text)


def normalize_barc_market(raw_value: Any, *, counters: dict[str, Counter[str]]) -> str | None:
    raw_text = clean_text(raw_value)
    if raw_text is None:
        return None

    lookup_key = normalize_lookup_key(raw_text)
    if lookup_key in BARC_MARKET_NORMALIZATION:
        return BARC_MARKET_NORMALIZATION[lookup_key]

    record_unmatched(counters, "barc_market", raw_text)
    return smart_title_case(raw_text)


def parse_sheet(
    sheet_name: str,
    rows: list[list[Any]],
    counters: dict[str, Counter[str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if len(rows) < 11:
        return [], []

    label_rows = detect_label_rows(rows)
    lcn_header_row = 10
    block_starts = detect_blocks(rows, lcn_header_row)

    if not block_starts:
        LOGGER.warning("Skipping sheet '%s' because no LCN blocks were detected.", sheet_name)
        return [], []

    distribution_rows: list[dict[str, Any]] = []
    channel_rows: list[dict[str, Any]] = []

    for block_start in block_starts:
        record_date = clean_date(get_cell(rows, label_rows.get("date"), block_start))
        network_name = clean_text(get_cell(rows, label_rows.get("network_name"), block_start))
        headend_location = normalize_headend_location(
            get_cell(rows, label_rows.get("headend_location"), block_start),
            counters=counters,
        )
        state = normalize_state(
            get_cell(rows, label_rows.get("state"), block_start) or sheet_name,
            sheet_name=sheet_name,
            headend_location=headend_location,
            counters=counters,
        )
        barc_market = normalize_barc_market(
            get_cell(rows, label_rows.get("barc_market"), block_start),
            counters=counters,
        )
        stbs = clean_int(get_cell(rows, label_rows.get("stbs"), block_start))
        landing_channel_1 = clean_text(get_cell(rows, label_rows.get("landing_channel_1"), block_start))
        landing_channel_2 = clean_text(get_cell(rows, label_rows.get("landing_channel_2"), block_start))
        barker_1 = clean_text(get_cell(rows, label_rows.get("barker_1"), block_start))
        barker_2 = clean_text(get_cell(rows, label_rows.get("barker_2"), block_start))

        if not network_name and not headend_location:
            continue

        headend_id = generate_headend_id(
            network_name=network_name,
            headend_location=headend_location,
            state=state,
        )

        distribution_rows.append(
            {
                "headend_id": headend_id,
                "date": record_date,
                "network_name": network_name,
                "headend_location": headend_location,
                "state": state,
                "barc_market": barc_market,
                "stbs": stbs,
                "landing_channel_1": landing_channel_1,
                "landing_channel_2": landing_channel_2,
                "barker_1": barker_1,
                "barker_2": barker_2,
                "source_sheet": sheet_name,
            }
        )

        for row_index in range(lcn_header_row + 1, len(rows)):
            lcn_no = clean_text(get_cell(rows, row_index, block_start))
            channel_name = clean_text(get_cell(rows, row_index, block_start + 1))

            if not lcn_no and not channel_name:
                continue

            if normalize_label(get_cell(rows, row_index, block_start)) in {"lcn", "lcnno", "lcnnumber"}:
                continue

            if lcn_no is None or channel_name is None:
                continue

            channel_rows.append(
                {
                    "headend_id": headend_id,
                    "date": record_date,
                    "network_name": network_name,
                    "headend_location": headend_location,
                    "state": state,
                    "barc_market": barc_market,
                    "lcn_no": lcn_no,
                    "channel_name": channel_name,
                }
            )

    return distribution_rows, channel_rows


def parse_workbook(workbook_path: Path) -> ParsedWorkbook:
    LOGGER.info("Reading workbook: %s", workbook_path)
    workbook_rows = workbook_to_rows(workbook_path)
    distribution_rows: list[dict[str, Any]] = []
    channel_rows: list[dict[str, Any]] = []
    counters = create_warning_counters()

    for sheet_name, rows in workbook_rows.items():
        LOGGER.info("Parsing sheet: %s", sheet_name)
        sheet_distribution, sheet_channels = parse_sheet(sheet_name, rows, counters)
        distribution_rows.extend(sheet_distribution)
        channel_rows.extend(sheet_channels)

    distribution_rows = list(
        {
            (
                row["headend_id"],
                row.get("date"),
                row.get("network_name"),
                row.get("headend_location"),
                row.get("state"),
            ): row
            for row in distribution_rows
        }.values()
    )
    channel_rows = list(
        {
            (
                row["headend_id"],
                row.get("date"),
                row.get("lcn_no"),
                row.get("channel_name"),
            ): row
            for row in channel_rows
        }.values()
    )
    week_dates = sorted({row["date"] for row in channel_rows if row.get("date")})

    LOGGER.info(
        "Parsed workbook into %s headend snapshots and %s channel rows across %s weeks.",
        len(distribution_rows),
        len(channel_rows),
        len(week_dates),
    )

    return ParsedWorkbook(
        distribution_rows=distribution_rows,
        channel_rows=channel_rows,
        week_dates=week_dates,
        normalization_warnings={
            field: dict(counter.most_common())
            for field, counter in counters.items()
            if counter
        },
    )
