#!/usr/bin/env python3
"""Build a daily VOD-key workbook with watch minutes and audience counts."""

from __future__ import annotations

import argparse
import csv
import io
from datetime import date, datetime, timedelta
from pathlib import Path

import duckdb
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
ETL_ROOT = HERE.parents[1]
DEFAULT_EVENTS = ETL_ROOT / "output" / "exports" / "vod_stream_query_events.csv"
DEFAULT_START_DATE = date(2026, 8, 24)
SHEETS = (
    ("Watch Minutes", "watch_minutes"),
    ("Distinct CLI IPs", "cli_ips"),
    ("Distinct Device IDs", "device_ids"),
)
VIDEO_TITLE_HEADERS = ("video report title", "video title")
VOD_KEY_HEADER = "vod key"


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as error:
        raise argparse.ArgumentTypeError("Dates must use YYYY-MM-DD.") from error


def output_path_for(source: Path) -> Path:
    return source.with_name(f"{source.stem}_details.xlsx")


def normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def parse_video_rows(
    headers: tuple[object, ...] | list[object],
    rows,
    source_label: str,
) -> list[tuple[str, str]]:
    header_map = {
        normalize_header(value): index
        for index, value in enumerate(headers)
    }
    title_indexes = [
        header_map[name] for name in VIDEO_TITLE_HEADERS if name in header_map
    ]
    if not title_indexes or VOD_KEY_HEADER not in header_map:
        missing = []
        if not title_indexes:
            missing.append("Video Title or Video Report Title")
        if VOD_KEY_HEADER not in header_map:
            missing.append("VOD Key")
        raise ValueError(f"Missing columns in {source_label}: {', '.join(missing)}")

    code_index = header_map[VOD_KEY_HEADER]
    videos: list[tuple[str, str]] = []
    seen_codes: set[str] = set()
    for excel_row, row in enumerate(rows, start=2):
        values = list(row)
        title = next(
            (
                str(values[index] or "").strip()
                for index in title_indexes
                if index < len(values) and str(values[index] or "").strip()
            ),
            "",
        )
        code = (
            str(values[code_index] or "").strip().lower()
            if code_index < len(values)
            else ""
        )
        if not title and not code:
            continue
        if not title or not code:
            raise ValueError(f"Row {excel_row} must contain both Video Title and VOD Key.")
        if code in seen_codes:
            raise ValueError(f"Duplicate VOD Key {code!r} at row {excel_row}.")
        seen_codes.add(code)
        videos.append((title, code))
    if not videos:
        raise ValueError(f"No videos were found in {source_label}.")
    return videos


def read_video_list(path: Path, sheet_name: str) -> list[tuple[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Video-list workbook was not found: {path}")

    if path.suffix.lower() == ".xls":
        raw = path.read_bytes()
        if raw.startswith(b"\xd0\xcf\x11\xe0"):
            raise ValueError(
                f"{path.name!r} is a binary legacy XLS file. Convert it to XLSX first."
            )
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp1252")
        rows = csv.reader(io.StringIO(text, newline=""), delimiter="\t")
        try:
            headers = next(rows)
        except StopIteration as error:
            raise ValueError(f"Tab-separated XLS export {path.name!r} is empty.") from error
        return parse_video_rows(headers, rows, repr(path.name))

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} was not found. Available: {', '.join(workbook.sheetnames)}"
        )
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as error:
        raise ValueError(f"Sheet {sheet_name!r} is empty.") from error
    return parse_video_rows(headers, rows, repr(sheet_name))


def date_range(start: date, end: date) -> list[date]:
    return [start + timedelta(days=offset) for offset in range((end - start).days + 1)]


def half_up_minutes(hours: float) -> int:
    return int(max(0.0, hours) * 60 + 0.5)


def query_metrics(
    events_path: Path,
    videos: list[tuple[str, str]],
    start_date: date,
    end_date: date | None,
) -> tuple[list[date], dict[tuple[str, date], dict[str, float | int]], dict[str, dict[str, float | int]], set[date]]:
    if not events_path.exists():
        raise FileNotFoundError(f"VOD event CSV was not found: {events_path}")
    connection = duckdb.connect(":memory:")
    try:
        connection.execute(
            "CREATE TEMP TABLE requested_videos (content_code VARCHAR PRIMARY KEY)"
        )
        connection.executemany(
            "INSERT INTO requested_videos VALUES (?)",
            [(code,) for _, code in videos],
        )
        source_sql = """
            SELECT
                try_cast(log_date AS DATE) AS log_date,
                lower(trim(COALESCE(content_code, ''))) AS content_code,
                COALESCE(try_cast(request_watch_hours AS DOUBLE), 0.0) AS watch_hours,
                NULLIF(trim(COALESCE(cli_ip, '')), '') AS cli_ip,
                NULLIF(trim(COALESCE(device_id, '')), '') AS device_id
            FROM read_csv_auto(?, all_varchar = true)
            WHERE try_cast(log_date AS DATE) IS NOT NULL
        """
        available_dates = {
            row[0]
            for row in connection.execute(
                f"SELECT DISTINCT log_date FROM ({source_sql}) ORDER BY log_date",
                [str(events_path)],
            ).fetchall()
        }
        if not available_dates:
            raise ValueError(f"No dated event rows were found in {events_path}.")
        resolved_end = end_date or max(available_dates)
        if resolved_end < start_date:
            raise ValueError(
                f"End date {resolved_end} is before start date {start_date}."
            )
        dates = date_range(start_date, resolved_end)
        daily_rows = connection.execute(
            f"""
            WITH source AS ({source_sql})
            SELECT
                source.content_code,
                source.log_date,
                sum(source.watch_hours) AS watch_hours,
                count(DISTINCT source.cli_ip) AS cli_ips,
                count(DISTINCT source.device_id) AS device_ids
            FROM source
            INNER JOIN requested_videos USING (content_code)
            WHERE source.log_date BETWEEN ? AND ?
            GROUP BY source.content_code, source.log_date
            """,
            [str(events_path), start_date, resolved_end],
        ).fetchall()
        total_rows = connection.execute(
            f"""
            WITH source AS ({source_sql})
            SELECT
                source.content_code,
                sum(source.watch_hours) AS watch_hours,
                count(DISTINCT source.cli_ip) AS cli_ips,
                count(DISTINCT source.device_id) AS device_ids
            FROM source
            INNER JOIN requested_videos USING (content_code)
            WHERE source.log_date BETWEEN ? AND ?
            GROUP BY source.content_code
            """,
            [str(events_path), start_date, resolved_end],
        ).fetchall()
    finally:
        connection.close()

    daily = {
        (code, row_date): {
            "watch_minutes": half_up_minutes(float(watch_hours or 0)),
            "cli_ips": int(cli_ips or 0),
            "device_ids": int(device_ids or 0),
        }
        for code, row_date, watch_hours, cli_ips, device_ids in daily_rows
    }
    totals = {
        code: {
            "watch_minutes": half_up_minutes(float(watch_hours or 0)),
            "cli_ips": int(cli_ips or 0),
            "device_ids": int(device_ids or 0),
        }
        for code, watch_hours, cli_ips, device_ids in total_rows
    }
    return dates, daily, totals, available_dates


def style_sheet(sheet, date_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="F2F6FA")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row_index in range(2, sheet.max_row + 1):
        for cell in sheet[row_index]:
            if row_index % 2 == 0:
                cell.fill = alternate_fill
            cell.alignment = Alignment(
                horizontal="left" if cell.column <= 2 else "center",
                vertical="center",
            )
            cell.border = border
            if cell.column > 2:
                cell.number_format = "0"
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    sheet.column_dimensions["A"].width = 60
    sheet.column_dimensions["B"].width = 15
    for column in range(3, 3 + date_count):
        sheet.column_dimensions[get_column_letter(column)].width = 13
    sheet.column_dimensions[get_column_letter(3 + date_count)].width = 18


def write_report(
    output_path: Path,
    videos: list[tuple[str, str]],
    dates: list[date],
    daily: dict[tuple[str, date], dict[str, float | int]],
    totals: dict[str, dict[str, float | int]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    date_headers = [value.strftime("%d %b %Y") for value in dates]
    for sheet_name, metric in SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Video Title", "VOD Key", *date_headers, "Total"])
        for title, code in videos:
            sheet.append(
                [title, code]
                + [int(daily.get((code, value), {}).get(metric, 0)) for value in dates]
                + [int(totals.get(code, {}).get(metric, 0))]
            )
        style_sheet(sheet, len(dates))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_report(
    source_workbook: Path,
    events_path: Path = DEFAULT_EVENTS,
    output_path: Path | None = None,
    sheet_name: str = "Video List",
    start_date: date = DEFAULT_START_DATE,
    end_date: date | None = None,
) -> tuple[Path, list[date], set[date], int]:
    videos = read_video_list(source_workbook, sheet_name)
    dates, daily, totals, available_dates = query_metrics(
        events_path, videos, start_date, end_date
    )
    resolved_output = output_path or output_path_for(source_workbook)
    write_report(resolved_output, videos, dates, daily, totals)
    return resolved_output, dates, available_dates, len(videos)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build daily watch-minute, CLI-IP, and device-ID sheets for VOD keys."
    )
    parser.add_argument("workbook", type=Path, help="XLSX containing Video Title and VOD Key columns.")
    parser.add_argument("--events", type=Path, default=DEFAULT_EVENTS)
    parser.add_argument("--out", type=Path, help="Defaults to <input_stem>_details.xlsx beside the input.")
    parser.add_argument("--sheet", default="Video List")
    parser.add_argument("--start-date", type=parse_date, default=DEFAULT_START_DATE)
    parser.add_argument("--end-date", type=parse_date)
    args = parser.parse_args()

    output, dates, available_dates, video_count = build_report(
        source_workbook=args.workbook,
        events_path=args.events,
        output_path=args.out,
        sheet_name=args.sheet,
        start_date=args.start_date,
        end_date=args.end_date,
    )
    missing_dates = [value for value in dates if value not in available_dates]
    print(
        f"Wrote {output} for {video_count} videos from {dates[0]} through {dates[-1]}."
    )
    if missing_dates:
        print(
            "Warning: no source event rows were available for: "
            + ", ".join(value.isoformat() for value in missing_dates)
        )


if __name__ == "__main__":
    main()
