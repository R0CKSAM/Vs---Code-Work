#!/usr/bin/env python3
"""Build an incremental VOD segment-analysis dashboard from event CSV exports."""

from __future__ import annotations

import argparse
import base64
import csv
import gzip
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote_plus
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
ETL_ROOT = HERE.parents[1]
DEFAULT_EVENTS = ETL_ROOT / "output" / "exports" / "vod_stream_query_events.csv"
DEFAULT_OUTPUT = ETL_ROOT / "output" / "exports" / "vod_stream_query_analysis_dashboard.html"
URL_DECODE_FIELDS = {
    "content_title", "category_name", "content_type", "channel", "platform",
    "device", "country", "state", "city",
}
BROWSER_FIELDS = (
    "log_date", "minute_ist", "request_ist", "last_request_ist", "content_title",
    "category_name", "channel", "platform", "device", "user_agent", "decode_status",
    "decode_confidence", "decoded_device_type", "decoded_form_factor", "decoded_brand",
    "decoded_model", "decoded_os", "decoded_browser", "decoded_player", "session_id",
    "device_id", "cli_ip", "country", "state", "city", "req_host", "req_path",
    "status_code", "cache_status", "content_code", "is_successful_segment",
    "request_watch_hours", "delivered_watch_hours", "segment_count",
    "successful_segment_count",
)


def clean_field(name: str, value: object) -> str:
    """Trim CSV values and decode only fields known to contain URL encoding."""
    text = str(value or "").strip()
    return unquote_plus(text) if name in URL_DECODE_FIELDS else text


def read_events(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Event CSV was not found: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {key: clean_field(key, value) for key, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def event_key(row: dict[str, str]) -> tuple[str, ...]:
    """Identify one CDN request without relying on inherited metadata."""
    return tuple(
        row.get(key, "")
        for key in (
            "request_ist", "last_request_ist", "cli_ip", "req_host", "req_path",
            "status_code", "content_code", "session_id", "device_id",
        )
    )


def merge_events(
    existing: list[dict[str, str]], incoming: list[dict[str, str]]
) -> list[dict[str, str]]:
    incoming_dates = {row.get("log_date", "") for row in incoming}
    merged: dict[tuple[str, ...], dict[str, str]] = {}
    retained = [row for row in existing if row.get("log_date", "") not in incoming_dates]
    for row in retained + incoming:
        merged[event_key(row)] = row
    return sorted(
        merged.values(),
        key=lambda row: (row.get("log_date", ""), row.get("request_ist", "")),
    )


def write_events(path: Path, rows: list[dict[str, str]]) -> None:
    fields: list[str] = []
    seen_fields: set[str] = set()
    for row in rows:
        for field in row:
            if field not in seen_fields:
                fields.append(field)
                seen_fields.add(field)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def number(value: str) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def canonicalize_content_titles(
    rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    """Use the segment-weighted dominant title for each request-path video key."""
    title_weights: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    invalid_titles = {"", "null", "undefined", "unknown", "unknown / na", "na"}
    for row in rows:
        code = row.get("content_code", "").strip().lower()
        title = row.get("content_title", "").strip()
        if code and title.lower() not in invalid_titles:
            title_weights[code][title] += number(row.get("segment_count", "1"))
    canonical_titles = {
        code: max(weights, key=lambda title: (weights[title], title))
        for code, weights in title_weights.items()
        if weights
    }
    normalized: list[dict[str, str]] = []
    for row in rows:
        item = dict(row)
        code = item.get("content_code", "").strip().lower()
        if code in canonical_titles:
            item["content_title"] = canonical_titles[code]
        normalized.append(item)
    return normalized


def day_summary(rows: list[dict[str, str]]) -> dict[str, object]:
    minute_ips: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row.get("cli_ip"):
            minute_ips[row.get("minute_ist", "")].add(row["cli_ip"])
    return {
        "media_segments": int(sum(number(row.get("segment_count", "1")) for row in rows)),
        "cli_ips": len({row["cli_ip"] for row in rows if row.get("cli_ip")}),
        "device_ids": len({row["device_id"] for row in rows if row.get("device_id")}),
        "session_ids": len({row["session_id"] for row in rows if row.get("session_id")}),
        "watch_hours": round(
            sum(number(row.get("request_watch_hours", "")) for row in rows), 6
        ),
        "delivered_watch_hours": round(
            sum(number(row.get("delivered_watch_hours", "")) for row in rows), 6
        ),
        "peak_concurrency": max((len(values) for values in minute_ips.values()), default=0),
    }


def is_davis_cup_row(row: dict[str, str]) -> bool:
    title = row.get("content_title", "").strip().lower()
    category = row.get("category_name", "").strip().lower().replace("_", " ")
    return "davis cup" in title or "davis cup" in category


def watch_minutes(hours: float) -> int:
    """Return whole watch minutes using the dashboard's half-up rounding."""
    return int(max(0.0, hours) * 60 + 0.5)


def davis_cup_performance(rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, object]]]:
    available_dates = sorted(
        {row.get("log_date", "") for row in rows if row.get("log_date")}
    )
    davis_codes = {
        row.get("content_code", "")
        for row in rows
        if is_davis_cup_row(row) and row.get("content_code")
    }
    davis_rows = [row for row in rows if row.get("content_code") in davis_codes]
    if not davis_rows:
        return [], []
    first_davis_date = min(row["log_date"] for row in davis_rows if row.get("log_date"))
    dates = [date for date in available_dates if date >= first_davis_date]
    groups: dict[str, dict[str, object]] = {}
    for row in davis_rows:
        code = row["content_code"]
        group = groups.setdefault(
            code,
            {
                "code": code,
                "first_seen": row.get("log_date", ""),
                "title_weights": defaultdict(float),
                "daily_hours": defaultdict(float),
                "daily_ips": defaultdict(set),
                "total_ips": set(),
            },
        )
        row_date = row.get("log_date", "")
        if row_date and row_date < group["first_seen"]:
            group["first_seen"] = row_date
        if is_davis_cup_row(row):
            title = row.get("content_title", "").strip() or code
            group["title_weights"][title] += number(row.get("segment_count", "1"))
        group["daily_hours"][row_date] += number(row.get("request_watch_hours", ""))
        cli_ip = row.get("cli_ip", "")
        if cli_ip:
            group["daily_ips"][row_date].add(cli_ip)
            group["total_ips"].add(cli_ip)

    performance: list[dict[str, object]] = []
    for group in groups.values():
        title_weights = group["title_weights"]
        title = max(title_weights, key=lambda value: (title_weights[value], value)) if title_weights else group["code"]
        if title.lower().startswith("davis cup_"):
            title = "Davis Cup | " + title[len("Davis Cup_"):]
        total_hours = sum(group["daily_hours"].values())
        performance.append(
            {
                **group,
                "title": title,
                "total_hours": total_hours,
            }
        )
    performance.sort(key=lambda group: (-group["total_hours"], group["title"]))
    return dates, performance


def style_performance_sheet(sheet, date_columns: int) -> None:
    dark_blue = PatternFill("solid", fgColor="1F4E78")
    alternate = PatternFill("solid", fgColor="F2F2F2")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in sheet[1]:
        cell.fill = dark_blue
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_index in range(2, sheet.max_row + 1):
        for cell in sheet[row_index]:
            if row_index % 2 == 0:
                cell.fill = alternate
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 1 else "center",
                vertical="center",
            )
            cell.border = border
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions["A"].width = 55
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 13
    for column in range(4, 4 + date_columns):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.column_dimensions[get_column_letter(4 + date_columns)].width = 20


def write_davis_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    dates, performance = davis_cup_performance(rows)
    workbook = Workbook()
    watch_sheet = workbook.active
    watch_sheet.title = "Watch Minutes"
    cli_sheet = workbook.create_sheet("CLI IPs")
    date_labels = [
        f"{value.strftime('%b')} {value.day}"
        for value in (datetime.strptime(date, "%Y-%m-%d") for date in dates)
    ]
    watch_sheet.append(
        ["Video Title", "Video Key", "First Seen"]
        + [f"Watch Minutes - {label}" for label in date_labels]
        + ["Total Watch Minutes"]
    )
    cli_sheet.append(
        ["Video Title", "Video Key", "First Seen"]
        + [f"Unique CLI IPs - {label}" for label in date_labels]
        + ["Total Unique CLI IPs"]
    )
    for group in performance:
        watch_sheet.append(
            [group["title"], group["code"], group["first_seen"]]
            + [watch_minutes(group["daily_hours"].get(date, 0.0)) for date in dates]
            + [watch_minutes(group["total_hours"])]
        )
        cli_sheet.append(
            [group["title"], group["code"], group["first_seen"]]
            + [len(group["daily_ips"].get(date, set())) for date in dates]
            + [len(group["total_ips"])]
        )
    style_performance_sheet(watch_sheet, len(dates))
    style_performance_sheet(cli_sheet, len(dates))
    for row in watch_sheet.iter_rows(
        min_row=2, min_col=4, max_col=watch_sheet.max_column
    ):
        for cell in row:
            cell.number_format = "0"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def encode_browser_rows(rows: list[dict[str, str]]) -> str:
    payload_json = json.dumps(rows, ensure_ascii=True, separators=(",", ":"))
    return base64.b64encode(
        gzip.compress(payload_json.encode("utf-8"), compresslevel=9)
    ).decode("ascii")


def script_json(value: object) -> str:
    """Serialize values safely inside an inline script element."""
    return (
        json.dumps(value, ensure_ascii=True, separators=(",", ":"))
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
        .replace("&", "\\u0026")
    )


def prepare_dashboard_data(
    rows: list[dict[str, str]],
) -> tuple[dict[str, object], dict[str, list[dict[str, str]]]]:
    display_rows = canonicalize_content_titles(rows)
    day_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    picker_values: dict[str, set[str]] = {
        "title": set(),
        "category": set(),
        "code": set(),
    }
    first_seen_by_code: dict[str, str] = {}
    davis_codes: set[str] = set()

    for row in display_rows:
        date = row.get("log_date", "")
        if not date:
            continue
        browser_row = {field: row.get(field, "") for field in BROWSER_FIELDS}
        day_rows[date].append(browser_row)
        title = row.get("content_title", "")
        category = row.get("category_name", "")
        code = row.get("content_code", "")
        if title:
            picker_values["title"].add(title)
        if category:
            picker_values["category"].add(category)
        if code:
            picker_values["code"].add(code)
            previous = first_seen_by_code.get(code)
            if previous is None or date < previous:
                first_seen_by_code[code] = date
            if is_davis_cup_row(row):
                davis_codes.add(code)

    dates = sorted(day_rows)
    index: dict[str, object] = {
        "dates": dates,
        "default_window_days": 7,
        "picker_values": {
            key: sorted(values, key=str.casefold)
            for key, values in picker_values.items()
        },
        "first_seen_by_code": first_seen_by_code,
        "davis_codes": sorted(davis_codes),
    }
    return index, dict(day_rows)


def write_dashboard_data(
    data_dir: Path,
    days: dict[str, list[dict[str, str]]],
) -> None:
    """Atomically replace date-partitioned browser payloads."""
    data_dir.parent.mkdir(parents=True, exist_ok=True)
    staged = data_dir.with_name(f"{data_dir.name}.tmp-{uuid4().hex}")
    staged.mkdir(parents=True)
    try:
        for date, rows in sorted(days.items()):
            payload = encode_browser_rows(rows)
            registration = (
                "window.__vodDayPayloads=window.__vodDayPayloads||{};"
                f"window.__vodDayPayloads[{json.dumps(date)}]={json.dumps(payload)};"
            )
            (staged / f"{date}.js").write_text(registration, encoding="ascii")
        if data_dir.exists():
            shutil.rmtree(data_dir)
        staged.replace(data_dir)
    finally:
        if staged.exists():
            shutil.rmtree(staged)


def dashboard_script() -> str:
    return """<script>
(async()=>{
async function loadPayload(encoded){
  if(typeof DecompressionStream!=='function')throw new Error('This dashboard requires a current Chrome or Edge browser.');
  const bytes=Uint8Array.from(atob(encoded),char=>char.charCodeAt(0));
  const stream=new Blob([bytes]).stream().pipeThrough(new DecompressionStream('gzip'));
  return JSON.parse(await new Response(stream).text());
}
const dashboardIndex=__DASHBOARD_INDEX__;
const dataDirectory=__DATA_DIRECTORY__;
window.__vodDayPayloads=Object.assign(window.__vodDayPayloads||{},__INLINE_PAYLOADS__);
const by=id=>document.getElementById(id);
const countFmt=new Intl.NumberFormat('en-IN');
const esc=value=>String(value??'').replace(/[&<>"']/g,char=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[char]));
const normalized=value=>String(value??'').trim().toLocaleLowerCase();
const segmentCount=row=>Number(row.segment_count||1);
const availableDates=[...(dashboardIndex.dates||[])].filter(Boolean).sort();
const availableDateSet=new Set(availableDates);
const filterState={title:new Set(),category:new Set(),code:new Set()};
const pickerValues={
  title:[...(dashboardIndex.picker_values?.title||[])],
  category:[...(dashboardIndex.picker_values?.category||[])],
  code:[...(dashboardIndex.picker_values?.code||[])],
};
const pickerLabels={title:'titles',category:'categories',code:'video keys'};
const firstSeenByCode=new Map(Object.entries(dashboardIndex.first_seen_by_code||{}));
const davisCodes=new Set(dashboardIndex.davis_codes||[]);
const dayCache=new Map();
const dayLoads=new Map();
let activeDays=[];
let activeRows=[];
let davisMode=false;
let loadVersion=0;
let renderVersion=0;
let renderTimer=0;

const yieldToBrowser=()=>new Promise(resolve=>setTimeout(resolve,0));

function setLoadStatus(message,busy=false){
  const status=by('loadStatus');
  status.textContent=message;
  status.classList.toggle('busy',busy);
  document.body.classList.toggle('dashboard-busy',busy);
}

function loadDayScript(date){
  return new Promise((resolve,reject)=>{
    if(window.__vodDayPayloads[date]){resolve();return;}
    if(!dataDirectory){reject(new Error(`No embedded dashboard data for ${date}.`));return;}
    const script=document.createElement('script');
    script.src=new URL(`${dataDirectory}/${date}.js`,document.baseURI).href;
    script.async=true;
    script.onload=()=>{script.remove();resolve();};
    script.onerror=()=>{script.remove();reject(new Error(`Could not load dashboard data for ${date}. Keep the ${dataDirectory} folder beside this HTML file.`));};
    document.head.append(script);
  });
}

async function loadDay(date){
  if(dayCache.has(date))return dayCache.get(date);
  if(dayLoads.has(date))return dayLoads.get(date);
  const loading=(async()=>{
    await loadDayScript(date);
    const encoded=window.__vodDayPayloads[date];
    if(!encoded)throw new Error(`Dashboard payload registration failed for ${date}.`);
    const rows=await loadPayload(encoded);
    delete window.__vodDayPayloads[date];
    const day={date,rows};
    dayCache.set(date,day);
    return day;
  })();
  dayLoads.set(date,loading);
  try{return await loading;}finally{dayLoads.delete(date);}
}

function watchTimeParts(hours){
  const value=Math.max(0,Number(hours)||0);
  const totalMinutes=Math.round(value*60);
  return {hours:Math.floor(totalMinutes/60),minutes:totalMinutes%60,positive:value>0};
}

function duration(hours){
  const parts=watchTimeParts(hours);
  if(parts.hours&&parts.minutes)return `${countFmt.format(parts.hours)} h ${parts.minutes} min`;
  if(parts.hours)return `${countFmt.format(parts.hours)} h`;
  if(parts.minutes)return `${parts.minutes} min`;
  return parts.positive?'<1 min':'0 min';
}

function decodedDeviceLabel(row){
  const status=normalized(row.decode_status);
  const decoded=status==='decoded_local'||status==='decoded_api';
  if(!decoded){
    return `Unknown decode / Raw: ${row.device||'Unknown / NA'} / App: ${row.platform||'Unknown / NA'}`;
  }
  const brand=row.decoded_brand&&row.decoded_brand!=='Unknown / NA'?row.decoded_brand:'';
  const model=row.decoded_model&&row.decoded_model!=='Unknown / NA'?row.decoded_model:'';
  const modelLabel=brand&&model&&!normalized(model).startsWith(normalized(brand))?`${brand} ${model}`:(model||brand);
  const parts=[row.decoded_device_type,modelLabel,row.decoded_os]
    .filter(value=>value&&value!=='Unknown / NA');
  parts.push(`App: ${row.platform||'Unknown / NA'}`);
  return parts.join(' / ');
}

function dateLabel(value){
  if(!value)return 'Unavailable';
  const [year,month,day]=value.split('-').map(Number);
  return new Intl.DateTimeFormat('en-GB',{day:'2-digit',month:'short',year:'numeric'}).format(new Date(Date.UTC(year,month-1,day)));
}

function dateDistance(from,to){
  return Math.round((Date.parse(`${to}T00:00:00Z`)-Date.parse(`${from}T00:00:00Z`))/86400000);
}

function nearestAvailableDate(value,boundary){
  if(!value||availableDateSet.has(value))return value;
  const directional=boundary==='from'
    ? availableDates.find(date=>date>=value)
    : [...availableDates].reverse().find(date=>date<=value);
  if(directional)return directional;
  return availableDates.reduce((best,date)=>Math.abs(dateDistance(value,date))<Math.abs(dateDistance(value,best))?date:best,availableDates[0]);
}

function renderDateAvailability(){
  const first=availableDates[0]||'',last=availableDates.at(-1)||'';
  const calendarDays=first&&last?dateDistance(first,last)+1:0;
  const missing=Math.max(0,calendarDays-availableDates.length);
  const from=by('dateFrom').value||first,to=by('dateTo').value||last;
  const selected=availableDates.filter(date=>date>=from&&date<=to);
  const selectedCalendarDays=from&&to?dateDistance(from,to)+1:0;
  const selectedMissing=Math.max(0,selectedCalendarDays-selected.length);
  by('dateAvailabilitySummary').textContent=`Data available: ${dateLabel(first)} to ${dateLabel(last)} | ${countFmt.format(availableDates.length)} dates | ${missing?`${missing} missing`:'continuous'}`;
  by('dateSelectionStatus').textContent=`Selected range contains ${countFmt.format(selected.length)} data dates${selectedMissing?` and ${countFmt.format(selectedMissing)} dates with no data`:''}.`;
  by('availableDateList').innerHTML=availableDates.map(date=>`<time datetime="${date}" class="${date>=from&&date<=to?'selected':''}">${dateLabel(date)}</time>`).join('');
}

function syncNativeSelect(type){
  const select=by(`${type}Filter`),selected=filterState[type];
  for(const option of select.options)option.selected=selected.has(option.value);
}

function updatePickerToggle(type){
  const selected=filterState[type],toggle=by(`${type}Toggle`);
  if(selected.size===0)toggle.textContent=`All ${pickerLabels[type]}`;
  else if(selected.size===1)toggle.textContent=[...selected][0];
  else toggle.textContent=`${countFmt.format(selected.size)} selected`;
  syncNativeSelect(type);
}

function visiblePickerValues(type){
  const query=normalized(by(`${type}Search`).value);
  return pickerValues[type].filter(value=>!query||normalized(value).includes(query));
}

function renderPickerOptions(type){
  const values=visiblePickerValues(type);
  const container=by(`${type}Options`);
  container.replaceChildren();
  for(const value of values){
    const label=document.createElement('label');
    label.className='picker-option';
    const checkbox=document.createElement('input');
    checkbox.type='checkbox';
    checkbox.value=value;
    checkbox.checked=filterState[type].has(value);
    const text=document.createElement('span');
    text.textContent=value;
    checkbox.addEventListener('change',()=>{
      if(checkbox.checked)filterState[type].add(value);
      else filterState[type].delete(value);
      updatePickerToggle(type);
      scheduleRender();
    });
    label.append(checkbox,text);
    container.append(label);
  }
  if(values.length===0){
    const empty=document.createElement('div');
    empty.className='picker-empty';
    empty.textContent='No matching options';
    container.append(empty);
  }
  by(`${type}ResultCount`).textContent=`${countFmt.format(values.length)} of ${countFmt.format(pickerValues[type].length)}`;
}

function populatePicker(type){
  const select=by(`${type}Filter`);
  for(const value of pickerValues[type]){
    const option=document.createElement('option');
    option.value=value;
    option.textContent=value;
    select.append(option);
  }
  updatePickerToggle(type);
  renderPickerOptions(type);
}

function closePickers(except=null){
  document.querySelectorAll('.picker.open').forEach(picker=>{
    if(picker===except)return;
    picker.classList.remove('open');
    picker.querySelector('.picker-toggle')?.setAttribute('aria-expanded','false');
  });
}

function togglePicker(type){
  const picker=by(`${type}Picker`),toggle=by(`${type}Toggle`);
  const opening=!picker.classList.contains('open');
  closePickers(picker);
  picker.classList.toggle('open',opening);
  toggle.setAttribute('aria-expanded',String(opening));
  if(opening){
    by(`${type}Search`).value='';
    renderPickerOptions(type);
    requestAnimationFrame(()=>by(`${type}Search`).focus());
  }
}

function setAll(type,selected){
  if(selected){
    for(const value of visiblePickerValues(type))filterState[type].add(value);
  }else{
    filterState[type].clear();
  }
  updatePickerToggle(type);
  renderPickerOptions(type);
  scheduleRender();
}

function selectedRows(){
  const titles=filterState.title,categories=filterState.category,codes=filterState.code;
  return activeDays
    .flatMap(day=>day.rows)
    .filter(row=>(titles.size===0||titles.has(row.content_title))
      &&(categories.size===0||categories.has(row.category_name))
      &&(codes.size===0||codes.has(row.content_code)));
}

function selectedDates(){
  const from=by('dateFrom').value||availableDates[0]||'';
  const to=by('dateTo').value||availableDates.at(-1)||'';
  return availableDates.filter(date=>date>=from&&date<=to);
}

async function refreshData(){
  const version=++loadVersion;
  const dates=selectedDates();
  const dateWord=dates.length===1?'date':'dates';
  const loaded=[];
  setLoadStatus(`Loading ${countFmt.format(dates.length)} selected ${dateWord}...`,true);
  for(let index=0;index<dates.length;index+=1){
    const date=dates[index];
    const day=await loadDay(date);
    if(version!==loadVersion)return;
    loaded.push(day);
    setLoadStatus(`Loading ${countFmt.format(index+1)} of ${countFmt.format(dates.length)} ${dateWord}...`,true);
    await yieldToBrowser();
  }
  if(version!==loadVersion)return;
  activeDays=loaded;
  const selectedSet=new Set(dates);
  for(const date of [...dayCache.keys()]){
    if(!selectedSet.has(date))dayCache.delete(date);
  }
  render();
}

function summarize(rows){
  const minuteIps=new Map();
  const titles=new Set(),cliIps=new Set(),devices=new Set(),sessions=new Set();
  let segments=0,watchHours=0,deliveredHours=0;
  for(const row of rows){
    segments+=segmentCount(row);
    watchHours+=Number(row.request_watch_hours||0);
    deliveredHours+=Number(row.delivered_watch_hours||0);
    if(row.content_title)titles.add(row.content_title);
    if(row.device_id)devices.add(row.device_id);
    if(row.session_id)sessions.add(row.session_id);
    if(row.cli_ip){
      cliIps.add(row.cli_ip);
      if(!minuteIps.has(row.minute_ist))minuteIps.set(row.minute_ist,new Set());
      minuteIps.get(row.minute_ist).add(row.cli_ip);
    }
  }
  return {
    segments,titles:titles.size,cliIps:cliIps.size,devices:devices.size,sessions:sessions.size,
    watchHours,deliveredHours,
    peak:Math.max(0,...[...minuteIps.values()].map(ips=>ips.size)),
  };
}

function grouped(rows,key,limit=50){
  const groups=new Map();
  for(const row of rows){
    const label=key(row)||'Unknown / NA';
    const group=groups.get(label)||{label,segments:0,hours:0,ips:new Set(),devices:new Set(),sessions:new Set()};
    group.segments+=segmentCount(row);
    group.hours+=Number(row.request_watch_hours||0);
    if(row.cli_ip)group.ips.add(row.cli_ip);
    if(row.device_id)group.devices.add(row.device_id);
    if(row.session_id)group.sessions.add(row.session_id);
    groups.set(label,group);
  }
  const values=[...groups.values()].sort((a,b)=>b.hours-a.hours||b.segments-a.segments);
  return limit>0?values.slice(0,limit):values;
}

function isDavisCupRow(row){
  const title=normalized(row.content_title);
  const category=normalized(row.category_name).replaceAll('_',' ');
  return title.includes('davis cup')||category.includes('davis cup');
}

function performanceData(){
  const rows=activeRows.filter(row=>row.content_code);
  const firstActivityDate=rows.reduce(
    (first,row)=>!first||row.log_date<first?row.log_date:first,
    '',
  );
  const from=by('dateFrom').value||availableDates[0]||'';
  const to=by('dateTo').value||availableDates.at(-1)||'';
  const dates=availableDates.filter(date=>firstActivityDate&&date>=firstActivityDate&&date>=from&&date<=to);
  const byCode=new Map();
  const dayGroups=new Map(dates.map(date=>[date,{ips:new Set(),minutes:new Map(),hours:0}]));
  for(const row of rows){
    const code=row.content_code;
    const group=byCode.get(code)||{
      code,first:firstSeenByCode.get(code)||row.log_date,titles:new Map(),hours:0,ips:new Set(),minutes:new Map(),daily:new Map()
    };
    const title=row.content_title||'Unknown / NA';
    group.titles.set(title,(group.titles.get(title)||0)+segmentCount(row));
    group.hours+=Number(row.request_watch_hours||0);
    if(row.cli_ip)group.ips.add(row.cli_ip);
    if(row.cli_ip){
      if(!group.minutes.has(row.minute_ist))group.minutes.set(row.minute_ist,new Set());
      group.minutes.get(row.minute_ist).add(row.cli_ip);
    }
    const daily=group.daily.get(row.log_date)||{hours:0,ips:new Set()};
    daily.hours+=Number(row.request_watch_hours||0);
    if(row.cli_ip)daily.ips.add(row.cli_ip);
    group.daily.set(row.log_date,daily);
    byCode.set(code,group);
    const day=dayGroups.get(row.log_date);
    if(day){
      day.hours+=Number(row.request_watch_hours||0);
      if(row.cli_ip){
        day.ips.add(row.cli_ip);
        if(!day.minutes.has(row.minute_ist))day.minutes.set(row.minute_ist,new Set());
        day.minutes.get(row.minute_ist).add(row.cli_ip);
      }
    }
  }
  const videos=[...byCode.values()].map(group=>({
    ...group,
    title:[...group.titles.entries()].sort((a,b)=>b[1]-a[1]||a[0].localeCompare(b[0]))[0]?.[0]||group.code,
    peak:Math.max(0,...[...group.minutes.values()].map(ips=>ips.size)),
  })).sort((a,b)=>b.hours-a.hours);
  const daily=dates.map(date=>{
    const day=dayGroups.get(date);
    return {date,cliIps:day.ips.size,peak:Math.max(0,...[...day.minutes.values()].map(ips=>ips.size)),hours:day.hours};
  });
  return {rows,videos,daily};
}

function xmlText(value){
  return String(value??'').replace(/[&<>"']/g,char=>({
    '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&apos;'
  }[char]));
}

function excelColumn(index){
  let value=index+1,label='';
  while(value){value-=1;label=String.fromCharCode(65+(value%26))+label;value=Math.floor(value/26);}
  return label;
}

function worksheetXml(rows){
  const columnCount=Math.max(1,...rows.map(row=>row.length));
  const lastCell=`${excelColumn(columnCount-1)}${Math.max(1,rows.length)}`;
  const columns=Array.from({length:columnCount},(_,index)=>{
    const width=index===0?55:index===1?18:index===2?14:19;
    return `<col min="${index+1}" max="${index+1}" width="${width}" customWidth="1"/>`;
  }).join('');
  const sheetRows=rows.map((row,rowIndex)=>{
    const cells=row.map((value,columnIndex)=>{
      const reference=`${excelColumn(columnIndex)}${rowIndex+1}`;
      const style=rowIndex===0?' s="1"':'';
      if(typeof value==='number'&&Number.isFinite(value))return `<c r="${reference}"${style}><v>${Math.round(value)}</v></c>`;
      return `<c r="${reference}" t="inlineStr"${style}><is><t xml:space="preserve">${xmlText(value)}</t></is></c>`;
    }).join('');
    return `<row r="${rowIndex+1}">${cells}</row>`;
  }).join('');
  return `<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews><cols>${columns}</cols><sheetData>${sheetRows}</sheetData><autoFilter ref="A1:${lastCell}"/></worksheet>`;
}

function crc32(bytes){
  let crc=0xffffffff;
  for(const byte of bytes){
    crc^=byte;
    for(let bit=0;bit<8;bit+=1)crc=(crc>>>1)^((crc&1)?0xedb88320:0);
  }
  return (crc^0xffffffff)>>>0;
}

function littleEndian16(value){
  const bytes=new Uint8Array(2);
  new DataView(bytes.buffer).setUint16(0,value,true);
  return bytes;
}

function littleEndian32(value){
  const bytes=new Uint8Array(4);
  new DataView(bytes.buffer).setUint32(0,value>>>0,true);
  return bytes;
}

function joinBytes(parts){
  const output=new Uint8Array(parts.reduce((size,part)=>size+part.length,0));
  let offset=0;
  for(const part of parts){output.set(part,offset);offset+=part.length;}
  return output;
}

function xlsxBlob(sheets){
  const encoder=new TextEncoder();
  const sheetNames=Object.keys(sheets);
  const contentOverrides=sheetNames.map((_,index)=>`<Override PartName="/xl/worksheets/sheet${index+1}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>`).join('');
  const workbookSheets=sheetNames.map((name,index)=>`<sheet name="${xmlText(name)}" sheetId="${index+1}" r:id="rId${index+1}"/>`).join('');
  const workbookRelationships=sheetNames.map((_,index)=>`<Relationship Id="rId${index+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet${index+1}.xml"/>`).join('');
  const files={
    '[Content_Types].xml':`<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>${contentOverrides}</Types>`,
    '_rels/.rels':'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>',
    'xl/workbook.xml':`<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets>${workbookSheets}</sheets></workbook>`,
    'xl/_rels/workbook.xml.rels':`<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">${workbookRelationships}<Relationship Id="rId${sheetNames.length+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>`,
    'xl/styles.xml':'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font></fonts><fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/><bgColor indexed="64"/></patternFill></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf></cellXfs><cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles></styleSheet>',
  };
  sheetNames.forEach((name,index)=>{files[`xl/worksheets/sheet${index+1}.xml`]=worksheetXml(sheets[name]);});
  const localParts=[],centralParts=[];
  let localOffset=0,centralSize=0;
  for(const [name,content] of Object.entries(files)){
    const nameBytes=encoder.encode(name),data=encoder.encode(content),checksum=crc32(data);
    const localHeader=joinBytes([
      littleEndian32(0x04034b50),littleEndian16(20),littleEndian16(0x0800),littleEndian16(0),
      littleEndian16(0),littleEndian16(33),littleEndian32(checksum),littleEndian32(data.length),
      littleEndian32(data.length),littleEndian16(nameBytes.length),littleEndian16(0),nameBytes,
    ]);
    const centralHeader=joinBytes([
      littleEndian32(0x02014b50),littleEndian16(20),littleEndian16(20),littleEndian16(0x0800),
      littleEndian16(0),littleEndian16(0),littleEndian16(33),littleEndian32(checksum),
      littleEndian32(data.length),littleEndian32(data.length),littleEndian16(nameBytes.length),
      littleEndian16(0),littleEndian16(0),littleEndian16(0),littleEndian16(0),littleEndian32(0),
      littleEndian32(localOffset),nameBytes,
    ]);
    localParts.push(localHeader,data);
    centralParts.push(centralHeader);
    localOffset+=localHeader.length+data.length;
    centralSize+=centralHeader.length;
  }
  const end=joinBytes([
    littleEndian32(0x06054b50),littleEndian16(0),littleEndian16(0),
    littleEndian16(Object.keys(files).length),littleEndian16(Object.keys(files).length),
    littleEndian32(centralSize),littleEndian32(localOffset),littleEndian16(0),
  ]);
  return new Blob([...localParts,...centralParts,end],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
}

async function exportUniversalExcel(){
  const button=by('universalExportExcel');
  const originalLabel=button.textContent;
  button.disabled=true;
  button.textContent='Preparing Excel...';
  await yieldToBrowser();
  try{
    const {videos,daily}=performanceData();
    if(!videos.length){setLoadStatus('No selected video data to export.');return;}
    const dateHeaders=daily.map(day=>dateLabel(day.date).replace(' 2026',''));
    const watchRows=[
      ['Video Title','Video Key','First Seen',...dateHeaders.map(value=>`Watch Minutes - ${value}`),'Total Watch Minutes'],
      ...videos.map(video=>[
        video.title,video.code,video.first,
        ...daily.map(day=>Math.round((video.daily.get(day.date)?.hours||0)*60)),
        Math.round(video.hours*60),
      ]),
    ];
    const cliRows=[
      ['Video Title','Video Key','First Seen',...dateHeaders.map(value=>`Unique CLI IPs - ${value}`),'Total Unique CLI IPs'],
      ...videos.map(video=>[
        video.title,video.code,video.first,
        ...daily.map(day=>(video.daily.get(day.date)?.ips.size||0)),
        video.ips.size,
      ]),
    ];
    const from=by('dateFrom').value||daily[0]?.date||'start';
    const to=by('dateTo').value||daily.at(-1)?.date||'end';
    const url=URL.createObjectURL(xlsxBlob({'Watch Minutes':watchRows,'CLI IPs':cliRows}));
    const link=document.createElement('a');
    link.href=url;
    link.download=`VOD_Performance_${from}_to_${to}.xlsx`;
    document.body.append(link);
    link.click();
    link.remove();
    setTimeout(()=>URL.revokeObjectURL(url),1000);
    setLoadStatus(`Exported ${countFmt.format(videos.length)} selected videos to Excel.`);
  }catch(error){
    setLoadStatus(`Excel export failed: ${error.message}`);
    console.error(error);
  }finally{
    button.disabled=false;
    button.textContent=originalLabel;
  }
}

function chartSvg(daily,series,valueFormatter){
  if(!daily.length)return '<div class="chart-empty">No matching video data available</div>';
  const width=720,height=250,left=58,right=20,top=22,bottom=42;
  const plotWidth=width-left-right,plotHeight=height-top-bottom;
  const allValues=series.flatMap(item=>daily.map(day=>Number(item.value(day))||0));
  const maxValue=Math.max(1,...allValues);
  const x=index=>left+(daily.length===1?plotWidth/2:index*plotWidth/(daily.length-1));
  const y=value=>top+plotHeight-(Number(value)||0)*plotHeight/maxValue;
  const ticks=[0,.25,.5,.75,1].map(ratio=>{
    const value=maxValue*ratio,py=y(value);
    return `<line x1="${left}" y1="${py}" x2="${width-right}" y2="${py}" class="chart-grid"/><text x="${left-9}" y="${py+4}" text-anchor="end" class="chart-axis">${esc(valueFormatter(value))}</text>`;
  }).join('');
  const labels=daily.map((day,index)=>`<text x="${x(index)}" y="${height-15}" text-anchor="middle" class="chart-axis">${esc(dateLabel(day.date).replace(' 2026',''))}</text>`).join('');
  const lines=series.map(item=>{
    const points=daily.map((day,index)=>`${x(index)},${y(item.value(day))}`).join(' ');
    const dots=daily.map((day,index)=>`<circle cx="${x(index)}" cy="${y(item.value(day))}" r="4" fill="${item.color}"><title>${esc(dateLabel(day.date))}: ${esc(valueFormatter(item.value(day)))}</title></circle>`).join('');
    return `<polyline points="${points}" fill="none" stroke="${item.color}" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>${dots}`;
  }).join('');
  return `<svg viewBox="0 0 ${width} ${height}" role="img" aria-label="${esc(series.map(item=>item.label).join(' and '))} by day">${ticks}<line x1="${left}" y1="${top}" x2="${left}" y2="${height-bottom}" class="chart-axis-line"/><line x1="${left}" y1="${height-bottom}" x2="${width-right}" y2="${height-bottom}" class="chart-axis-line"/>${labels}${lines}</svg>`;
}

function renderDavisCup(){
  const {rows,videos,daily}=performanceData();
  const summary=summarize(rows);
  const first=videos.length?videos.reduce((value,video)=>video.first<value?video.first:value,videos[0].first):'';
  const values=[
    ['Videos',countFmt.format(videos.length)],
    ['First Seen',first?dateLabel(first):'Unavailable'],
    ['Unique CLI IPs',countFmt.format(summary.cliIps)],
    ['Est. Watch Time',duration(summary.watchHours)],
    ['Peak Concurrent CLI IPs',countFmt.format(summary.peak)],
  ];
  by('davisKpis').innerHTML=values.map(([label,value])=>`<div class="kpi"><span>${label}</span><strong>${value}</strong></div>`).join('');
  const audienceSeries=[
    {label:'Distinct CLI IPs',color:'#0f766e',value:day=>day.cliIps},
  ];
  by('davisAudienceLegend').innerHTML=audienceSeries.map(item=>`<span><i style="background:${item.color}"></i>${esc(item.label)}</span>`).join('');
  by('davisAudienceChart').innerHTML=chartSvg(daily,audienceSeries,value=>countFmt.format(Math.round(value)));
  const watchSeries=[{label:'Estimated watch hours',color:'#2563eb',value:day=>day.hours}];
  by('davisWatchLegend').innerHTML=watchSeries.map(item=>`<span><i style="background:${item.color}"></i>${esc(item.label)}</span>`).join('');
  by('davisWatchChart').innerHTML=chartSvg(daily,watchSeries,value=>duration(value));
  by('davisHeaderStatus').textContent=daily.length
    ?`Video performance: ${dateLabel(daily[0].date)} to ${dateLabel(daily.at(-1).date)}`
    :'No matching video data available';
  by('davisTitlesHead').innerHTML=`<tr><th>Video title</th><th>Video key</th><th>First seen</th>${daily.map(day=>`<th class="num">${esc(dateLabel(day.date).replace(' 2026',''))}<br>Watch / CLI IPs</th>`).join('')}<th class="num">Total watch</th><th class="num">CLI IPs</th><th class="num">Peak concurrent</th></tr>`;
  by('davisTitles').innerHTML=rowsHtml(videos,video=>`<tr><td>${esc(video.title)}</td><td>${esc(video.code)}</td><td>${esc(dateLabel(video.first))}</td>${daily.map(day=>{const value=video.daily.get(day.date)||{hours:0,ips:new Set()};return `<td class="num daily-metric"><strong>${duration(value.hours)}</strong><span>${countFmt.format(value.ips.size)} CLI IPs</span></td>`;}).join('')}<td class="num">${duration(video.hours)}</td><td class="num">${countFmt.format(video.ips.size)}</td><td class="num">${countFmt.format(video.peak)}</td></tr>`,daily.length+6);
}

function selectDavisCup(){
  filterState.code=new Set(davisCodes);
  updatePickerToggle('code');
  renderPickerOptions('code');
  scheduleRender();
}

function setDavisMode(enabled){
  davisMode=enabled;
  by('overviewView').hidden=enabled;
  by('davisView').hidden=!enabled;
  by('davisViewToggle').textContent=enabled?'Back to overview':'Video performance';
  by('davisViewToggle').setAttribute('aria-pressed',String(enabled));
  document.body.classList.toggle('davis-mode',enabled);
  scheduleRender(0);
}

function rowsHtml(items,renderer,colspan){
  return items.length?items.map(renderer).join(''):`<tr><td colspan="${colspan}" class="empty-cell">No matching data</td></tr>`;
}

function renderKpis(rows){
  const summary=summarize(rows);
  const values=[
    ['Unique CLI IPs',countFmt.format(summary.cliIps)],
    ['Est. Watch Time (All Statuses)',duration(summary.watchHours)],
    ['Delivered Segment Time',duration(summary.deliveredHours)],
    ['Device IDs',countFmt.format(summary.devices)],
    ['Session IDs',countFmt.format(summary.sessions)],
    ['Media Segments',countFmt.format(summary.segments)],
    ['Content Titles',countFmt.format(summary.titles)],
    ['Peak Media CLI IPs',countFmt.format(summary.peak)],
  ];
  by('kpis').innerHTML=values.map(([label,value])=>`<div class="kpi"><span>${label}</span><strong>${value}</strong></div>`).join('');
}

function renderTitles(rows){
  const titleRows=grouped(rows,row=>[row.content_title,row.category_name].filter(Boolean).join(' | '));
  by('titles').innerHTML=rowsHtml(titleRows,group=>`<tr><td>${esc(group.label)}</td><td class="num">${countFmt.format(group.segments)}</td><td class="num">${countFmt.format(group.ips.size)}</td><td class="num">${duration(group.hours)}</td><td class="num">${countFmt.format(group.devices.size)}</td><td class="num">${countFmt.format(group.sessions.size)}</td></tr>`,6);
}

function renderRegions(rows){
  const regionRows=grouped(rows,row=>[row.country,row.state].filter(Boolean).join(' / '));
  by('regions').innerHTML=rowsHtml(regionRows,group=>`<tr><td>${esc(group.label)}</td><td class="num">${countFmt.format(group.ips.size)}</td><td class="num">${countFmt.format(group.sessions.size)}</td><td class="num">${countFmt.format(group.devices.size)}</td><td class="num">${duration(group.hours)}</td></tr>`,5);
}

function renderDevices(rows){
  const deviceRows=grouped(rows,decodedDeviceLabel);
  by('devices').innerHTML=rowsHtml(deviceRows,group=>`<tr><td>${esc(group.label)}</td><td class="num">${countFmt.format(group.ips.size)}</td><td class="num">${countFmt.format(group.devices.size)}</td><td class="num">${countFmt.format(group.sessions.size)}</td><td class="num">${duration(group.hours)}</td><td class="num">${countFmt.format(group.segments)}</td></tr>`,6);
}

function renderRoutes(rows){
  const routeRows=grouped(rows,row=>[row.req_host,row.content_code].filter(Boolean).join(' / '));
  by('routes').innerHTML=rowsHtml(routeRows,group=>`<tr><td>${esc(group.label)}</td><td class="num">${countFmt.format(group.segments)}</td><td class="num">${countFmt.format(group.ips.size)}</td><td class="num">${duration(group.hours)}</td></tr>`,4);
}

function renderMinutes(rows){
  const minuteRows=grouped(rows,row=>row.minute_ist,0)
    .sort((a,b)=>b.ips.size-a.ips.size||b.segments-a.segments)
    .slice(0,50);
  by('minutes').innerHTML=rowsHtml(minuteRows,group=>`<tr><td>${esc(group.label)}</td><td class="num">${countFmt.format(group.ips.size)}</td><td class="num">${countFmt.format(group.segments)}</td></tr>`,3);
}

function renderSessions(rows){
  const groups=new Map();
  for(const row of rows){
    const hasSession=Boolean(row.session_id);
    const session=hasSession?row.session_id:'Missing session ID';
    const key=[row.content_title,row.category_name,session,row.device_id||'',row.cli_ip||'',row.state||''].join('\u0001');
    const rowLast=row.last_request_ist||row.request_ist;
    const group=groups.get(key)||{title:row.content_title||'Unknown / NA',category:row.category_name||'Unknown / NA',session,device:row.device_id||'Unknown / NA',ip:row.cli_ip||'Unknown / NA',state:row.state||'Unknown / NA',first:row.request_ist,last:rowLast,watch:0,hasSession};
    if(row.request_ist<group.first)group.first=row.request_ist;
    if(rowLast>group.last)group.last=rowLast;
    group.watch+=Number(row.request_watch_hours||0);
    groups.set(key,group);
  }
  const sessions=[...groups.values()].sort((a,b)=>b.watch-a.watch).slice(0,1000);
  by('sessionDetails').innerHTML=rowsHtml(sessions,group=>{
    const start=Date.parse(group.first.replace(' ','T')),end=Date.parse(group.last.replace(' ','T'));
    const span=group.hasSession&&Number.isFinite(start)&&Number.isFinite(end)?duration(Math.max(0,(end-start)/3600000)):'Unavailable';
    return `<tr><td>${esc(group.title)}</td><td>${esc(group.category)}</td><td>${esc(group.session)}</td><td>${esc(group.device)}</td><td>${esc(group.ip)}</td><td>${esc(group.state)}</td><td>${esc(group.first)}</td><td>${esc(group.last)}</td><td class="num">${span}</td><td class="num">${duration(group.watch)}</td></tr>`;
  },10);
}

function renderLedger(){
  const query=normalized(by('requestSearch').value);
  const rows=activeRows.filter(row=>!query||normalized(Object.values(row).join(' ')).includes(query)).slice(0,1000);
  const visibleSegments=rows.reduce((sum,row)=>sum+segmentCount(row),0);
  const totalSegments=activeRows.reduce((sum,row)=>sum+segmentCount(row),0);
  by('ledgerCount').textContent=`Showing ${countFmt.format(rows.length)} activity rows / ${countFmt.format(visibleSegments)} of ${countFmt.format(totalSegments)} matching segments`;
  by('ledger').innerHTML=rowsHtml(rows,row=>`<tr><td>${esc(row.request_ist)}</td><td class="num">${countFmt.format(segmentCount(row))}</td><td>${esc(row.content_title)}</td><td>${esc(row.category_name)}</td><td>${esc(row.status_code)}</td><td>${esc(row.channel)}</td><td>${esc(decodedDeviceLabel(row))}</td><td>${esc([row.platform,row.device].filter(Boolean).join(' / '))}</td><td>${esc(row.cli_ip)}</td><td>${esc(row.device_id)}</td><td>${esc(row.session_id)}</td><td>${esc([row.country,row.state,row.city].filter(Boolean).join(' / '))}</td><td>${esc([row.req_host,row.req_path].filter(Boolean).join(' / '))}</td></tr>`,13);
}

function render(){
  const version=++renderVersion;
  activeRows=selectedRows();
  setLoadStatus('Updating dashboard panels...',true);
  renderKpis(activeRows);
  const tasks=[
    ()=>renderTitles(activeRows),
    ()=>renderRegions(activeRows),
    ()=>renderDevices(activeRows),
    ()=>renderRoutes(activeRows),
    ()=>renderMinutes(activeRows),
    ()=>renderSessions(activeRows),
    ()=>renderLedger(),
  ];
  if(davisMode)tasks.push(()=>renderDavisCup());
  const runNext=()=>{
    if(version!==renderVersion)return;
    const task=tasks.shift();
    if(!task){
      const dateWord=activeDays.length===1?'date':'dates';
      setLoadStatus(`${countFmt.format(activeDays.length)} ${dateWord} loaded | ${countFmt.format(activeRows.length)} matching activity rows`);
      return;
    }
    task();
    setTimeout(runNext,0);
  };
  setTimeout(runNext,0);
}

function scheduleRender(delay=80){
  renderVersion+=1;
  clearTimeout(renderTimer);
  renderTimer=setTimeout(render,delay);
}

for(const type of ['title','category','code'])populatePicker(type);
const firstDate=availableDates[0]||'',lastDate=availableDates.at(-1)||'';
const defaultWindow=Math.max(1,Number(dashboardIndex.default_window_days)||7);
const defaultFrom=availableDates[Math.max(0,availableDates.length-defaultWindow)]||firstDate;
for(const id of ['dateFrom','dateTo']){by(id).min=firstDate;by(id).max=lastDate;}
by('dateFrom').value=defaultFrom;
by('dateTo').value=lastDate;
by('dateFrom').addEventListener('change',()=>{by('dateFrom').value=nearestAvailableDate(by('dateFrom').value,'from');if(by('dateFrom').value>by('dateTo').value)by('dateTo').value=by('dateFrom').value;renderDateAvailability();refreshData().catch(error=>{setLoadStatus(error.message);console.error(error);});});
by('dateTo').addEventListener('change',()=>{by('dateTo').value=nearestAvailableDate(by('dateTo').value,'to');if(by('dateTo').value<by('dateFrom').value)by('dateFrom').value=by('dateTo').value;renderDateAvailability();refreshData().catch(error=>{setLoadStatus(error.message);console.error(error);});});
for(const type of ['title','category','code']){
  by(`${type}Toggle`).addEventListener('click',event=>{event.stopPropagation();togglePicker(type);});
  by(`${type}Search`).addEventListener('input',()=>renderPickerOptions(type));
  by(`${type}Search`).addEventListener('keydown',event=>{
    if(event.key==='Enter'){
      const first=by(`${type}Options`).querySelector('input');
      if(first){event.preventDefault();first.click();}
    }
    if(event.key==='Escape'){closePickers();by(`${type}Toggle`).focus();}
  });
  by(`${type}All`).addEventListener('click',()=>setAll(type,true));
  by(`${type}Clear`).addEventListener('click',()=>setAll(type,false));
}
let ledgerTimer=0;
by('requestSearch').addEventListener('input',()=>{clearTimeout(ledgerTimer);ledgerTimer=setTimeout(renderLedger,160);});
by('davisPreset').addEventListener('click',selectDavisCup);
by('davisViewToggle').addEventListener('click',()=>setDavisMode(!davisMode));
by('universalExportExcel').addEventListener('click',exportUniversalExcel);
document.addEventListener('click',event=>{if(!event.target.closest('.picker'))closePickers();});
document.addEventListener('keydown',event=>{if(event.key==='Escape')closePickers();});
renderDateAvailability();
await refreshData();
})().catch(error=>{
  document.body.innerHTML=`<main><section><h2>Dashboard could not load</h2><p class="note">${String(error?.message||error)}</p></section></main>`;
  console.error(error);
});
</script>"""


def render_dashboard_html(
    index: dict[str, object],
    inline_payloads: dict[str, str],
    data_directory: str,
    davis_workbook_name: str = "Davis_Cup_Performance.xlsx",
) -> str:
    template = """<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>VOD Stream Analysis</title>
<style>
*{box-sizing:border-box}:root{--canvas:#f4f6f8;--surface:#fff;--ink:#172033;--muted:#64748b;--line:#dbe3ea;--teal:#0f766e;--orange:#b45309}
body{margin:0;background:var(--canvas);color:var(--ink);font:13px Inter,Segoe UI,Arial,sans-serif}button,input{font:inherit}
header{position:sticky;top:0;z-index:10;padding:12px max(16px,calc((100% - 1440px)/2));background:#fff;border-top:3px solid #333;border-bottom:1px solid var(--line)}
.filters{display:grid;grid-template-columns:28fr 24fr 24fr 24fr;gap:10px;align-items:center;width:100%}.date-range{display:grid;grid-template-columns:minmax(0,1fr) auto minmax(0,1fr);align-items:center;gap:7px;min-width:0}.date-range span{color:var(--muted);font-size:10px;font-weight:800;text-align:center;text-transform:uppercase}
.date-availability{margin-top:7px;color:var(--muted);font-size:10px}.date-availability summary{display:inline-flex;align-items:center;gap:6px;font-weight:800;cursor:pointer;list-style-position:inside}.date-availability summary:hover{color:var(--teal)}.date-selection-status{margin-left:8px;font-weight:600}.available-date-list{display:flex;flex-wrap:wrap;gap:4px;margin-top:7px;padding-top:7px;border-top:1px solid #edf1f4}.available-date-list time{padding:3px 5px;border:1px solid #dbe3ea;background:#f8fafc;color:var(--muted);font-size:9px;font-weight:700}.available-date-list time.selected{border-color:#75aaa5;background:#e6f2f0;color:var(--teal)}
input{height:33px;min-width:0;border:1px solid #b8c7d8;border-radius:4px;background:#fff;color:var(--ink);padding:6px 8px}.date-range input{width:100%}.picker{position:relative;min-width:0}
.picker-toggle{position:relative;width:100%;height:33px;overflow:hidden;border:1px solid #b8c7d8;border-radius:4px;padding:6px 30px 6px 9px;background:#fff;color:var(--ink);text-align:left;text-overflow:ellipsis;white-space:nowrap;cursor:pointer}.picker-toggle::after{content:'v';position:absolute;right:10px;color:var(--muted);font-weight:800}.picker-toggle:hover,.picker-toggle:focus,.picker.open .picker-toggle{border-color:var(--teal);outline:none;box-shadow:0 0 0 2px rgba(15,118,110,.12)}
.picker-menu{display:none;position:absolute;z-index:30;top:calc(100% + 4px);left:0;width:max(100%,320px);max-width:calc(100vw - 24px);overflow:hidden;border:1px solid #b8c7d8;border-radius:5px;background:#fff;box-shadow:0 12px 28px rgba(15,23,42,.18)}.picker.open .picker-menu{display:block}.picker-search{display:block;width:calc(100% - 12px);margin:6px}.picker-actions{display:flex;align-items:center;gap:6px;padding:0 6px 6px;border-bottom:1px solid #edf1f4}.picker-actions button{border:1px solid #cbd5e1;border-radius:3px;background:#f8fafc;color:var(--ink);padding:4px 8px;font-size:10px;font-weight:800;cursor:pointer}.picker-actions button:hover{border-color:var(--teal);background:#e6f2f0;color:var(--teal)}.picker-result-count{margin-left:auto;color:var(--muted);font-size:10px;font-weight:700}.picker-options{max-height:260px;overflow:auto;overscroll-behavior:contain;padding:4px}.picker-option{display:flex;align-items:flex-start;gap:7px;margin:0;padding:7px 6px;border-radius:3px;color:var(--ink);font-size:11px;font-weight:600;cursor:pointer}.picker-option:hover{background:#eef6f5}.picker-option input{width:14px;height:14px;flex:0 0 auto;margin:0;padding:0}.picker-option span{overflow-wrap:anywhere}.picker-empty{padding:14px 8px;color:var(--muted);font-size:11px;text-align:center}.native-filter{position:absolute!important;width:1px!important;height:1px!important;overflow:hidden!important;opacity:0!important;pointer-events:none!important;clip-path:inset(50%)}
main{max-width:1440px;margin:auto;padding:16px}.kpis{display:grid;grid-template-columns:repeat(8,minmax(0,1fr));gap:10px}.kpi{min-width:0;border:1px solid var(--line);border-top:3px solid var(--teal);background:var(--surface);padding:10px}.kpi:nth-child(2){border-top-color:var(--orange)}.kpi span{display:block;color:var(--muted);font-size:9px;font-weight:800;text-transform:uppercase}.kpi strong{display:block;overflow-wrap:anywhere;margin-top:6px;font-size:22px;font-variant-numeric:tabular-nums}
.view-bar{display:flex;align-items:center;justify-content:flex-end;gap:8px;margin-bottom:10px}.load-status{margin-right:auto;color:var(--muted);font-size:10px;font-weight:700}.load-status.busy::before{content:'';display:inline-block;width:10px;height:10px;margin-right:6px;border:2px solid #cbd5e1;border-top-color:var(--teal);border-radius:50%;vertical-align:-2px;animation:spin .8s linear infinite}@keyframes spin{to{transform:rotate(360deg)}}.view-toggle{height:33px;border:1px solid var(--teal);border-radius:4px;background:#fff;color:var(--teal);padding:6px 12px;font-weight:800;cursor:pointer}.view-toggle:hover,.view-toggle[aria-pressed="true"],.preset-button:hover{background:var(--teal);color:#fff}.preset-button{height:33px;border:1px solid #b8c7d8;border-radius:4px;background:#fff;color:var(--ink);padding:6px 12px;font-weight:800;cursor:pointer}.davis-header-status{display:none;margin-top:7px;color:var(--ink);font-size:12px;font-weight:800}.davis-mode .date-availability{display:none}.davis-mode .davis-header-status{display:block}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px}section{min-width:0;border:1px solid var(--line);background:var(--surface)}h2{margin:0;padding:10px 12px;border-bottom:1px solid var(--line);font-size:13px}.section-heading{display:flex;align-items:center;justify-content:space-between;gap:10px;border-bottom:1px solid var(--line);padding:7px 10px 7px 12px}.section-heading h2{border:0;padding:0}.export-button{display:inline-flex;align-items:center;height:33px;border:1px solid #b8c7d8;border-radius:4px;background:#fff;color:var(--ink);padding:6px 12px;font-size:11px;font-weight:800;text-decoration:none;cursor:pointer}.export-button:hover,.export-button:focus{border-color:var(--teal);color:var(--teal);outline:none;box-shadow:0 0 0 2px rgba(15,118,110,.12)}.export-button:disabled{cursor:wait;opacity:.65}.wide-section{margin-top:12px}.table-wrap{max-height:335px;overflow:auto}table{width:100%;border-collapse:collapse;font-size:11px}th,td{padding:7px 9px;border-bottom:1px solid #edf1f4;text-align:left;vertical-align:top}th{position:sticky;top:0;z-index:1;background:#f8fafc;color:var(--muted);font-size:9px;text-transform:uppercase}td.num,th.num{text-align:right;font-variant-numeric:tabular-nums}.empty-cell{padding:18px;color:var(--muted);text-align:center}.event-controls{display:grid;grid-template-columns:minmax(240px,520px) 1fr;align-items:center;gap:10px;padding:9px 12px;border-bottom:1px solid var(--line)}.ledger-count{color:var(--muted);font-size:10px;text-align:right}.note{margin:12px 0 0;color:var(--muted);font-size:11px}
.davis-kpis{grid-template-columns:repeat(5,minmax(0,1fr))}.chart-panel{padding-bottom:8px}.chart-legend{display:flex;flex-wrap:wrap;gap:14px;padding:9px 12px 0;color:var(--muted);font-size:10px;font-weight:700}.chart-legend span{display:inline-flex;align-items:center;gap:5px}.chart-legend i{width:14px;height:3px}.chart{width:100%;min-height:250px;padding:0 6px}.chart svg{display:block;width:100%;height:auto;min-height:250px}.chart-grid{stroke:#e7edf2;stroke-width:1}.chart-axis-line{stroke:#94a3b8;stroke-width:1}.chart-axis{fill:#64748b;font:10px Inter,Segoe UI,Arial,sans-serif}.chart-empty{display:grid;min-height:250px;place-items:center;color:var(--muted)}.davis-table{max-height:440px}.daily-metric strong,.daily-metric span{display:block;white-space:nowrap}.daily-metric span{margin-top:2px;color:var(--muted);font-size:9px}.davis-note{padding:0 2px}
@media(max-width:1000px){.kpis,.davis-kpis{grid-template-columns:repeat(3,minmax(0,1fr))}.grid{grid-template-columns:1fr}}@media(max-width:760px){.filters{grid-template-columns:1fr 1fr}.date-range{grid-column:span 2}.kpis,.davis-kpis{grid-template-columns:repeat(2,minmax(0,1fr))}}@media(max-width:520px){header{position:static}.filters{grid-template-columns:1fr}.date-range{grid-column:auto}.picker-menu{width:100%}main{padding:10px}.event-controls{grid-template-columns:1fr}.ledger-count{text-align:left}.view-bar{display:grid;grid-template-columns:1fr 1fr}.load-status{grid-column:1/-1}.view-toggle,.preset-button,.export-button{width:100%}}
</style></head><body>
<header><div class="filters">
<div class="date-range"><input id="dateFrom" type="date" aria-label="From date"><span>to</span><input id="dateTo" type="date" aria-label="To date"></div>
<div class="picker" id="titlePicker"><button class="picker-toggle" id="titleToggle" type="button" aria-haspopup="listbox" aria-expanded="false">All titles</button><div class="picker-menu"><input class="picker-search" id="titleSearch" type="search" autocomplete="off" placeholder="Search titles" aria-label="Search title options"><div class="picker-actions"><button id="titleAll" type="button">Select all</button><button id="titleClear" type="button">Clear</button><span class="picker-result-count" id="titleResultCount"></span></div><div class="picker-options" id="titleOptions" role="listbox" aria-multiselectable="true"></div></div><select class="native-filter" id="titleFilter" multiple aria-hidden="true" tabindex="-1"></select></div>
<div class="picker" id="categoryPicker"><button class="picker-toggle" id="categoryToggle" type="button" aria-haspopup="listbox" aria-expanded="false">All categories</button><div class="picker-menu"><input class="picker-search" id="categorySearch" type="search" autocomplete="off" placeholder="Search categories" aria-label="Search category options"><div class="picker-actions"><button id="categoryAll" type="button">Select all</button><button id="categoryClear" type="button">Clear</button><span class="picker-result-count" id="categoryResultCount"></span></div><div class="picker-options" id="categoryOptions" role="listbox" aria-multiselectable="true"></div></div><select class="native-filter" id="categoryFilter" multiple aria-hidden="true" tabindex="-1"></select></div>
<div class="picker" id="codePicker"><button class="picker-toggle" id="codeToggle" type="button" aria-haspopup="listbox" aria-expanded="false">All video keys</button><div class="picker-menu"><input class="picker-search" id="codeSearch" type="search" autocomplete="off" placeholder="Search video keys" aria-label="Search video key options"><div class="picker-actions"><button id="codeAll" type="button">Select all</button><button id="codeClear" type="button">Clear</button><span class="picker-result-count" id="codeResultCount"></span></div><div class="picker-options" id="codeOptions" role="listbox" aria-multiselectable="true"></div></div><select class="native-filter" id="codeFilter" multiple aria-hidden="true" tabindex="-1"></select></div>
</div><details class="date-availability"><summary id="dateAvailabilitySummary">Available data dates</summary><span class="date-selection-status" id="dateSelectionStatus"></span><div class="available-date-list" id="availableDateList"></div></details><div class="davis-header-status" id="davisHeaderStatus">Video performance follows the active filters</div></header>
<main><div class="view-bar"><span class="load-status busy" id="loadStatus" role="status" aria-live="polite">Preparing dashboard...</span><button class="export-button" id="universalExportExcel" type="button">Export selected Excel</button><button class="preset-button" id="davisPreset" type="button">Select Davis Cup</button><button class="view-toggle" id="davisViewToggle" type="button" aria-pressed="false">Video performance</button></div><div id="overviewView"><div class="kpis" id="kpis"></div><div class="grid">
<section><h2>Content Titles</h2><div class="table-wrap"><table><thead><tr><th>Title / Category</th><th class="num">Segments</th><th class="num">CLI IPs</th><th class="num">Watch time</th><th class="num">Devices</th><th class="num">Sessions</th></tr></thead><tbody id="titles"></tbody></table></div></section>
<section><h2>Regions</h2><div class="table-wrap"><table><thead><tr><th>Country / State</th><th class="num">CLI IPs</th><th class="num">Sessions</th><th class="num">Devices</th><th class="num">Watch time</th></tr></thead><tbody id="regions"></tbody></table></div></section>
<section><h2>Decoded Device and Platform</h2><div class="table-wrap"><table><thead><tr><th>Device type / Brand and model / OS / App platform</th><th class="num">CLI IPs</th><th class="num">Device IDs</th><th class="num">Sessions</th><th class="num">Watch time</th><th class="num">Segments</th></tr></thead><tbody id="devices"></tbody></table></div></section>
<section><h2>Delivery Routes</h2><div class="table-wrap"><table><thead><tr><th>Host / Content code</th><th class="num">Segments</th><th class="num">CLI IPs</th><th class="num">Watch time</th></tr></thead><tbody id="routes"></tbody></table></div></section>
<section><h2>Minute Concurrency</h2><div class="table-wrap"><table><thead><tr><th>IST minute</th><th class="num">Concurrent CLI IPs</th><th class="num">Segments</th></tr></thead><tbody id="minutes"></tbody></table></div></section>
</div>
<section class="wide-section"><h2>Session Detail</h2><div class="table-wrap"><table><thead><tr><th>Title</th><th>Category</th><th>Session ID</th><th>Device ID</th><th>CLI IP</th><th>State</th><th>First request</th><th>Last request</th><th class="num">Request span</th><th class="num">Segment watch time</th></tr></thead><tbody id="sessionDetails"></tbody></table></div></section>
<section class="wide-section"><h2>Segment Activity Ledger</h2><div class="event-controls"><input id="requestSearch" type="search" autocomplete="off" placeholder="Search title, category, status, CLI IP, decoded device, raw device, session, region, host, or path"><span class="ledger-count" id="ledgerCount"></span></div><div class="table-wrap"><table><thead><tr><th>First request IST</th><th class="num">Segments</th><th>Title</th><th>Category</th><th>Status</th><th>Channel</th><th>Decoded device</th><th>App platform / Raw device</th><th>CLI IP</th><th>Device ID</th><th>Session ID</th><th>Region</th><th>Host / Sample Path</th></tr></thead><tbody id="ledger"></tbody></table></div></section>
<p class="note">Estimated watch time counts every mapped VOD .ts request, regardless of status, at 6 seconds per request. Delivered segment time counts only 2xx requests. Compact dashboard rows combine repeated segment activity at minute, viewer, content, identity, and geography grain; segment totals remain weighted by the original requests. Both time measures are request-based estimates rather than player telemetry; retries and failed requests can overstate viewing. Request span is shown only when a real session ID exists.</p></div>
<div id="davisView" hidden>
<div class="kpis davis-kpis" id="davisKpis"></div>
<div class="grid">
<section class="chart-panel"><h2>Daily Distinct CLI IPs</h2><div class="chart-legend" id="davisAudienceLegend"></div><div class="chart" id="davisAudienceChart"></div></section>
<section class="chart-panel"><h2>Daily Estimated Watch Time</h2><div class="chart-legend" id="davisWatchLegend"></div><div class="chart" id="davisWatchChart"></div></section>
</div>
<section class="wide-section"><h2>Video Performance</h2><div class="table-wrap davis-table"><table><thead id="davisTitlesHead"></thead><tbody id="davisTitles"></tbody></table></div></section>
<p class="note davis-note"><strong>First seen</strong> is the earliest CDN request date available in this dashboard and is used as the upload-date proxy; it is not CMS upload metadata. Estimated watch time counts all mapped .ts requests at 6 seconds each. Peak concurrency is the largest distinct CLI IP count observed in one IST minute.</p>
</div></main>
__SCRIPT__
</body></html>"""
    return (
        template.replace("__SCRIPT__", dashboard_script())
        .replace("__DASHBOARD_INDEX__", script_json(index))
        .replace("__INLINE_PAYLOADS__", script_json(inline_payloads))
        .replace("__DATA_DIRECTORY__", script_json(data_directory))
        .replace("__DAVIS_WORKBOOK__", davis_workbook_name)
    )


def render_html(
    rows: list[dict[str, str]],
    davis_workbook_name: str = "Davis_Cup_Performance.xlsx",
) -> str:
    """Build a self-contained dashboard for small previews and unit tests."""
    index, days = prepare_dashboard_data(rows)
    inline_payloads = {
        date: encode_browser_rows(day_rows)
        for date, day_rows in days.items()
    }
    return render_dashboard_html(index, inline_payloads, "", davis_workbook_name)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the incremental VOD segment dashboard.")
    parser.add_argument("--events", type=Path, default=DEFAULT_EVENTS)
    parser.add_argument("--append", type=Path, action="append", default=[])
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--davis-xlsx", type=Path)
    parser.add_argument("--data-dir", type=Path)
    parser.add_argument(
        "--data-url",
        help="Relative browser path to --data-dir; defaults to the data directory name.",
    )
    args = parser.parse_args()
    existing = read_events(args.events) if args.events.exists() else []
    incoming = [row for path in args.append for row in read_events(path)]
    rows = merge_events(existing, incoming) if incoming else existing
    if incoming:
        write_events(args.events, rows)
    if not rows:
        raise ValueError("No VOD event rows were supplied. Provide --append for the first extract.")
    args.out.parent.mkdir(parents=True, exist_ok=True)
    davis_xlsx = args.davis_xlsx or args.out.with_name("Davis_Cup_Performance.xlsx")
    data_dir = args.data_dir or args.out.with_name(f"{args.out.stem}_data")
    data_url = args.data_url or data_dir.name
    write_davis_workbook(davis_xlsx, rows)
    index, days = prepare_dashboard_data(rows)
    write_dashboard_data(data_dir, days)
    args.out.write_text(
        render_dashboard_html(index, {}, data_url, davis_xlsx.name),
        encoding="utf-8",
    )
    latest = max((row.get("log_date", "") for row in rows), default="no dates")
    segments = int(sum(number(row.get("segment_count", "1")) for row in rows))
    print(
        f"Wrote {args.out} with {segments:,} media segments "
        f"across {len(rows):,} activity rows through {latest}. "
        f"Lazy data: {data_dir}. Davis Cup workbook: {davis_xlsx}."
    )


if __name__ == "__main__":
    main()
