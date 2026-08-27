from __future__ import annotations

import csv
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from zoneinfo import ZoneInfo

import duckdb
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
TOOLS = ROOT / "ETL" / "src" / "tools"
sys.path.insert(0, str(TOOLS))

from build_vod_query_dashboard import (  # noqa: E402
    canonicalize_content_titles,
    merge_events,
    render_html,
    watch_minutes,
    write_davis_workbook,
)


def epoch(ist_time: str) -> str:
    value = datetime.strptime(ist_time, "%Y-%m-%d %H:%M:%S").replace(
        tzinfo=ZoneInfo("Asia/Kolkata")
    )
    return str(value.timestamp())


def source_row(
    request_time: str,
    cli_ip: str,
    path: str,
    query: str = "",
    status: str = "200",
) -> dict[str, str]:
    return {
        "reqTimeSec": epoch(request_time),
        "reqHost": "vod.example.net",
        "reqPath": path,
        "queryStr": query,
        "cliIP": cli_ip,
        "country": "IN",
        "state": "Delhi",
        "city": "NEWDELHI",
        "asn": "64500",
        "statusCode": status,
        "cacheStatus": "1",
        "UA": "TestUA/1.0",
    }


def write_source_parquet(path: Path) -> None:
    content_path = "library/2026/08/content-code/asset"
    rows = [
        source_row(
            "2026-08-25 10:00:00",
            "10.0.0.1",
            f"{content_path}/master.m3u8",
            "content_type=Vod&content_title=Example%20Title&category_name=Drama"
            "&platform=android&device=Phone-A&session_id=session-a&device_id=device-a",
        ),
        source_row(
            "2026-08-25 10:01:00",
            "10.0.0.2",
            f"{content_path}/master.m3u8",
            "content_type=Vod&content_title=Example%20Title&category_name=Drama"
            "&platform=ios&device=Phone-B&session_id=session-b&device_id=device-b",
        ),
        source_row("2026-08-25 10:02:00", "10.0.0.1", f"{content_path}/segment-1.ts"),
        source_row("2026-08-25 10:02:05", "10.0.0.2", f"{content_path}/segment-1.ts"),
        source_row(
            "2026-08-25 10:02:10",
            "10.0.0.1",
            f"{content_path}/failed.ts",
            status="404",
        ),
    ]
    connection = duckdb.connect()
    connection.execute(
        """
        CREATE TABLE source (
            reqTimeSec VARCHAR, reqHost VARCHAR, reqPath VARCHAR,
            queryStr VARCHAR, cliIP VARCHAR, country VARCHAR,
            state VARCHAR, city VARCHAR, asn VARCHAR,
            statusCode VARCHAR, cacheStatus VARCHAR, UA VARCHAR
        )
        """
    )
    connection.executemany(
        "INSERT INTO source VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [tuple(row.values()) for row in rows],
    )
    connection.execute("COPY source TO ? (FORMAT PARQUET)", [str(path)])
    connection.close()


def write_ua_lookup(path: Path) -> None:
    connection = duckdb.connect()
    connection.execute(
        """
        COPY (
            SELECT
                'TestUA/1.0'::VARCHAR AS ua_norm,
                'decoded_local'::VARCHAR AS decode_status,
                'high'::VARCHAR AS confidence,
                'smart_tv'::VARCHAR AS device_type,
                'TV'::VARCHAR AS form_factor,
                'Example'::VARCHAR AS brand,
                'Living Room TV'::VARCHAR AS model,
                'ExampleOS'::VARCHAR AS os_name,
                '1.0'::VARCHAR AS os_version,
                'Example Browser'::VARCHAR AS browser_name,
                '2.0'::VARCHAR AS browser_version,
                'Example Player'::VARCHAR AS app_player,
                ''::VARCHAR AS api_device_type,
                ''::VARCHAR AS api_brand,
                ''::VARCHAR AS api_model,
                ''::VARCHAR AS api_os_name,
                ''::VARCHAR AS api_browser_name,
                ''::VARCHAR AS api_browser_version,
                '2026-08-25T00:00:00Z'::VARCHAR AS decoded_at_utc
        ) TO ? (FORMAT PARQUET)
        """,
        [str(path)],
    )
    connection.close()


def test_export_keeps_viewer_identity_separate_and_includes_all_statuses() -> None:
    with TemporaryDirectory(dir=ROOT / "ETL" / "output") as folder:
        folder_path = Path(folder)
        source = folder_path / "source.parquet"
        lookup = folder_path / "ua_lookup.parquet"
        output = folder_path / "events.csv"
        write_source_parquet(source)
        write_ua_lookup(lookup)
        subprocess.run(
            [
                sys.executable,
                str(TOOLS / "export_vod_query_events.py"),
                "--input",
                str(source),
                "--date",
                "2026-08-25",
                "--out",
                str(output),
                "--ua-lookup",
                str(lookup),
            ],
            check=True,
        )
        with output.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))

    assert len(rows) == 3
    assert {row["session_id"] for row in rows} == {"session-a", "session-b"}
    assert {row["device_id"] for row in rows} == {"device-a", "device-b"}
    assert {row["identity_source"] for row in rows} == {"same_cli_manifest"}
    assert {row["status_code"] for row in rows} == {"200", "404"}
    assert {row["decoded_device_type"] for row in rows} == {"smart_tv"}
    assert {row["decoded_model"] for row in rows} == {"Living Room TV"}
    failed = next(row for row in rows if row["status_code"] == "404")
    assert float(failed["request_watch_hours"]) == 6 / 3600
    assert float(failed["delivered_watch_hours"]) == 0
    assert failed["is_successful_segment"] == "0"


def test_compact_export_preserves_weighted_segments_and_identity() -> None:
    with TemporaryDirectory(dir=ROOT / "ETL" / "output") as folder:
        folder_path = Path(folder)
        source = folder_path / "source.parquet"
        lookup = folder_path / "ua_lookup.parquet"
        output = folder_path / "activity.csv"
        write_source_parquet(source)
        write_ua_lookup(lookup)
        subprocess.run(
            [
                sys.executable,
                str(TOOLS / "export_vod_query_events.py"),
                "--input",
                str(source),
                "--date",
                "2026-08-25",
                "--out",
                str(output),
                "--ua-lookup",
                str(lookup),
                "--compact",
            ],
            check=True,
        )
        with output.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))

    assert len(rows) == 2
    assert sum(int(row["segment_count"]) for row in rows) == 3
    assert sum(int(row["successful_segment_count"]) for row in rows) == 2
    assert {row["session_id"] for row in rows} == {"session-a", "session-b"}
    assert {row["decoded_os"] for row in rows} == {"ExampleOS 1.0"}
    viewer_a = next(row for row in rows if row["session_id"] == "session-a")
    assert viewer_a["status_code"] == "200, 404"
    assert float(viewer_a["request_watch_hours"]) == 12 / 3600
    assert float(viewer_a["delivered_watch_hours"]) == 6 / 3600


def test_dashboard_has_searchable_multiselect_and_safe_payload() -> None:
    row = {
        "log_date": "2026-08-25",
        "minute_ist": "2026-08-25 10:02:00",
        "request_ist": "2026-08-25 10:02:01.000",
        "content_title": "Title </script><script>alert(1)</script>",
        "category_name": "Drama",
        "content_code": "video-key-01",
        "request_watch_hours": str(6 / 3600),
        "cli_ip": "10.0.0.1",
        "device_id": "device-a",
        "session_id": "session-a",
    }
    html = render_html([row])
    assert 'id="titleSearch"' in html
    assert 'id="titleAll"' in html
    assert 'id="titleClear"' in html
    assert 'id="codeSearch"' in html
    assert 'id="codeAll"' in html
    assert 'id="codeClear"' in html
    assert 'id="codeFilter"' in html
    assert "All video keys" in html
    assert 'id="dateAvailabilitySummary"' in html
    assert 'id="availableDateList"' in html
    assert 'id="davisViewToggle"' in html
    assert 'id="davisPreset"' in html
    assert 'id="davisAudienceChart"' in html
    assert 'id="davisWatchChart"' in html
    assert 'id="davisTitles"' in html
    assert 'id="davisTitlesHead"' in html
    assert 'id="davisHeaderStatus"' in html
    assert 'id="davisExportExcel"' in html
    assert 'href="Davis_Cup_Performance.xlsx"' in html
    assert "Export Davis Cup Excel" in html
    assert "isDavisCupRow" in html
    assert "watchTimeParts" in html
    assert "const DAVIS_TO" not in html
    assert "24 Aug<br>" not in html
    assert "Watch / CLI IPs" in html
    assert "daily-metric" in html
    assert "First seen" in html
    assert "performanceData" in html
    assert "selectDavisCup" in html
    assert "firstSeenByCode" in html
    assert "grouped(rows,row=>row.minute_ist,0)" in html
    assert "codes.size===0||codes.has(row.content_code)" in html
    assert "activeRows.filter(row=>row.content_code)" in html
    assert ".davis-mode .filters" not in html
    assert "nearestAvailableDate" in html
    assert "container.replaceChildren()" in html
    assert "option.hidden" not in html
    assert "DecompressionStream" in html
    assert "Title </script>" not in html


def test_dashboard_uses_dominant_title_for_each_video_key() -> None:
    rows = [
        {
            "content_code": "r78srgxh",
            "content_title": "IND vs NED Match 3 Highlights | Doubles Match",
            "segment_count": "30",
        },
        {
            "content_code": "r78srgxh",
            "content_title": "BraveHearts Banner | Home Page | Top Banner",
            "segment_count": "2",
        },
        {
            "content_code": "r78srgxh",
            "content_title": "null",
            "segment_count": "100",
        },
    ]
    normalized = canonicalize_content_titles(rows)

    assert {row["content_title"] for row in normalized} == {
        "IND vs NED Match 3 Highlights | Doubles Match"
    }
    assert rows[1]["content_title"] == "BraveHearts Banner | Home Page | Top Banner"


def test_davis_workbook_has_watch_hours_and_cli_ip_sheets() -> None:
    rows = [
        {
            "log_date": "2026-08-24",
            "content_code": "match-five",
            "content_title": "Davis Cup_IND vs NED Match 5 Highlights",
            "category_name": "davis_cup_qualifiers_highlights",
            "request_watch_hours": "0.5",
            "segment_count": "300",
            "cli_ip": "10.0.0.1",
        },
        {
            "log_date": "2026-08-25",
            "content_code": "match-five",
            "content_title": "Davis Cup_IND vs NED Match 5 Highlights",
            "category_name": "davis_cup_qualifiers_highlights",
            "request_watch_hours": "1.0",
            "segment_count": "600",
            "cli_ip": "10.0.0.1",
        },
        {
            "log_date": "2026-08-25",
            "content_code": "match-five",
            "content_title": "Davis Cup_IND vs NED Match 5 Highlights",
            "category_name": "davis_cup_qualifiers_highlights",
            "request_watch_hours": "0.5",
            "segment_count": "300",
            "cli_ip": "10.0.0.2",
        },
    ]
    with TemporaryDirectory(dir=ROOT / "ETL" / "output") as folder:
        output = Path(folder) / "Davis_Cup_Performance.xlsx"
        write_davis_workbook(output, rows)
        workbook = load_workbook(output, data_only=True)

    assert workbook.sheetnames == ["Watch Minutes", "CLI IPs"]
    watch = workbook["Watch Minutes"]
    cli_ips = workbook["CLI IPs"]
    assert [cell.value for cell in watch[1]] == [
        "Match Title", "Date Posted", "Watch Minutes - Aug 24",
        "Watch Minutes - Aug 25", "Total Watch Minutes",
    ]
    assert [cell.value for cell in watch[2]] == [
        "Davis Cup | IND vs NED Match 5 Highlights", "2026-08-24",
        30, 90, 120,
    ]
    assert [cell.value for cell in cli_ips[2]] == [
        "Davis Cup | IND vs NED Match 5 Highlights", "2026-08-24", 1, 2, 2,
    ]
    assert watch.freeze_panes == "A2"
    assert cli_ips.freeze_panes == "A2"
    assert watch.auto_filter.ref == watch.dimensions
    assert all(cell.number_format == "0" for cell in watch[2][2:])


def test_watch_minutes_matches_dashboard_half_up_rounding() -> None:
    assert watch_minutes(0) == 0
    assert watch_minutes(2.5 / 60) == 3
    assert watch_minutes(2.49 / 60) == 2


def test_reimported_date_replaces_old_rows() -> None:
    old = [{"log_date": "2026-08-25", "request_ist": "old"}]
    new = [{"log_date": "2026-08-25", "request_ist": "new"}]
    assert merge_events(old, new) == new
