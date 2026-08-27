from __future__ import annotations

import csv
import sys
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
TOOLS = ROOT / "ETL" / "src" / "tools"
sys.path.insert(0, str(TOOLS))

from build_vod_key_details_report import build_report  # noqa: E402


def test_report_has_three_daily_metric_sheets_and_true_distinct_totals() -> None:
    with TemporaryDirectory(dir=ROOT / "ETL" / "output") as folder:
        root = Path(folder)
        source = root / "video-list.xlsx"
        events = root / "events.csv"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Video List"
        sheet.append(["Video Title", "VOD Key"])
        sheet.append(["Example A", "code-a"])
        sheet.append(["Example B", "code-b"])
        workbook.save(source)

        rows = [
            ["2026-08-24", "code-a", "0.5", "ip-1", "device-1"],
            ["2026-08-24", "code-a", "0.25", "ip-2", "device-2"],
            ["2026-08-25", "code-a", "0.5", "ip-1", "device-1"],
            ["2026-08-26", "other-code", "1", "ip-9", "device-9"],
        ]
        with events.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                ["log_date", "content_code", "request_watch_hours", "cli_ip", "device_id"]
            )
            writer.writerows(rows)

        output, dates, _, count = build_report(
            source,
            events_path=events,
            start_date=date(2026, 8, 24),
        )
        result = load_workbook(output, data_only=True)

    assert output.name == "video-list_details.xlsx"
    assert count == 2
    assert dates == [date(2026, 8, 24), date(2026, 8, 25), date(2026, 8, 26)]
    assert result.sheetnames == ["Watch Minutes", "Distinct CLI IPs", "Distinct Device IDs"]
    assert [cell.value for cell in result["Watch Minutes"][1]] == [
        "Video Title", "VOD Key", "24 Aug 2026", "25 Aug 2026", "26 Aug 2026", "Total"
    ]
    assert [cell.value for cell in result["Watch Minutes"][2]] == [
        "Example A", "code-a", 45, 30, 0, 75
    ]
    assert [cell.value for cell in result["Distinct CLI IPs"][2]] == [
        "Example A", "code-a", 2, 1, 0, 2
    ]
    assert [cell.value for cell in result["Distinct Device IDs"][2]] == [
        "Example A", "code-a", 2, 1, 0, 2
    ]
    assert [cell.value for cell in result["Watch Minutes"][3]] == [
        "Example B", "code-b", 0, 0, 0, 0
    ]
    assert result["Watch Minutes"].freeze_panes == "C2"
    assert result["Watch Minutes"].auto_filter.ref == result["Watch Minutes"].dimensions
